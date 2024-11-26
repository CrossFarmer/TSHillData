import logging
import sys 
import tkinter as tk
from tkinter import ttk
from tkinter import font
import tkinter.font as tkFont
from tkinter import filedialog
from tkinter import messagebox
import ttkthemes
import json 
import threading
import os 
import time
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.cell import _writer
import openpyxl
import re
from fpdf import FPDF


root_width = 1080
root_height = 720

def universal_locate_file_fn(file_name):
    """Finds the correct path to the file whether in script mode or PyInstaller .exe"""
    try:
        # When the script is bundled with PyInstaller
        base_path = sys._MEIPASS
    except AttributeError:
        # When running the script normally
        base_path = os.path.abspath(".")

    # Build the complete file path
    return os.path.join(base_path, file_name)

def load_json_list(file_name):
    file_path = universal_locate_file_fn(file_name)
    with open(file_path, 'r') as file:
        return json.load(file)

def load_json_dict(file_name):
    file_path = universal_locate_file_fn(file_name)
    with open(file_path, 'r') as file:
        json_lines = [line for line in file if not line.strip().startswith(("//", "#"))]
    json_data = ''.join(json_lines)
    data = json.loads(json_data)
    if not isinstance(data, dict):
        raise ValueError(f"Expected a JSON dictionary in {file_name}, but got {type(data)}")
    return data

def resource_path(relative_path):
    try:
        # When the application is bundled by PyInstaller, the sys._MEIPASS attribute is added
        base_path = sys._MEIPASS
    except Exception:
        # Fallback to the directory where the script is located
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

# Determine the directory of the executable
exe_dir = os.path.dirname(os.path.abspath(sys.executable))

# Change the date format to MM.DD.YYYY
log_filename = os.path.join(exe_dir, f"error_log_{datetime.now().strftime('%m.%d.%Y')}.txt")

# Set up logging
try:
    logging.basicConfig(
        level=logging.ERROR,
        format='%(asctime)s [%(levelname)s] %(message)s',
        handlers=[
            logging.FileHandler(log_filename, mode='a'),  # 'a' for append mode
            logging.StreamHandler()  # This will also print to console if you run the script directly
        ]
    )
    logging.error("Logging setup complete. Test log entry.")
except Exception as e:
    print(f"Failed to set up logging: {e}")
    with open(os.path.join(exe_dir, "log_setup_error.txt"), "w") as f:
        f.write(f"Failed to set up logging: {e}")

        
class WidgetManager:
    def __init__(self, root):
        self.root = root
        self.font_widgets = {}
        self.resize_scheduled = False
        self.last_width = root_width
        self.last_height = root_height
        self.tubing_or_drillpipe_menu = {'buttons': {}, 'labels': {}}
        self.saved_file_select_window = {'display_frames': {}, 'buttons': {}, 'labels': {}, 'listboxes': {}, 'scrollbars': {}, 'meta_labels': {}}
        self.branch_select_menu = {'buttons': {}, 'labels': {}}
        self.spreadsheet_exist_menu = {'buttons': {}, 'labels': {}}
        self.select_tx_ss_style_menu = {'buttons': {}, 'labels': {}}
        self.drillpipe_tabs_usemenu = {'buttons': {}, 'labels': {}, 'check_buttons': {}}
        self.select_active_tab_fresh_menu = {'buttons': {}, 'labels': {}}
        self.ask_datentry_file_window = {'buttons': {}, 'labels': {}}
        self.add_notes_window = {'buttons': {}, 'text': {}, 'labels': {}}
        self.main_metadata_widgets = {
            'labels': {}, 'entries': {}, 'buttons': {}, 'dropdown_menus': {},
            'review_widgets': {
                'review_buttons': {},
                'review_labels': {}
            }
        }
        self.table_select_page_widgets = {'buttons': {}, 'labels': {}}
        self.custom_column_selection_page_widgets = {'dropdown_menus': {}, 'labels': {}, 'buttons': {}}
        self.main_table_display_widgets = {
            'display_frames': {},
            'buttons': {},
            'back_buttons': {},
            'labels': {},
            'header_col_labels': {},
            'col_entry_labels': {}
        }

        self.continue_existing_report_screen = {
            'display_frames': {}, 
            'listboxes': {}, 
            'scrollbars': {},
            'buttons': {},
            'labels': {}
        }
        self.branch_var = tk.StringVar()
        self.branch_var.set("")
        self.branch_label = ttk.Label(self.root, textvariable=self.branch_var, font=('Arial', 24))

    #Creating Report Menu
        self.pdpir_boolean = tk.BooleanVar()
        self.hwdp_boolean = tk.BooleanVar()
        self.subs_boolean = tk.BooleanVar()
        self.actual_odid_boolean = tk.BooleanVar()

    #Selecting Saved Report File Window
        self.operator_msv = tk.StringVar()
        self.date_msv = tk.StringVar()
        self.contractor_msv = tk.StringVar()
        self.invoice_msv = tk.StringVar()
        self.location_msv = tk.StringVar()
        self.inspecttype_msv = tk.StringVar()
        self.connectsize_msv = tk.StringVar()
        self.conntype_msv = tk.StringVar()
        self.grade_info_tube = tk.StringVar()
        self.grade_info_tube.set("")


    # Main Data Table Window
        self.tab_data_header = tk.StringVar()
        self.tab_data_header.set("-")
        self.spreadsheet_type_prework = tk.StringVar()
        self.spreadsheet_type_prework.set("-")
        self.current_joint_number = tk.StringVar(value='1')
        self.magni_header = tk.StringVar()
        self.magni_content = tk.StringVar()
        self.micro_negative_two_stvar = tk.StringVar()
        self.micro_negative_one_stvar = tk.StringVar()
        self.micro_plus_one_stvar = tk.StringVar()
        self.micro_plus_two_stvar = tk.StringVar()
        self.first_entry_widget = None

        self.prev_first_row_label_list = []
        self.prev_sec_row_label_list = []
        self.next_first_row_label_list = []
        self.next_second_row_label_list = []

    def set_data_manager(self, data_manager):
        self.data_manager = data_manager

    def store_and_place(self, widget_dict, key, widget, font_changer=False, **params):
        widget_dict[key] = {'widget': widget, 'params': params, 'visible': True}
        widget.place(**params)
        if font_changer:
            initial_font = widget.cget("font")
            initial_wraplength = int(float(str(widget.cget("wraplength")))) if "wraplength" in widget.keys() else None
            widget_dict_id = id(widget_dict)
            text_content = widget.cget("text") if widget.cget("text") else None
            text_var = widget.cget("textvariable") if widget.cget("textvariable") else None
            if widget_dict_id not in self.font_widgets:
                self.font_widgets[widget_dict_id] = {}
            self.font_widgets[widget_dict_id][key] = {
                'widget': widget,
                'params': params,
                'visible': True,
                'initial_font': initial_font,
                'initial_wraplength': initial_wraplength,
                'text_content': text_content,
                'text_var': text_var
            }



    def store_and_grid(self, widget_dict, key, widget, font_changer=False, **params):
        widget_dict[key] = {'widget': widget, 'params': params, 'visible': True}
        widget.grid(**params)
        if font_changer:
            initial_font = widget.cget("font")
            initial_wraplength = int(float(str(widget.cget("wraplength")))) if "wraplength" in widget.keys() else None
            widget_dict_id = id(widget_dict)
            text_content = widget.cget("text") if widget.cget("text") else None
            text_var = widget.cget("textvariable") if widget.cget("textvariable") else None
            if widget_dict_id not in self.font_widgets:
                self.font_widgets[widget_dict_id] = {}
            self.font_widgets[widget_dict_id][key] = {
                'widget': widget,
                'params': params,
                'visible': True,
                'initial_font': initial_font,
                'initial_wraplength': initial_wraplength,
                'text_content': text_content,
                'text_var': text_var
            }



    def store_and_pack(self, widget_dict, key, widget, font_changer=False, **params):
        widget_dict[key] = {'widget': widget, 'params': params, 'visible': True}
        widget.pack(**params)
        if font_changer:
            initial_font = widget.cget("font")
            initial_wraplength = int(float(str(widget.cget("wraplength")))) if "wraplength" in widget.keys() else None
            widget_dict_id = id(widget_dict)
            text_content = widget.cget("text") if widget.cget("text") else None
            text_var = widget.cget("textvariable") if widget.cget("textvariable") else None
            if widget_dict_id not in self.font_widgets:
                self.font_widgets[widget_dict_id] = {}
            self.font_widgets[widget_dict_id][key] = {
                'widget': widget,
                'params': params,
                'visible': True,
                'initial_font': initial_font,
                'initial_wraplength': initial_wraplength,
                'text_content': text_content,
                'text_var': text_var
            }

    def hide_widget_grouping(self, widget_grouping):
        for key, widget_info in widget_grouping.items():
            widget = widget_info['widget']
            widget.place_forget()
            if key in self.font_widgets:
                self.font_widgets[key]['visible'] = False

    def place_back_widgets(self, widget_grouping):
        for key, widget_info in widget_grouping.items():
            widget = widget_info['widget']
            widget.place(**widget_info['params'])
            if key in self.font_widgets:
                self.font_widgets[key]['visible'] = True

    def get_label_widget(self, widmg, frame_key, label_key):
        label_info = widmg.main_table_display_widgets['display_frames'][frame_key][label_key]['widget']
        return label_info

  # Interactions
    def on_focus(self, event, root):
        def adjust_font(label, text_variable, base_font_family, base_font_size, base_font_weight):
            text = text_variable.get()
            label_height = int(label.winfo_height())
            label_width = int(label.winfo_width())
            wrap_length = int(label_width * 0.98)

            # Create a test label with similar configuration
            test_label = tk.Label(label.master, text=text, bd=1, relief="solid", font=(base_font_family, base_font_size, base_font_weight))
            test_label.config(wraplength=wrap_length)
            test_label.update_idletasks()

            # Adjust font size to fit text within the label
            while (test_label.winfo_reqwidth() > wrap_length or test_label.winfo_reqheight() > label_height) and base_font_size > 5:
                base_font_size -= 1
                test_label.config(font=(base_font_family, base_font_size, base_font_weight))
                test_label.update_idletasks()

            # Destroy the test label after adjustments
            test_label.destroy()

            # Apply the adjusted font size to the actual label
            label.config(font=(base_font_family, base_font_size, base_font_weight))

            # Update the widmg.font_widgets dictionary
            widget_dict_id = id(self.main_table_display_widgets['labels'])
            label_key = label._name
            if widget_dict_id not in self.font_widgets:
                self.font_widgets[widget_dict_id] = {}
            self.font_widgets[widget_dict_id][label_key] = {
                'widget': label,
                'params': {},
                'visible': True,
                'initial_font': (base_font_family, base_font_size, base_font_weight)
            }

        col_header_display_frame = self.main_table_display_widgets['display_frames']['col_header_display_frame']['widget']
        if event.widget.winfo_parent() == str(col_header_display_frame):
            active_tab = self.data_manager.json_data_dict['active_tab']
            column = event.widget.grid_info()['column']
            headers_list = self.data_manager.json_data_dict['report_data'][active_tab]['users_column_select']
            header = headers_list[column]
            self.magni_header.set(header)
            self.magni_content.set(event.widget.get())
            event.widget.config(highlightthickness=3, highlightbackground='blue', highlightcolor='green')  # Change border on focus

            # Adjust font for magni_header_label
            magni_header_label = self.main_table_display_widgets['labels']['magni_header_label']['widget']
            adjust_font(magni_header_label, self.magni_header, "Arial", 45, 'bold')

            # Adjust font for magni_content_label
            magni_content_label = self.main_table_display_widgets['labels']['magni_content_label']['widget']
            adjust_font(magni_content_label, self.magni_content, "Helvetica", 38, 'italic')

    def on_focus_out(self, event):
        event.widget.config(highlightthickness=1, highlightbackground='black', highlightcolor='black')  # Revert border on focus loss


    def on_key_release(self, event):
        col_header_display_frame = self.main_table_display_widgets['display_frames']['col_header_display_frame']['widget']
        if event.widget.winfo_parent() == str(col_header_display_frame):
            self.magni_content.set(event.widget.get())

    def update_magnifier(self, root):
        def adjust_font(label, text_variable, base_font_family, base_font_size, base_font_weight):
            text = text_variable.get()
            label_height = int(label.winfo_height())
            label_width = int(label.winfo_width())
            wrap_length = int(label_width * 0.98)

            # Create a test label with similar configuration
            test_label = tk.Label(label.master, text=text, bd=1, relief="solid", font=(base_font_family, base_font_size, base_font_weight))
            test_label.config(wraplength=wrap_length)
            test_label.update_idletasks()

            # Adjust font size to fit text within the label
            while (test_label.winfo_reqwidth() > wrap_length or test_label.winfo_reqheight() > label_height) and base_font_size > 5:
                base_font_size -= 1
                test_label.config(font=(base_font_family, base_font_size, base_font_weight))
                test_label.update_idletasks()

            # Destroy the test label after adjustments
            test_label.destroy()

            # Apply the adjusted font size to the actual label
            label.config(font=(base_font_family, base_font_size, base_font_weight))

            # Update the widmg.font_widgets dictionary
            widget_dict_id = id(self.main_table_display_widgets['labels'])
            label_key = label._name
            if widget_dict_id not in self.font_widgets:
                self.font_widgets[widget_dict_id] = {}
            self.font_widgets[widget_dict_id][label_key] = {
                'widget': label,
                'params': {},
                'visible': True,
                'initial_font': (base_font_family, base_font_size, base_font_weight)
            }

        # Existing logic to update magnifier content
        focus_widget = root.focus_get()
        if isinstance(focus_widget, tk.Entry):
            try:
                col_header_display_frame = self.main_table_display_widgets['display_frames']['col_header_display_frame']['widget']
                if focus_widget.winfo_parent() == str(col_header_display_frame):
                    column = focus_widget.grid_info()['column']
                    active_tab = self.data_manager.json_data_dict['active_tab']
                    headers_list = self.data_manager.json_data_dict['report_data'][active_tab]['users_column_select']
                    header = headers_list[column]
                    self.magni_header.set(header)
                    self.magni_content.set(focus_widget.get())

                    # Adjust font for magni_header_label
                    magni_header_label = self.main_table_display_widgets['labels']['magni_header_label']['widget']
                    adjust_font(magni_header_label, self.magni_header, "Arial", 45, 'bold')

                    # Adjust font for magni_content_label
                    magni_content_label = self.main_table_display_widgets['labels']['magni_content_label']['widget']
                    adjust_font(magni_content_label, self.magni_content, "Helvetica", 38, 'italic')
            except KeyError:
                print("Error: The focused widget is not configured properly.")



    def schedule_resize_fonts(self, event=None):
        current_width = event.width if event else self.root.winfo_width()
        current_height = event.height if event else self.root.winfo_height()

        if not self.resize_scheduled and (current_width != self.last_width or current_height != self.last_height):
            self.resize_scheduled = True
            self.last_width = current_width
            self.last_height = current_height
            self.root.after(100, self.resize_fonts)
            self.root.after(100, self.adjust_wraplength)

    def adjust_wraplength(self, event=None):
        current_width = event.width if event else self.root.winfo_width()
        width_ratio = current_width / root_width

        for widget_dict_id, widgets in self.font_widgets.items():
            for key, widget_info in widgets.items():
                if widget_info['visible']:
                    widget = widget_info['widget']
                    initial_wraplength = widget_info.get('initial_wraplength')
                    
                    if initial_wraplength is not None:
                        new_wraplength = int(initial_wraplength * width_ratio)
                        widget.config(wraplength=new_wraplength)


    def resize_fonts(self, event=None):
        self.resize_scheduled = False
        current_width = event.width if event else self.root.winfo_width()
        current_height = event.height if event else self.root.winfo_height()
        width_ratio = current_width / root_width
        height_ratio = current_height / root_height
        scaling_factor = min(width_ratio, height_ratio)

        def adjust_font_to_fit_widget(widget, initial_font):
            base_font_family, base_font_size, base_font_weight = initial_font
            label_height = int(widget.winfo_height())
            label_width = int(widget.winfo_width())
            wrap_length = int(label_width * 0.98)

            # Create a test label with similar configuration
            test_label = tk.Label(widget.master, text=widget.cget("text"), bd=1, relief="solid", font=(base_font_family, base_font_size, base_font_weight))
            test_label.config(wraplength=wrap_length)
            test_label.update_idletasks()

            # Adjust font size to fit text within the label
            while (test_label.winfo_reqwidth() > wrap_length or test_label.winfo_reqheight() > label_height) and base_font_size > 5:
                base_font_size -= 1
                test_label.config(font=(base_font_family, base_font_size, base_font_weight))
                test_label.update_idletasks()

            # Destroy the test label after adjustments
            test_label.destroy()

            # Apply the adjusted font size to the actual widget
            widget.config(font=(base_font_family, base_font_size, base_font_weight))

        for widget_dict_id, widgets in self.font_widgets.items():
            for key, widget_info in widgets.items():
                if widget_info['visible']:
                    widget = widget_info['widget']
                    initial_font = widget_info['initial_font']
                    
                    if isinstance(initial_font, tuple):
                        family = initial_font[0]
                        size = initial_font[1]
                        styles = initial_font[2:]  # Capture additional styles
                        
                        # Initialize default styles
                        weight = 'normal'
                        slant = 'roman'
                        underline = 0
                        overstrike = 0
                        
                        # Parse styles
                        for style in styles:
                            if style == 'bold':
                                weight = 'bold'
                            elif style == 'italic':
                                slant = 'italic'
                            elif style == 'underline':
                                underline = 1
                            elif style == 'overstrike':
                                overstrike = 1
                        
                        adjust_font_to_fit_widget(widget, (family, size, weight))

                    elif isinstance(initial_font, str):
                        parts = initial_font.split()
                        family = parts[0]
                        size = int(parts[1]) if parts[1].isdigit() else 10
                        style_parts = parts[2:] if len(parts) > 2 else []
                        
                        weight = 'normal'
                        slant = 'roman'
                        underline = 0
                        overstrike = 0
                        
                        for style in style_parts:
                            if style == 'bold':
                                weight = 'bold'
                            elif style == 'italic':
                                slant = 'italic'
                            elif style == 'underline':
                                underline = 1
                            elif style == 'overstrike':
                                overstrike = 1
                        
                    adjust_font_to_fit_widget(widget, (family, size, weight))


class DataManager:
    def __init__(self):
        self.json_data_dict = {
            "branch": "", 
            "report_type": "", 
            "active_tab": "", 
            "report_user_metadata": {}, 
            "report_data": {}
        }
        self.is_saving = False
        self.lock = threading.Lock()
        self.save_thread = None
        self.xel_file_path = None
        self.new_excel_fp = None
        self.excel_files_tct = None
        self.json_tct_filepath = None
        self.filename = None
        self.notes_prefill = {"Tubing/Casing Report": "Missing Caps: () BOX ; () PIN"}
        self.keyword_tally_dict = {"Keyword Tallies": {}, "Joint Tallies": {}}
        self.nd_column_types = load_json_dict(resource_path('external_files/nd_column_types.json'))
        self.tx_column_types = load_json_dict(resource_path('external_files/tx_column_types.json'))
        self.tubing_report = False
        self.dpinspection_report = False
        self.tubing_inspection_type_list = load_json_list(resource_path('external_files/tubing_inspection_type_list.json'))
        self.dp_inspection_type_list = load_json_list(resource_path('external_files/dp_inspection_type_list.json'))
        self.dp_inspection_type_list_short = ["Cat 2", "Cat 2 w/Blacklight", "Cat 3", "Cat 3-5", "Cat 4", "Cat 4 w/Blacklight", "Cat 5", "API RP 7G"]
        self.tubing_conn_size_list = ["1.9\"", "2 3/8\"", "2 7/8\"", "4.5\"", "3.5\""]
        self.dp_conn_size_list = ["4.0", "4.5", "5.0", "5.25", "5.5", "5.5 - 24.7lb", "2.875"]
        self.tubing_conn_type_list = ["PH6", "EUE", "TTWS", "FATBOY-PH6", "CS8", "AOH", "BTC"]
        self.dp_conn_type_list = load_json_list(resource_path('external_files/dp_conn_type_list.json'))
        self.dp_conn_size_nom_rel_dict = load_json_dict(resource_path('external_files/dp_conn_size_nom_rel_dict.json'))
        self.tube_conn_size_nom_rel_dict = load_json_dict(resource_path('external_files/tube_conn_size_nom_rel_dict.json'))
        self.invalid_combinations_cs_dt = load_json_dict(resource_path('external_files/invalid_combinations_cs_dt.json'))
        self.dp_conn_type_vals_dict = load_json_dict(resource_path('external_files/dp_conn_type_vals_dict.json'))
        self.color_code_dict = load_json_list(resource_path('external_files/color_code_dict.json'))
        self.nd_operator_vals = load_json_list(resource_path('external_files/nd_operator_vals.json'))
        self.editing_spec_tab = False

    def set_widget_manager(self, widget_manager):
        self.widget_manager = widget_manager


# Data Storage File METHODS
    def save_dict_to_file_start(self):
        # Ensure the directory is created before saving
        directory = os.path.join(os.getcwd(), 'data_entry_files')
        if not os.path.exists(directory):
            os.makedirs(directory)
        
        self.filename = os.path.join(directory, os.path.basename(self.filename))

        with open(self.filename, 'w') as f:
            json.dump(self.json_data_dict, f)

    def save_dict_to_file(self):
        # Skip saving if another save is in progress
        if self.is_saving:
            print("A save operation is already in progress. Skipping this save request.")
            return

        # Set the flag and begin the background save
        self.is_saving = True
        self.save_thread = threading.Thread(target=self._save_operation)
        self.save_thread.start()

    def _save_operation(self):
        directory = os.path.join(os.getcwd(), 'data_entry_files')
        if not os.path.exists(directory):
            os.makedirs(directory)
        self.filename = os.path.join(directory, os.path.basename(self.filename))

        with self.lock:  # Ensure thread-safe execution
            try:
                tmp_filename = f"{self.filename}.tmp"
                bak_filename = f"{self.filename}.bak"

                # Step 1: Write to a temporary file
                with open(tmp_filename, 'w') as tmp_file:
                    json.dump(self.json_data_dict, tmp_file)
                    tmp_file.flush()  # Flush the data
                    os.fsync(tmp_file.fileno())  # Ensure it's written to disk

                # Step 2: Make a backup of the existing file
                if os.path.exists(self.filename):
                    os.replace(self.filename, bak_filename)

                # Step 3: Safely replace the original file with the temporary file
                os.replace(tmp_filename, self.filename)

            except Exception as e:
                print(f"Error during the save operation: {e}")

            finally:
                # Reset the flag to allow future saves
                self.is_saving = False
                
    def load_file_to_dict(self, filename):
        try:
            with open(filename, 'r') as file:
                self.json_data_dict = json.load(file)
                self.filename = filename
        except Exception as e:
            print(f"Error reading file {filename}: {e}")

    def initialize_json(self, widmg):
        #Create a JSON File To Store The Input Data For Report, Saving it as the Date and Time For Filename
        md = self.json_data_dict['report_user_metadata']
        active_tab = self.json_data_dict['active_tab']

        current_date = datetime.now().strftime('%m.%d.%Y')
        current_time = datetime.now().strftime('%m.%d.%Y_%I.%M.%p')
        md['create_date'] = current_date
        md['create_time'] = current_time


        if self.json_data_dict['report_type'] == "Drill Pipe Inspection Report":
            base_filename = f"{md['date_entry'].replace('/', '.')}_INV{md['invoice_entry']}_{md['connection_size_selection']} Inch DP Inspection Report_{md['operator_entry']}_{md['contractor_entry']}"
        elif self.json_data_dict['report_type'] == "Tubing/Casing Report":
            if md['connection_size_selection'] == "2 7/8\"":
                file_con_sizesel = '2.875'
            elif md['connection_size_selection'] == "2 3/8\"":
                file_con_sizesel = '2.375'
            else:
                file_con_sizesel = md['connection_size_selection'].strip('"')

            if 'grade_info' in md and md['grade_info'] != None:
                grade_info = f"{md['grade_info']}_"
            else:
                grade_info = f"{md['grade_info']}_" if 'grade_info' in md else ""

            base_filename = f"{md['date_entry'].replace('/', '.')}_INV{md['invoice_entry']}_{md['operator_entry']}_{file_con_sizesel}_{grade_info}Tubing_{md['contractor_entry']}_{md['connection_type_selection']}"

        self.filename = f"{base_filename}.json"
        self.save_dict_to_file_start()



    def save_tab_status_table_to_json(self, file_name):
        with open(file_name, 'w') as json_file:
            json.dump(self.excel_files_tct, json_file, indent=4)  

    def load_tab_status_table_to_dict(self, excel_filename):
        raw_filename = excel_filename.replace(".xlsx", "")
        tct_filename = f"{raw_filename}_tct.json"
        try:
            with open(tct_filename, 'r') as file:
                self.excel_files_tct = json.load(file)
                self.json_tct_filepath = tct_filename
        except Exception as e:
            print(f"Error reading file {tct_filename}: {e}")

    def update_user_report_metadata(self):
        widges = self.widget_manager.main_metadata_widgets

        combined_widgets = list(widges['entries'].items()) + list(widges['dropdown_menus'].items())

        for key, widget_info in combined_widgets:
            user_input = widget_info['widget'].get()
            self.json_data_dict['report_user_metadata'][key] = user_input


#-----------------------------------------------------------------------

def initialize_main_window():
    root = tk.Tk()
    style = ttkthemes.ThemedStyle(root)
    style.set_theme("arc") #arc
    style.configure('Large.TButton', font=('Arial', 18))
    style.configure('Combobox', font=('Arial', 16))
    root.title("TS-Hill Data Entry Device")
    root.geometry(f"{root_width}x{root_height}")
    return root

def branch_select_screen(widmg, datmg, root):
    page = widmg.branch_select_menu

    branch_select_label = ttk.Label(root, text="Select Your Branch", font=('Arial', 32))
    widmg.store_and_place(page['labels'], "branch_select_label", branch_select_label, relx=0.5, rely=0.15, relheight=0.15, anchor='c')
   
    select_nd_button = ttk.Button(root, text="ND", command=lambda: select_nd_but_to_tdpselscreen('ND', widmg, datmg, root), style='Large.TButton')
    widmg.store_and_place(page['buttons'], "select_nd_button", select_nd_button, relx=0.5, rely=0.30, relwidth=0.50, relheight=0.2, anchor='center')

    select_tx_button = ttk.Button(root, text="TX", command=lambda: select_tx_but_to_tdpselscreen('TX', widmg, datmg, root), style='Large.TButton')
    widmg.store_and_place(page['buttons'], "select_tx_button", select_tx_button, relx=0.5, rely=0.55, relwidth=0.50, relheight=0.2, anchor='center')


def select_nd_but_to_tdpselscreen(branch, widmg, datmg, root):
    datmg.json_data_dict['branch'] = branch
    widmg.branch_var.set(branch)
    page = widmg.branch_select_menu
    widmg.hide_widget_grouping(page['labels'])
    widmg.hide_widget_grouping(page['buttons'])
    create_tubing_drillpipe_selection_screen(widmg, datmg, root)

def select_tx_but_to_tdpselscreen(branch, widmg, datmg, root):
    datmg.json_data_dict['branch'] = branch
    widmg.branch_var.set(branch)
    datmg.json_data_dict['report_type'] = "Drill Pipe Inspection Report"
    page = widmg.branch_select_menu
    widmg.hide_widget_grouping(page['labels'])
    widmg.hide_widget_grouping(page['buttons'])
    does_spreadsheet_exist_screen(widmg, datmg, root)


def create_tubing_drillpipe_selection_screen(widmg, datmg, root):
    widmg.branch_label.place(relx=0.92, rely=0.02, relheight=0.07, relwidth=0.07)
    page = widmg.tubing_or_drillpipe_menu

    menu_selection_header_label = ttk.Label(root, text="SELECT YOUR REPORT TYPE", font=('Arial', 32))
    widmg.store_and_place(page['labels'], "menu_selection_header_label", menu_selection_header_label, relx=0.5, rely=0.15, relheight=0.15, anchor='c')

    tube_casing_button = ttk.Button(root, text="Tubing/Casing Report", command=lambda: tubing_button_to_ssescreen(widmg, datmg, root), style='Large.TButton')
    widmg.store_and_place(page['buttons'], "tube_casing_button", tube_casing_button, relx=0.5, rely=0.27, relwidth=0.75, relheight=0.20, anchor='n')
   
    drill_pipe_button = ttk.Button(root, text="Drill Pipe Inspection Report", command=lambda: drill_pipe_button_to_ssescreen(widmg, datmg, root), style='Large.TButton')
    widmg.store_and_place(page['buttons'], "drill_pipe_button", drill_pipe_button, relx=0.5, rely=0.52, relwidth=0.75, relheight=0.20, anchor='n')

    back_to_branch_btn = ttk.Button(root, text="BACK", command=lambda: back_to_branch_action(widmg, datmg, root), style='Large.TButton')
    widmg.store_and_place(page['buttons'], "back_to_branch_btn", back_to_branch_btn, relx=0.5, rely=0.85, relwidth=0.40, relheight=0.15, anchor='center')

def back_to_branch_action(widmg, datmg, root):
    page = widmg.tubing_or_drillpipe_menu
    widmg.hide_widget_grouping(page['labels'])
    widmg.hide_widget_grouping(page['buttons'])
    branch_select_screen(widmg, datmg, root)

def tubing_button_to_ssescreen(widmg, datmg, root):
    datmg.json_data_dict['report_type'] = "Tubing/Casing Report"
    datmg.json_data_dict['active_tab'] = 'Tubing Insp Report'
    datmg.json_data_dict['report_data'] = {
        "Tubing Insp Report": {
            'joint_count': 0,
            'users_column_select': [],
            "is_complete": False,
            "inspection_type_data": {
                "inspection_type_selection": "", 
                "inspection_type_addodid": "", 
                "inspection_type_additional": ""
            },
            "joint_data": {}
        }
    }
    widmg.spreadsheet_type_prework.set("Tubing Spreadsheet")
    page = widmg.tubing_or_drillpipe_menu

    widmg.hide_widget_grouping(page['labels'])
    widmg.hide_widget_grouping(page['buttons'])
    
    does_spreadsheet_exist_screen(widmg, datmg, root)



def drill_pipe_button_to_ssescreen(widmg, datmg, root):
    datmg.json_data_dict['report_type'] = "Drill Pipe Inspection Report"

    widmg.spreadsheet_type_prework.set("Drill Pipe Spreadsheet")
    page = widmg.tubing_or_drillpipe_menu

    widmg.hide_widget_grouping(page['labels'])
    widmg.hide_widget_grouping(page['buttons'])
    
    does_spreadsheet_exist_screen(widmg, datmg, root)

#### START HERE ###
## IMPLEMENTING TEXAS' "DOES SPREADSHEET EXIST" SCREEN


def continue_existing_report_btn(widmg, datmg, root):
    page = widmg.spreadsheet_exist_menu
    widmg.hide_widget_grouping(page['labels'])
    widmg.hide_widget_grouping(page['buttons'])

    display_incomplete_reports(widmg, datmg, root)



### Function that allows user to continue working on an existing report. This function currently works correctly assuming that
### the datmg.json_data_dict/self.filename is in the proper format. Currently, having issues due to this.



def display_incomplete_reports(widmg, datmg, root):
    import os
    import tkinter as tk
    from tkinter import ttk
    
    # Define the page variable for easy reference
    page = widmg.continue_existing_report_screen
    
    # Path to the folder containing incomplete report files
    folder_path = os.path.join(os.path.dirname(__file__), 'data_entry_files')
    
    # Label to instruct user
    select_existing_report = ttk.Label(root, text="SELECT EXISTING REPORT FILE", font=('Arial', 32))
    widmg.store_and_place(page['labels'], "select_existing_report", select_existing_report, relx=0.5, rely=0.05, relheight=0.10, anchor='c')
    
    # Frame for holding the filename list
    file_display_frame = ttk.Frame(root)
    widmg.store_and_place(page['display_frames'], "file_display_frame", file_display_frame, relx=0.5, rely=0.225, relwidth=0.8, relheight=0.1, anchor='center')
    
    # Listbox for displaying Excel filenames
    file_listbox = tk.Listbox(file_display_frame, selectmode=tk.SINGLE)
    widmg.store_and_place(page['listboxes'], "file_listbox", file_listbox, relx=0, rely=0, relwidth=1.0, relheight=1.0)
    
    # Scrollbar for the listbox
    file_scrollbar = tk.Scrollbar(file_display_frame, orient=tk.VERTICAL, command=file_listbox.yview)
    widmg.store_and_place(page['scrollbars'], "file_scrollbar", file_scrollbar, relx=0.98, rely=0, relheight=1.0)
    file_listbox.config(yscrollcommand=file_scrollbar.set)
    
    # Populate listbox with .json filenames from the specified folder
    for filename in os.listdir(folder_path):
        if filename.endswith('.json'):
            file_listbox.insert(tk.END, filename)
    
    # Bind the selection event to load data and determine frames dynamically
    file_listbox.bind("<<ListboxSelect>>", lambda event: handle_file_selection(root, datmg, widmg, event))
    
    # Create the 'BACK' button
    back_button = ttk.Button(root, text="BACK", command=lambda: back_from_displayincompletereports(widmg, datmg, root), style='Large.TButton')
    widmg.store_and_place(page['buttons'], "back_button", back_button, relx=0.5, rely=0.95, relwidth=0.33, relheight=0.075, anchor='center')
    
    def handle_file_selection(root, datmg, widmg, event):
        # Retrieve the selected file from the listbox
        selected_index = file_listbox.curselection()
        if not selected_index:
            return
        selected_file = file_listbox.get(selected_index[0])
        
        sel_file_full_path = os.path.join(folder_path, selected_file)
        
        # Load data from the file
        datmg.load_file_to_dict(sel_file_full_path)
        datmg.filename = sel_file_full_path
        
        # Retrieve tabs information
        included_tabs = get_report_tabs_info(datmg)
        create_tab_display_frames(root, datmg, widmg, included_tabs)

    def get_report_tabs_info(datmg):
        included_tabs = []
        for tab in datmg.json_data_dict['report_data']:
            included_tabs.append(tab)
        return included_tabs


    def create_tab_display_frames(root, datmg, widmg, included_tabs):
        # Determine frame configurations based on included_tabs length
        frame_count = min(len(included_tabs), 3)  # Limit to a maximum of 3 frames
        
        # Define the page variable for easy reference
        page = widmg.continue_existing_report_screen

        for frame_key in list(page['display_frames'].keys()):
            frame = page['display_frames'].pop(frame_key)
            if isinstance(frame, ttk.Frame) or isinstance(frame, tk.Frame):
                frame.destroy()
        
        # Calculate frame width based on the number of frames
        frame_width = 0.8 / frame_count  # Equal width division
        
        # Display the report type in a separate frame
        report_type_display_frame = ttk.Frame(root)
        report_type_label = ttk.Label(report_type_display_frame, text=datmg.json_data_dict['report_type'], font=('Arial', 20))
        report_type_label.pack(anchor='center')
        widmg.store_and_place(page['display_frames'], "report_type_display_frame", report_type_display_frame, relx=0.5, rely=0.35, relwidth=0.8, relheight=0.05, anchor='center')
        
        # Get user metadata to display in each tab frame
        user_metadata = datmg.json_data_dict['report_user_metadata']
        metadata_list = [
            user_metadata['operator_entry'], user_metadata['contractor_entry'], user_metadata['location_entry'], 
            user_metadata['date_entry'], user_metadata['invoice_entry'], user_metadata['connection_size_selection'], 
            user_metadata['connection_type_selection'], user_metadata['inspected_by_entry'], user_metadata.get('grade_info')
        ]
        metadata_text = "\n".join(str(item) for item in metadata_list if item)  # Join metadata into one block of text

        # Dynamically create frames based on the number of included tabs
        for i, tab_key in enumerate(included_tabs):
            frame_name = f"tab_display_frame_{i + 1}"
            tab_label_name = f"tab_label_name_{i + 1}"
            tdf_completion_lbl_name = f"tdf_completion_lbl_name_{i + 1}"
            tdf_metadata_lbl_name = f"tdf_metadata_lbl_name_{i + 1}"
            tdf_inspection_lbl_name = f"tdf_inspection_lbl_name_{i + 1}"
            tdf_button_name = f"tdf_button_name_{i + 1}"

            frame = ttk.Frame(root)
            relx_position = 0.1 + i * frame_width + frame_width / 2
            widmg.store_and_place(page['display_frames'], frame_name, frame, relx=relx_position, rely=0.65, relwidth=frame_width, relheight=0.45, anchor='center')
            
            # Add the tab name at the top
            tab_label = ttk.Label(frame, text=tab_key, font=('Arial', 14, 'bold'))
            widmg.store_and_place(page['labels'], tab_label_name, tab_label, relx=0.5, rely=0.04, relwidth=0.95, relheight=0.10, anchor='center')
            
            # Check if tab is complete and set completion status
            tab_data = datmg.json_data_dict['report_data'][tab_key]
            is_complete = tab_data['is_complete']
            completion_text = "Complete" if is_complete else "Incomplete"
            completion_color = "green" if is_complete else "red"

            completion_label = ttk.Label(frame, text=completion_text, font=('Arial', 14, 'bold'), foreground=completion_color)
            widmg.store_and_place(page['labels'], tdf_completion_lbl_name, completion_label, relx=0.5, rely=0.12, relwidth=0.95, relheight=0.10, anchor='center')

            # Display user metadata
            metadata_label = ttk.Label(frame, text=metadata_text, font=('Arial', 10), wraplength=frame_width*500)  # Wrap text to fit frame
            widmg.store_and_place(page['labels'], tdf_metadata_lbl_name, metadata_label, relx=0.5, rely=0.40, relwidth=0.95, relheight=0.30, anchor='center')
            
            # Display unique inspection data for this tab
            inspection_data = tab_data["inspection_type_data"]
            inspection_parts = []
            if inspection_data.get("inspection_type_selection"):
                inspection_parts.append(inspection_data["inspection_type_selection"])
            if inspection_data.get("inspection_type_addodid"):
                inspection_parts.append(inspection_data["inspection_type_addodid"])
            if inspection_data.get("inspection_type_additional"):
                inspection_parts.append(inspection_data["inspection_type_additional"])
            # Filter out any None values in inspection_parts
            inspection_text = " ".join(str(part) for part in inspection_parts if part is not None)

            
            inspection_label = ttk.Label(frame, text=inspection_text, font=('Arial', 12, 'italic'), wraplength=frame_width*300)
            widmg.store_and_place(page['labels'], tdf_inspection_lbl_name, inspection_label, relx=0.5, rely=0.71, relwidth=0.95, relheight=0.15, anchor='center')
            
            # Add button based on completion status
            button_text = "EDIT" if is_complete else "START"
            action_button = ttk.Button(frame, text=button_text, command=lambda tab=tab_key: bring_to_main_report_scrn(widmg, datmg, root, tab))
            widmg.store_and_place(page['buttons'], tdf_button_name, action_button, relx=0.5, rely=0.91, relwidth=0.87, relheight=0.09, anchor='center')



def bring_to_main_report_scrn(widmg, datmg, root, tab_key):
    page = widmg.continue_existing_report_screen
    widmg.hide_widget_grouping(page['display_frames'])
    widmg.hide_widget_grouping(page['listboxes'])
    widmg.hide_widget_grouping(page['scrollbars'])
    widmg.hide_widget_grouping(page['buttons'])
    widmg.hide_widget_grouping(page['labels'])

    if tab_key == "Prop Drill Pipe Inp Report":
        widmg.tab_data_header.set("PDPIR")
    elif tab_key == "Prop HWDP Inp Report":
        widmg.tab_data_header.set("HWDP")
    elif tab_key == "Prop Subs Inp Report":
        widmg.tab_data_header.set("SUBS")
    elif tab_key == "Tubing Insp Report":
        widmg.tab_data_header.set(datmg.json_data_dict['report_user_metadata']['connection_type_selection'])

    datmg.json_data_dict['active_tab'] = tab_key
    active_tab = datmg.json_data_dict['active_tab']

    datmg.json_data_dict['report_data'][tab_key]['is_complete'] = False
    datmg.save_dict_to_file()

    if datmg.json_data_dict['report_data'][active_tab]['users_column_select'] == []:
        show_table_selection_screen(widmg, datmg, root)
    else:
        display_main_report_screen(widmg, datmg, root)
    
    

def back_from_displayincompletereports(widmg, datmg, root):
    # Clear the widgets from the 'continue_existing_report_screen'
    datmg.filename = None
    branch = datmg.json_data_dict['branch']
    report_type = datmg.json_data_dict['report_type']
    datmg.json_data_dict = {
        "branch": branch, 
        "report_type": report_type, 
        "active_tab": "", 
        "report_user_metadata": {}, 
        "report_data": {}
    }

    page = widmg.continue_existing_report_screen
    widmg.hide_widget_grouping(page['display_frames'])
    widmg.hide_widget_grouping(page['listboxes'])
    widmg.hide_widget_grouping(page['scrollbars'])
    widmg.hide_widget_grouping(page['buttons'])
    widmg.hide_widget_grouping(page['labels'])
    
    # Navigate to the 'does_spreadsheet_exist_screen'
    does_spreadsheet_exist_screen(widmg, datmg, root)




def does_spreadsheet_exist_screen(widmg, datmg, root):
    widmg.branch_label.place(relx=0.92, rely=0.02, relheight=0.07, relwidth=0.07)
    page = widmg.spreadsheet_exist_menu

    spread_exist_label = ttk.Label(root, text="START AN INSPECTION", font=('Arial', 32))
    widmg.store_and_place(page['labels'], "spread_exist_label", spread_exist_label, relx=0.5, rely=0.15, relheight=0.15, anchor='c')

    create_new_ss_button = ttk.Button(root, text="Create New Report", command=lambda: create_new_ss_bridge(widmg, datmg, root), style='Large.TButton')
    widmg.store_and_place(page['buttons'], "create_new_ss_button", create_new_ss_button, relx=0.5, rely=0.27, relwidth=0.75, relheight=0.20, anchor='n')

    use_existing_ss_button = ttk.Button(root, text="Continue Existing Report", command=lambda: continue_existing_report_btn(widmg, datmg, root), style='Large.TButton')
    widmg.store_and_place(page['buttons'], "use_existing_ss_button", use_existing_ss_button, relx=0.5, rely=0.52, relwidth=0.75, relheight=0.20, anchor='n')

    backto_ss_exist_screen = ttk.Button(root, text="BACK", command=lambda: back_to_dptubesel_action(widmg, datmg, root), style='Large.TButton')
    widmg.store_and_place(page['buttons'], "backto_ss_exist_screen", backto_ss_exist_screen, relx=0.5, rely=0.85, relwidth=0.5, relheight=0.15, anchor='center')  


def ss_exist_back_bridge_tx(widmg, datmg, root):
    page = widmg.select_tx_ss_style_menu
    widmg.hide_widget_grouping(page['labels'])
    widmg.hide_widget_grouping(page['buttons'])

    does_spreadsheet_exist_screen(widmg, datmg, root)

def back_to_dptubesel_action(widmg, datmg, root):
    branch = datmg.json_data_dict['branch']
    page = widmg.spreadsheet_exist_menu
    widmg.hide_widget_grouping(page['labels'])
    widmg.hide_widget_grouping(page['buttons'])
    if branch == 'ND':
        create_tubing_drillpipe_selection_screen(widmg, datmg, root)
    elif branch == 'TX':
        branch_select_screen(widmg, datmg, root)
    
def create_new_ss_bridge(widmg, datmg, root):
    page = widmg.spreadsheet_exist_menu
    widmg.hide_widget_grouping(page['labels'])
    widmg.hide_widget_grouping(page['buttons'])

    if datmg.json_data_dict['branch'] == 'ND':
        if datmg.json_data_dict['report_type'] == 'Drill Pipe Inspection Report':
            create_tab_select_screen(widmg, datmg, root)
        elif datmg.json_data_dict['report_type'] == 'Tubing/Casing Report':
            widmg.tab_data_header.set('TUBING/CASING')
            create_report_metadata_input_widgets(widmg, datmg, root, editing=False)
    elif datmg.json_data_dict['branch'] == 'TX':
        select_tx_ss_style_screen(widmg, datmg, root)
        


def select_tx_ss_style_screen(widmg, datmg, root):
    page = widmg.select_tx_ss_style_menu

    spread_exist_label = ttk.Label(root, text="SELECT SPREADSHEET STYLE", font=('Arial', 32))
    widmg.store_and_place(page['labels'], "spread_exist_label", spread_exist_label, relx=0.5, rely=0.15, relheight=0.15, anchor='c')

    class_2_dbr_btn = ttk.Button(root, text="Class 2 DBR", command=lambda: tx_ss_style_to_tab_select(widmg, datmg, root, 'Class 2 DBR'), style='Large.TButton')
    widmg.store_and_place(page['buttons'], "class_2_dbr_btn", class_2_dbr_btn, relx=0.5, rely=0.25, relwidth=0.75, relheight=0.15, anchor='n')

    class_2_not_dbr_btn = ttk.Button(root, text="Class 2 NOT DBR", command=lambda: tx_ss_style_to_tab_select(widmg, datmg, root, 'Class 2 NOT DBR'), style='Large.TButton')
    widmg.store_and_place(page['buttons'], "class_2_not_dbr_btn", class_2_not_dbr_btn, relx=0.5, rely=0.45, relwidth=0.75, relheight=0.15, anchor='n')

    full_dimensional_btn = ttk.Button(root, text="Full Dimensional", command=lambda: tx_ss_style_to_tab_select(widmg, datmg, root, 'Full Dimensional'), style='Large.TButton')
    widmg.store_and_place(page['buttons'], "full_dimensional_btn", full_dimensional_btn, relx=0.5, rely=0.65, relwidth=0.75, relheight=0.15, anchor='n')

    backto_dptubesel_btn = ttk.Button(root, text="BACK", command=lambda: ss_exist_back_bridge_tx(widmg, datmg, root), style='Large.TButton')
    widmg.store_and_place(page['buttons'], "backto_dptubesel_btn", backto_dptubesel_btn, relx=0.5, rely=0.90, relwidth=0.5, relheight=0.10, anchor='center')  


def tx_ss_style_to_tab_select(widmg, datmg, root, ss_style):
    page = widmg.select_tx_ss_style_menu
    widmg.hide_widget_grouping(page['labels'])
    widmg.hide_widget_grouping(page['buttons'])

    datmg.json_data_dict['report_style'] = ss_style
    type_storage = datmg.json_data_dict['report_style']

    if ss_style == 'Class 2 DBR':
        type_storage = 'Class 2 DBR'
    elif ss_style == 'Class 2 NOT DBR':
        type_storage = 'Class 2 NOT DBR'
    elif ss_style == 'Full Dimensional':
        type_storage = 'Full Dimensional'

    create_tab_select_screen(widmg, datmg, root)


def use_existing_ss_bridge(widmg, datmg, root):
    report_type = datmg.json_data_dict['report_type']
    branch = datmg.json_data_dict['branch']
    datmg.xel_file_path = None
    datmg.new_excel_fp = None
    page = widmg.spreadsheet_exist_menu

    widmg.hide_widget_grouping(page['labels'])
    widmg.hide_widget_grouping(page['buttons'])

    avail_tabs = return_sheets_active_tabs_first(widmg, datmg, root)
    workbook = load_workbook(filename=datmg.xel_file_path, read_only=True)

    if branch == 'ND':
        if report_type == 'Drill Pipe Inspection Report':
            create_tab_select_existing_screen(avail_tabs, widmg, datmg, root)
        elif report_type == 'Tubing/Casing Report':
            ask_for_data_entry_file_screen(widmg, datmg, root)
    elif branch == 'TX':
        if 'DATA SHEET' in workbook.sheetnames:
            data_sheet = workbook['DATA SHEET']
            value_j2 = data_sheet['J2'].value  # Store the value in J2
            datmg.json_data_dict['report_style'] = value_j2

        create_tab_select_existing_screen(avail_tabs, widmg, datmg, root)



def return_sheets_active_tabs_first(widmg, datmg, root):
    # Open the file dialog to select an Excel file
    file_path = filedialog.askopenfilename(
        title="Select the spreadsheet you want to use",
        filetypes=[("Excel files", "*.xlsx")]
    )
    datmg.xel_file_path = file_path
    if not file_path:  # If no file is selected, return an empty list
        datmg.xel_file_path = None
        does_spreadsheet_exist_screen(widmg, datmg, root)

    # Load the workbook and get all sheets
    workbook = load_workbook(filename=file_path, read_only=True)



    # List to hold the names of visible tabs
    visible_tabs = [sheet.title for sheet in workbook.worksheets if not sheet.sheet_state == 'hidden']

    return visible_tabs

def return_sheets_active_tabs_after(widmg, datmg, root):
    # Load the workbook and get all sheets
    workbook = load_workbook(filename=datmg.xel_file_path, read_only=True)

    # List to hold the names of visible tabs
    visible_tabs = [sheet.title for sheet in workbook.worksheets if not sheet.sheet_state == 'hidden']

    return visible_tabs


def create_tab_select_screen(widmg, datmg, root):
    page = widmg.drillpipe_tabs_usemenu
    style = ttk.Style()
    style.configure('LargeFont.TButton', font=('Helvetica', 20))


    tabs_select_header_label = ttk.Label(root, text="INCLUDE IN THIS REPORT: ", font=('Arial', 24))
    widmg.store_and_place(page['labels'], "tabs_select_header_label", tabs_select_header_label, relx=0.5, rely=0.15, relheight=0.15, anchor='c')


    pdpir_checkbutton = ttk.Checkbutton(root, text="Drill Pipe Inspection (PDPIR)", variable=widmg.pdpir_boolean, style='LargeFont.TCheckbutton')
    widmg.store_and_place(page['check_buttons'], "pdpir_checkbutton", pdpir_checkbutton, relx=0.5, rely=0.24, relwidth=0.75, relheight=0.17, anchor='n')

    hwdp_checkbutton = ttk.Checkbutton(root, text="Heavy Weight Drill Pipe Inspection (HWDP)", variable=widmg.hwdp_boolean, style='LargeFont.TCheckbutton')
    widmg.store_and_place(page['check_buttons'], "hwdp_checkbutton", hwdp_checkbutton, relx=0.5, rely=0.42, relwidth=0.75, relheight=0.17, anchor='n')

    subs_checkbutton = ttk.Checkbutton(root, text="Subs Inspection (SUBS)", variable=widmg.subs_boolean, style='LargeFont.TCheckbutton')
    widmg.store_and_place(page['check_buttons'], "subs_checkbutton", subs_checkbutton, relx=0.50, rely=0.60, relwidth=0.75, relheight=0.17, anchor='n')

    tabs_select_back_button = ttk.Button(root, text="BACK", command=lambda: tabs_select_back_action(widmg, datmg, root), style='Large.TButton')
    widmg.store_and_place(page['buttons'], "tabs_select_back_button", tabs_select_back_button, relx=0.34, rely=0.87, relwidth=0.26, relheight=0.08, anchor='n')

    next_after_tabs_btn = ttk.Button(root, text="NEXT", command=lambda: tabs_select_next_action(widmg, datmg, root), style='Large.TButton')
    widmg.store_and_place(page['buttons'], "next_after_tabs_btn", next_after_tabs_btn, relx=0.62, rely=0.87, relwidth=0.26, relheight=0.08, anchor='n')


def create_tab_select_existing_screen(visible_tabs, widmg, datmg, root):
    branch = datmg.json_data_dict['branch']
    page = widmg.drillpipe_tabs_usemenu
    style = ttk.Style()
    style.configure('LargeFont.TButton', font=('Helvetica', 28))


    tabs_select_header_label = ttk.Label(root, text="WHICH WOULD YOU LIKE TO START?", font=('Arial', 32))
    widmg.store_and_place(page['labels'], "tabs_select_header_label", tabs_select_header_label, relx=0.5, rely=0.15, relheight=0.15, anchor='c')


    for tab in visible_tabs:
        if tab == 'Prop Drill Pipe Inp Report':
            pdpir_strt_button = ttk.Button(root, text="Use PDPIR Tab", command=lambda: sel_btn_to_create_dp_spread("Prop Drill Pipe Inp Report", widmg, datmg, root, existing=True), style='LargeFont.TButton')
            widmg.store_and_place(page['buttons'], "pdpir_strt_button", pdpir_strt_button, relx=0.5, rely=0.24, relwidth=0.75, relheight=0.17, anchor='n')

        elif tab == 'Prop HWDP Inp Report':
            hwdp_strt_button = ttk.Button(root, text="Use HWDP Tab", command=lambda: sel_btn_to_create_dp_spread("Prop HWDP Inp Report", widmg, datmg, root, existing=True), style='LargeFont.TButton')
            widmg.store_and_place(page['buttons'], "hwdp_strt_button", hwdp_strt_button, relx=0.5, rely=0.42, relwidth=0.75, relheight=0.17, anchor='n')

        elif tab == 'Prop Subs Inp Report':
            subs_strt_button = ttk.Button(root, text="Use SUBS Tab", command=lambda: sel_btn_to_create_dp_spread("Prop Subs Inp Report", widmg, datmg, root, existing=True), style='LargeFont.TButton')
            widmg.store_and_place(page['buttons'], "subs_strt_button", subs_strt_button, relx=0.5, rely=0.60, relwidth=0.75, relheight=0.17, anchor='n')

    for tab in visible_tabs:
        if tab == "Tubing Insp Report":
            no_tabs_found_label = ttk.Label(root, text="No Tabs Found (Must Select Drill Pipe Report)", font=('Arial', 24))
            widmg.store_and_place(page['labels'], "no_tabs_found_label", no_tabs_found_label, relx=0.5, rely=0.5, relwidth=0.95, relheight=0.15, anchor='center')
            break

    tabs_select_back_button = ttk.Button(root, text="BACK", command=lambda: tabs_select_back_action(widmg, datmg, root), style='Large.TButton')
    widmg.store_and_place(page['buttons'], "tabs_select_back_button", tabs_select_back_button, relx=0.25, rely=0.87, relwidth=0.3, relheight=0.08, anchor='n')


def tabs_select_next_action(widmg, datmg, root):
    save_tabs_screen_clear(widmg, datmg, root)
    if widmg.pdpir_boolean.get():
        datmg.json_data_dict['report_data']['Prop Drill Pipe Inp Report'] = {
        'joint_count': 0,
        'users_column_select': [],
        'is_complete': False,
        'inspection_type_data': {"inspection_type_selection": None, "inspection_type_addodid": None, "inspection_type_additional": None },
        'joint_data': {}
        }
    if widmg.hwdp_boolean.get():
        datmg.json_data_dict['report_data']['Prop HWDP Inp Report'] = {
        'joint_count': 0,
        'users_column_select': [],
        'is_complete': False,
        'inspection_type_data': {"inspection_type_selection": None, "inspection_type_addodid": None, "inspection_type_additional": None },
        'joint_data': {}
        }
    if widmg.subs_boolean.get():
        datmg.json_data_dict['report_data']['Prop Subs Inp Report'] = {
        'joint_count': 0,
        'users_column_select': [],
        'is_complete': False,
        'inspection_type_data': {"inspection_type_selection": None, "inspection_type_addodid": None, "inspection_type_additional": None },
        'joint_data': {}
        }

    select_active_tab_menu(widmg, datmg, root)

def tabs_select_back_action(widmg, datmg, root):
    branch = datmg.json_data_dict['report_user_metadata']['branch']
    save_tabs_clear(widmg, datmg, root)
    if branch == 'ND':
        does_spreadsheet_exist_screen(widmg, datmg, root)
    elif branch == 'TX':
        select_tx_ss_style_screen(widmg, datmg, root)



def save_tabs_screen_clear(widmg, datmg, root):
    page = widmg.drillpipe_tabs_usemenu

    widmg.hide_widget_grouping(page['check_buttons'])
    widmg.hide_widget_grouping(page['buttons'])
    widmg.hide_widget_grouping(page['labels'])

def select_active_tab_menu(widmg, datmg, root):
    page = widmg.select_active_tab_fresh_menu

    active_tab_fresh_select_header = ttk.Label(root, text="WHICH WOULD YOU LIKE TO START?", font=('Arial', 24))
    widmg.store_and_place(page['labels'], "active_tab_fresh_select_header", active_tab_fresh_select_header, relx=0.5, rely=0.15, relheight=0.15, anchor='c')

    if 'Prop Drill Pipe Inp Report' in datmg.json_data_dict['report_data']:
        pdpir_sel_button = ttk.Button(root, text="Drill Pipe Inspection (PDPIR)", command=lambda: sel_btn_to_create_dp_spread('Prop Drill Pipe Inp Report', widmg, datmg, root), style='Large.TButton')
        widmg.store_and_place(page['buttons'], "pdpir_sel_button", pdpir_sel_button, relx=0.5, rely=0.24, relwidth=0.75, relheight=0.17, anchor='n')
    if 'Prop HWDP Inp Report' in datmg.json_data_dict['report_data']:
        hwdp_sel_button = ttk.Button(root, text="Heavy Weight Drill Pipe Inspection (HWDP)", command=lambda: sel_btn_to_create_dp_spread('Prop HWDP Inp Report', widmg, datmg, root), style='Large.TButton')
        widmg.store_and_place(page['buttons'], "hwdp_sel_button", hwdp_sel_button, relx=0.5, rely=0.42, relwidth=0.75, relheight=0.17, anchor='n')
    if 'Prop Subs Inp Report' in datmg.json_data_dict['report_data']: 
        subs_sel_button = ttk.Button(root, text="Subs Inspection (SUBS)", command=lambda: sel_btn_to_create_dp_spread('Prop Subs Inp Report', widmg, datmg, root), style='LargeFont.TButton')
        widmg.store_and_place(page['buttons'], "subs_sel_button", subs_sel_button, relx=0.50, rely=0.60, relwidth=0.75, relheight=0.17, anchor='n')

    actab_select_back_button = ttk.Button(root, text="BACK", command=lambda: sel_act_tab_bridge_back(widmg, datmg, root), style='Large.TButton')
    widmg.store_and_place(page['buttons'], "actab_select_back_button", actab_select_back_button, relx=0.25, rely=0.87, relwidth=0.3, relheight=0.08, anchor='n')



def sel_act_tab_bridge_back(widmg, datmg, root):
    page = widmg.select_active_tab_fresh_menu
    widmg.hide_widget_grouping(page['buttons'])
    widmg.hide_widget_grouping(page['labels'])
    create_tab_select_screen(widmg, datmg, root)

def sel_btn_to_create_dp_spread(type_rep, widmg, datmg, root, existing=False):
    datmg.json_data_dict['active_tab'] = type_rep
    if type_rep == 'Prop Drill Pipe Inp Report':
        tab_set = 'PDPIR'
    elif type_rep == 'Prop HWDP Inp Report':
        tab_set = 'HWDP'
    elif type_rep == 'Prop Subs Inp Report':
        tab_set = 'SUBS'

    widmg.tab_data_header.set(tab_set)

    if existing:
        page = widmg.drillpipe_tabs_usemenu
    else:   
        page = widmg.select_active_tab_fresh_menu
    widmg.hide_widget_grouping(page['buttons'])
    widmg.hide_widget_grouping(page['labels']) 

    create_report_metadata_input_widgets(widmg, datmg, root, editing=False)

def ask_for_data_entry_file_screen(widmg, datmg, root):
    page = widmg.ask_datentry_file_window

    de_fileask_label = ttk.Label(root, text="How Would You Like To Proceed For ", font=('Arial', 32))
    widmg.store_and_place(page['labels'], "de_fileask_label", de_fileask_label, relx=0.5, rely=0.15, relheight=0.15, anchor='c')

    create_new_defile_btn = ttk.Button(root, text="Begin Fresh Session", command=lambda: create_new_defile_action(widmg, datmg, root), style='Large.TButton')
    widmg.store_and_place(page['buttons'], "create_new_defile_btn", create_new_defile_btn, relx=0.5, rely=0.24, relwidth=0.75, relheight=0.17, anchor='n')

    upload_previous_defile_btn = ttk.Button(root, text="Upload Previous File", command=lambda: load_previous_defile_button(widmg, datmg, root), style='Large.TButton')
    widmg.store_and_place(page['buttons'], "upload_previous_defile_btn", upload_previous_defile_btn, relx=0.5, rely=0.42, relwidth=0.75, relheight=0.17, anchor='n')

    defile_back_button = ttk.Button(root, text="BACK", command=lambda: defile_back_action(widmg, datmg, root), style='Large.TButton')
    widmg.store_and_place(page['buttons'], "defile_back_button", defile_back_button, relx=0.5, rely=0.87, relwidth=0.5, relheight=0.08, anchor='n')

def defile_back_action(widmg, datmg, root):
    report_type = datmg.json_data_dict['report_user_metadata']['report_type']
    branch = datmg.json_data_dict['report_user_metadata']['branch']

    page = widmg.ask_datentry_file_window
    widmg.hide_widget_grouping(page['labels'])
    widmg.hide_widget_grouping(page['buttons'])

    if branch == 'ND':
        if datmg.xel_file_path != None:
            if report_type == 'Drill Pipe Inspection Report':
                active_tabs = return_sheets_active_tabs_after(widmg, datmg, root)
                create_tab_select_existing_screen(active_tabs, widmg, datmg, root)
            elif report_type == 'Tubing/Casing Report':
                does_spreadsheet_exist_screen(widmg, datmg, root)
        else:
            if report_type == 'Drill Pipe Inspection Report':
                select_active_tab_menu(widmg, datmg, root)
            elif report_type == 'Tubing/Casing Report':
                does_spreadsheet_exist_screen(widmg, datmg, root)
    elif branch == 'TX':
        if datmg.xel_file_path != None:
            active_tabs = return_sheets_active_tabs_after(widmg, datmg, root)
            create_tab_select_existing_screen(active_tabs, widmg, datmg, root)
        else:
            select_active_tab_menu(widmg, datmg, root)



def create_new_defile_action(widmg, datmg, root):
    hide_de_file_ask_screen(widmg, datmg, root)
    create_report_metadata_input_widgets(widmg, datmg, root, editing=False)
    
def hide_de_file_ask_screen(widmg, datmg, root):
    page = widmg.ask_datentry_file_window   
    widmg.hide_widget_grouping(page['labels'])
    widmg.hide_widget_grouping(page['buttons'])

def load_previous_defile_button(widmg, datmg, root):
    hide_de_file_ask_screen(widmg, datmg, root)
    page = widmg.tubing_or_drillpipe_menu
    widmg.hide_widget_grouping(page['buttons'])
    widmg.hide_widget_grouping(page['labels'])
    sel_saved_report_scrn(widmg, datmg, root)


def sel_saved_report_scrn(widmg, datmg, root):
    page = widmg.saved_file_select_window
    file_selecting_frame = tk.Frame(root, bd=1, relief=tk.GROOVE)
    widmg.store_and_place(page['display_frames'], "file_selecting_frame", file_selecting_frame, relx=.02, rely=0.15, relwidth=0.40, relheight=0.70)
    sel_metadisp_frame = tk.Frame(root, bd=1, relief=tk.GROOVE)
    widmg.store_and_place(page['display_frames'], "sel_metadisp_frame", sel_metadisp_frame, relx=.45, rely=0.15, relwidth=0.50, relheight=0.70)
    report_type_header = tk.Label(root, textvariable=widmg.spreadsheet_type_prework, bd=1, relief='solid', font=('Arial', 16, 'bold'), wraplength=(root_width * 0.18))
    widmg.store_and_place(page['labels'], "report_type_header", report_type_header, font_changer=True, relx=0.73, rely=0.01, relwidth=0.18, relheight=0.12)

    file_listbox = tk.Listbox(file_selecting_frame)
    widmg.store_and_place(page['listboxes'], "file_listbox", file_listbox, relx=0.001, rely=0.001, relwidth=0.9, relheight=0.99)

    scrollbar = tk.Scrollbar(file_selecting_frame, orient='vertical', command=file_listbox.yview)
    widmg.store_and_place(page['scrollbars'], "scrollbar", scrollbar, relx=0.90, rely=0.001, relwidth=0.09, relheight=0.99)
    file_listbox.bind('<<ListboxSelect>>', lambda event: on_file_select(event, widmg))
    file_listbox.config(yscrollcommand=scrollbar.set)

    for file in list_json_files('data_entry_files'):
        file_listbox.insert(tk.END, file)


    def on_file_select(event, widmg):
        selection = file_listbox.curselection()
        if selection:
            filename = file_listbox.get(selection[0])
            update_metadata_display(filename, widmg)

    select_file_header = tk.Label(root, text="Upload Report File", font=('Arial', 32))
    widmg.store_and_place(page['labels'], "select_file_header", select_file_header, relx=0.20, rely=0.01, relwidth=0.5, relheight=0.12)

    meta_operator = tk.Label(sel_metadisp_frame, textvariable=widmg.operator_msv, bd=1, relief="solid", font=('Arial', 18))
    widmg.store_and_place(page['labels'], "meta_operator", meta_operator, relx=0.01, rely=0.01, relwidth=0.98, relheight=0.12)
    meta_date = tk.Label(sel_metadisp_frame, textvariable=widmg.date_msv, bd=1, relief="solid", font=('Arial', 18))
    widmg.store_and_place(page['labels'], "meta_date", meta_date, relx=0.01, rely=0.12, relwidth=0.98, relheight=0.14)
    meta_contractor = tk.Label(sel_metadisp_frame, textvariable=widmg.contractor_msv, bd=1, relief="solid", font=('Arial', 18))
    widmg.store_and_place(page['labels'], "meta_contractor", meta_contractor, relx=0.01, rely=0.26, relwidth=0.98, relheight=0.12)
    meta_invoice = tk.Label(sel_metadisp_frame, textvariable=widmg.invoice_msv, bd=1, relief="solid", font=('Arial', 18))
    widmg.store_and_place(page['labels'], "meta_invoice", meta_invoice, relx=0.01, rely=0.38, relwidth=0.98, relheight=0.12)
    meta_location = tk.Label(sel_metadisp_frame, textvariable=widmg.location_msv, bd=1, relief="solid", font=('Arial', 18))
    widmg.store_and_place(page['labels'], "meta_location", meta_location, relx=0.01, rely=0.50, relwidth=0.98, relheight=0.12)
    meta_inspecttype = tk.Label(sel_metadisp_frame, textvariable=widmg.inspecttype_msv, bd=1, relief="solid", font=('Arial', 16), wraplength=350)
    widmg.store_and_place(page['labels'], "meta_inspecttype", meta_inspecttype, relx=0.01, rely=0.62, relwidth=0.98, relheight=0.14)
    meta_connectsize = tk.Label(sel_metadisp_frame, textvariable=widmg.connectsize_msv, bd=1, relief="solid", font=('Arial', 18))
    widmg.store_and_place(page['labels'], "meta_connectsize", meta_connectsize, relx=0.01, rely=0.76, relwidth=0.98, relheight=0.13)
    meta_conntype = tk.Label(sel_metadisp_frame, textvariable=widmg.conntype_msv, bd=1, relief="solid", font=('Arial', 18))
    widmg.store_and_place(page['labels'], "meta_conntype", meta_conntype, relx=0.01, rely=0.87, relwidth=0.98, relheight=0.12)

    choose_report_btn = ttk.Button(root, text="Load Report", command=lambda: choose_report_action(widmg, datmg, root), style='Large.TButton')
    widmg.store_and_place(page['buttons'], "choose_report_btn", choose_report_btn, relx=0.70, rely=0.87, relwidth=0.25, relheight=0.10)

    back_to_main_btn = ttk.Button(root, text="<< BACK", command=lambda: back_to_main_action(widmg, datmg, root), style='Large.TButton')
    widmg.store_and_place(page['buttons'], "back_to_main_btn", back_to_main_btn, relx=0.35, rely=0.87, relwidth=0.25, relheight=0.10)

    delete_file_btn = ttk.Button(root, text="Delete File", command=lambda: delete_selected_file(widmg), style='Large.TButton')
    widmg.store_and_place(page['buttons'], "delete_file_btn", delete_file_btn, relx=0.02, rely=0.87, relwidth=0.20, relheight=0.10)

def choose_report_action(widmg, datmg, root):
    file_listbox = widmg.saved_file_select_window['listboxes']['file_listbox']['widget']
    selection = file_listbox.curselection()
    if selection:
        datmg.filename = file_listbox.get(selection[0])
        datmg.load_file_to_dict(datmg.filename)
    else:
        print("No file selected")
    hide_saved_file_select_widgets(widmg, datmg, root)
    
    if datmg.json_data_dict['report_user_metadata']['active_tab'] == 'Tubing Insp Report':
            widmg.tab_data_header.set(datmg.json_data_dict['report_user_metadata']['connection_type_selection'])

    display_main_report_screen(widmg, datmg, root)
    load_row_data(widmg, datmg, root)
    update_all_row_cells(widmg, datmg, root)

def list_json_files(directory='data_entry_files'):
    return [f for f in os.listdir(directory) if f.endswith('.json')]

def update_metadata_display(filename, widmg):
    try:
        with open(filename, 'r') as file:
            data = json.load(file)
            metadata = data.get('report_user_metadata', {})

        # Update each StringVar with the corresponding metadata value
        widmg.operator_msv.set(f"{metadata.get('operator_entry', '')}")
        widmg.date_msv.set(f"{metadata.get('date_entry', '')}")
        widmg.contractor_msv.set(f"{metadata.get('contractor_entry', '')}")
        widmg.invoice_msv.set(f"{metadata.get('invoice_entry', '')}")
        widmg.location_msv.set(f"{metadata.get('location_entry', '')}")
        widmg.inspecttype_msv.set(f"{metadata.get('inspection_type_selection', '')}")
        widmg.connectsize_msv.set(f"{metadata.get('connection_size_selection', '')}")
        widmg.conntype_msv.set(f"{metadata.get('connection_type_selection', '')}")
    except Exception as e:
        print(f"Error reading file {filename}: {e}")


def hide_saved_file_select_widgets(widmg, datmg, root):
    page = widmg.saved_file_select_window
    widmg.hide_widget_grouping(page['display_frames'])
    widmg.hide_widget_grouping(page['buttons'])
    widmg.hide_widget_grouping(page['labels'])
    widmg.hide_widget_grouping(page['listboxes'])
    widmg.hide_widget_grouping(page['scrollbars'])
    widmg.hide_widget_grouping(page['meta_labels'])

def back_to_main_action(widmg, datmg, root):
    hide_saved_file_select_widgets(widmg, datmg, root)
    ask_for_data_entry_file_screen(widmg, datmg, root)

def delete_selected_file(widmg):
    file_listbox = widmg.saved_file_select_window['listboxes']['file_listbox']['widget']
    selection = file_listbox.curselection()
    if selection:
        filename = file_listbox.get(selection[0])
        if os.path.exists(filename):
            os.remove(filename)
            file_listbox.delete(selection[0])
            print(f"File {filename} deleted.")
        else:
            print(f"File {filename} not found.")
    else:
        print("No file selected.")

def select_back_to_choose_button(widmg, datmg, root, editing=False):
    report_type = datmg.json_data_dict['report_user_metadata']['report_type']
    page = widmg.main_metadata_widgets
    widmg.hide_widget_grouping(page['entries'])
    widmg.hide_widget_grouping(page['labels'])
    widmg.hide_widget_grouping(page['buttons'])
    widmg.hide_widget_grouping(page['dropdown_menus'])

    
    if editing:
        display_main_report_screen(widmg, datmg, root, editing=editing)
    else:
        if report_type == "Drill Pipe Inspection Report":
            select_active_tab_menu(widmg, datmg, root)
        else:
            does_spreadsheet_exist_screen(widmg, datmg, root)
        


def create_report_metadata_input_widgets(widmg, datmg, root, editing=False):
    report_type = datmg.json_data_dict['report_type']
    tab_type = datmg.json_data_dict['active_tab']

    page = widmg.main_metadata_widgets

    if datmg.json_data_dict['report_type'] == 'Tubing/Casing Report':
        inspection_type_list, connection_size_list, connection_type_list = datmg.tubing_inspection_type_list, datmg.tubing_conn_size_list, datmg.tubing_conn_type_list 
    elif datmg.json_data_dict['report_type'] == 'Drill Pipe Inspection Report':
        if datmg.json_data_dict['branch'] == 'TX':
            inspection_type_list, connection_size_list, connection_type_list = datmg.dp_inspection_type_list_short, datmg.dp_conn_size_list, datmg.dp_conn_type_list
        else:
            inspection_type_list, connection_size_list, connection_type_list = datmg.dp_inspection_type_list_short, datmg.dp_conn_size_list, datmg.dp_conn_type_list

    report_type_label = ttk.Label(root, text=f"Enter Metadata for this {widmg.tab_data_header.get()} Report", font=('Helvetica', 26))
    widmg.store_and_place(page['labels'], "report_type_label", report_type_label, relx=0.01, rely=0.07, relwidth=0.90, relheight=0.14, anchor='w')

    # Helper function to prepopulate Entry widgets
    def create_prepopulated_entry(key, root, font, relx, rely, relwidth, relheight):
        entry_value = datmg.json_data_dict['report_user_metadata'].get(key, "")
        entry = ttk.Entry(root, font=font)
        entry.insert(0, entry_value)
        widmg.store_and_place(page['entries'], key, entry, relx=relx, rely=rely, relwidth=relwidth, relheight=relheight)

    # Helper function to prepopulate ComboBox widgets
    def create_prepopulated_combobox(key, values, root, font, relx, rely, relwidth, relheight):
        combobox_value = datmg.json_data_dict['report_user_metadata'].get(key, "")
        combobox = ttk.Combobox(root, values=values, font=font, state='normal')  # Allows typing and selection
        combobox.set(combobox_value)
        widmg.store_and_place(page['dropdown_menus'], key, combobox, relx=relx, rely=rely, relwidth=relwidth, relheight=relheight)

    # Operator - updated to use create_prepopulated_combobox
    operator_label = ttk.Label(root, text="Operator:", font=('Arial', 16))
    widmg.store_and_place(page['labels'], "operator_label", operator_label, relx=0.03, rely=0.15, relwidth=0.30, relheight=0.1)
    
    # Use the existing helper function to create the combobox
    create_prepopulated_combobox("operator_entry", datmg.nd_operator_vals, root, ('Arial', 16), 0.18, 0.15, 0.36, 0.1)

    # Date
    date_label = ttk.Label(root, text="Date:", font=('Arial', 16))
    widmg.store_and_place(page['labels'], "date_label", date_label, relx=0.66, rely=0.15, relwidth=0.3, relheight=0.1)
    create_prepopulated_entry("date_entry", root, ('Arial', 16), 0.74, 0.15, 0.23, 0.1)
    
    # Contractor
    contractor_label = ttk.Label(root, text="Contractor/Rig:", font=('Arial', 16))
    widmg.store_and_place(page['labels'], "contractor_label", contractor_label, relx=0.03, rely=0.30, relwidth=0.30, relheight=0.1)
    create_prepopulated_entry("contractor_entry", root, ('Arial', 16), 0.18, 0.30, 0.36, 0.1)

    
    # Field Invoice
    invoice_label = ttk.Label(root, text="Field Invoice:", font=('Arial', 16))
    widmg.store_and_place(page['labels'], "invoice_label", invoice_label, relx=0.56, rely=0.30, relwidth=0.3, relheight=0.1)
    create_prepopulated_entry("invoice_entry", root, ('Arial', 16), 0.74, 0.30, 0.23, 0.1)
    
    # Location
    location_label = ttk.Label(root, text="Location:", font=('Arial', 16))
    widmg.store_and_place(page['labels'], "location_label", location_label, relx=0.03, rely=0.44, relwidth=0.37, relheight=0.1)
    create_prepopulated_entry("location_entry", root, ('Arial', 16), 0.18, 0.44, 0.36, 0.1)
    
    add_notes_button = ttk.Button(root, text="Add/Edit Notes", command=lambda: add_notes_button_bridge(widmg, datmg, root), style='Large.TButton')
    widmg.store_and_place(page['buttons'], "add_notes_button", add_notes_button, relx=0.56, rely=0.44, relwidth=0.40, relheight=0.1)
    
    # Inspection Type
    inspection_label = ttk.Label(root, text="Inspection Type:", font=('Arial', 16))
    widmg.store_and_place(page['labels'], "inspection_label", inspection_label, relx=0.03, rely=0.58, relwidth=0.33, relheight=0.1)
    create_prepopulated_combobox("inspection_type_selection", inspection_type_list, root, ('Arial', 12), 0.18, 0.58, 0.55, 0.1)
    
    add_inspectinfo_btn = ttk.Button(root, text="Add Additional Info", command=lambda: create_addinfo_window(widmg, datmg, root), style='Large.TButton')
    widmg.store_and_place(page['buttons'], "add_inspectinfo_btn", add_inspectinfo_btn, relx=0.74, rely=0.58, relwidth=0.23, relheight=0.1)
    
    # Connection Size
    connection_size_label = ttk.Label(root, text="Connection Size:", font=('Arial', 16))
    widmg.store_and_place(page['labels'], "connection_size_label", connection_size_label, relx=0.03, rely=0.70, relwidth=0.3, relheight=0.1)
    create_prepopulated_combobox("connection_size_selection", connection_size_list, root, ('Arial', 16), 0.26, 0.70, 0.22, 0.1)
    
    # Connection Type
    connection_type_label = ttk.Label(root, text="Connection Type:", font=('Arial', 16))
    widmg.store_and_place(page['labels'], "connection_type_label", connection_type_label, relx=0.50, rely=0.70, relwidth=0.3, relheight=0.1)
    create_prepopulated_combobox("connection_type_selection", connection_type_list, root, ('Arial', 16), 0.74, 0.70, 0.23, 0.1)
    
    add_grade_info_label = ttk.Button(root, text="Add Grade Info", command=lambda: add_grade_info_btn_bridge(widmg, datmg, root), style='Small.TButton')
    widmg.store_and_place(page['buttons'], "add_grade_info_label", add_grade_info_label, relx=0.50, rely=0.80, relwidth=0.23, relheight=0.05)
    grade_addition_label = ttk.Label(root, textvariable=widmg.grade_info_tube, font=('Arial', 12, 'italic'))
    widmg.store_and_place(page['labels'], "grade_addition_label", grade_addition_label, relx=0.74, rely=0.80, relwidth=0.23, relheight=0.05)
    
    inspected_by_label = ttk.Label(root, text="Inspected By:", font=('Arial', 16))
    widmg.store_and_place(page['labels'], "inspected_by_label", inspected_by_label, relx=0.56, rely=0.86, relwidth=0.15, relheight=0.1)
    create_prepopulated_entry("inspected_by_entry", root, ('Arial', 16), 0.71, 0.86, 0.26, 0.1)
    
    # Submit Button
    submit_button = ttk.Button(root, text="Submit Data", command=lambda: validate_and_submit_data(widmg, datmg, root, editing=editing), style='Large.TButton')
    widmg.store_and_place(page['buttons'], "submit_button", submit_button, relx=0.03, rely=0.87, relwidth=0.25, relheight=0.10)
    
    # Back Button (To DP/Tube Select)
    back_to_choose_button = ttk.Button(root, text="Go Back", command=lambda: select_back_to_choose_button(widmg, datmg, root, editing=editing), style='Large.TButton')
    widmg.store_and_place(page['buttons'], "back_to_choose_button", back_to_choose_button, relx=0.29, rely=0.87, relwidth=0.25, relheight=0.10)



def validate_and_submit_data(widmg, datmg, root, editing=False):
    connection_size = widmg.main_metadata_widgets['dropdown_menus']['connection_size_selection']['widget'].get()
    connection_type = widmg.main_metadata_widgets['dropdown_menus']['connection_type_selection']['widget'].get()
    
    if connection_size in datmg.invalid_combinations_cs_dt and connection_type in datmg.invalid_combinations_cs_dt[connection_size]:
        # Display warning message
        messagebox.showwarning("Invalid Selection", f"{connection_size} and {connection_type} is not a valid selection.")
    else:
        # Call the original submit function if the combination is valid
        submit_data_button(widmg, datmg, root, editing)



def add_notes_button_bridge(widmg, datmg, root): 
    page = widmg.main_metadata_widgets
    widmg.hide_widget_grouping(page['entries'])
    widmg.hide_widget_grouping(page['labels'])
    widmg.hide_widget_grouping(page['buttons'])
    widmg.hide_widget_grouping(page['dropdown_menus'])

    open_add_notes_window(widmg, datmg, root)

def open_add_notes_window(widmg, datmg, root):
    datmg.update_user_report_metadata()

    page = widmg.add_notes_window

    add_notes_label = ttk.Label(root, text="Add Notes:", font=('Helvetica', 30))
    widmg.store_and_place(page['labels'], "add_notes_label", add_notes_label, relx=0.03, rely=0.03, relwidth=0.65, relheight=0.10)

    notes_entry = tk.Text(root, font=('Arial', 16))
    widmg.store_and_place(page['text'], "notes_entry", notes_entry, relx=0.03, rely=0.15, relwidth=0.90, relheight=0.65)

    report_type = datmg.json_data_dict['report_user_metadata'].get('report_type', '')

    if 'notes_entry' in datmg.json_data_dict['report_user_metadata']:
        notes_text = datmg.json_data_dict['report_user_metadata']['notes_entry']
        notes_entry.insert('1.0', notes_text)
    else:
        # If no previous notes, check if pre-fill text exists for the report_type
        if report_type in datmg.notes_prefill:
            pre_fill_text = datmg.notes_prefill[report_type]
            notes_entry.insert('1.0', pre_fill_text)
            # Save the pre-filled text as the current notes entry
            datmg.json_data_dict['report_user_metadata']['notes_entry'] = pre_fill_text


    notes_save_button = ttk.Button(root, text="SAVE", command=lambda: save_notes_button(notes_entry, widmg, datmg, root), style='Large.TButton')
    widmg.store_and_place(page['buttons'], "notes_save_button", notes_save_button, relx=0.53, rely=0.85, relwidth=0.35, relheight=0.10)

    notes_back_button = ttk.Button(root, text="CANCEL", command=lambda: back_from_notes_button(widmg, datmg, root), style='Large.TButton')
    widmg.store_and_place(page['buttons'], "notes_back_button", notes_back_button, relx=0.12, rely=0.85, relwidth=0.35, relheight=0.1)


def bring_back_metadata_from_notes(widmg, datmg, root):
    page = widmg.main_metadata_widgets

    widmg.place_back_widgets(page['buttons'])
    widmg.place_back_widgets(page['labels'])


    combined_widgets = list(page['entries'].items()) + list(page['dropdown_menus'].items())
    for key, widget_info in combined_widgets:
        widget = widget_info['widget']
        user_input = datmg.json_data_dict['report_user_metadata'].get(key, "")
        widget.delete(0, 'end')  # clear the current content
        widget.insert(0, user_input)  # insert the user input at the beginning of the entry field
        widget.place(**widget_info['params'])



def save_notes_button(notes_entry, widmg, datmg, root):
    datmg.json_data_dict['report_user_metadata']['notes_entry'] = notes_entry.get("1.0", "end")
    datmg.save_dict_to_file()
    back_from_notes_button(widmg, datmg, root)


def back_from_notes_button(widmg, datmg, root):
    page = widmg.add_notes_window 

    widmg.hide_widget_grouping(page['labels'])
    widmg.hide_widget_grouping(page['text'])
    widmg.hide_widget_grouping(page['buttons'])
    bring_back_metadata_from_notes(widmg, datmg, root)

def add_grade_info_btn_bridge(widmg, datmg, root):
    if hasattr(widmg, 'addgradeinfo_window') and widmg.addgradeinfo_window.winfo_exists():
        # If the window already exists, bring it to the front
        widmg.addgradeinfo_window.lift()
        return

    metadata = datmg.json_data_dict['report_user_metadata']
    widmg.addgradeinfo_window = tk.Toplevel(root)
    widmg.addgradeinfo_window.title("Add Additional Inspection Type Info")
    widmg.addgradeinfo_window.geometry("485x275")
    additional_grade_info = metadata.get('grade_info')

    entry_label = ttk.Label(widmg.addgradeinfo_window, text="Add Grade Info", font=('Arial', 14, "bold"))
    entry_label.place(relx=0.50, rely=0.30, relwidth=0.30, relheight=0.15, anchor='center')

    widmg.manual_addgrade_entry = ttk.Entry(widmg.addgradeinfo_window, font=('Arial', 14))
    if additional_grade_info:
        widmg.manual_addgrade_entry.insert(0, additional_grade_info)
    widmg.manual_addgrade_entry.place(relx=0.5, rely=0.50, relwidth=0.95, relheight=0.15, anchor='center')

    confirm_button = ttk.Button(widmg.addgradeinfo_window, text="Confirm", command=lambda: add_grade_bridge(widmg, datmg, root))
    confirm_button.place(relx=0.33, rely=0.75, relwidth=0.25, relheight=0.15, anchor='center')

    cancel_button = ttk.Button(widmg.addgradeinfo_window, text="Cancel", command=widmg.addgradeinfo_window.destroy)
    cancel_button.place(relx=0.67, rely=0.75, relwidth=0.25, relheight=0.15, anchor='center')

    # Ensure the window is properly cleaned up when closed
    widmg.addgradeinfo_window.protocol("WM_DELETE_WINDOW", lambda: on_addgradeinfo_window_close(widmg))

def on_addgradeinfo_window_close(widmg):
    widmg.addgradeinfo_window.destroy()
    del widmg.addgradeinfo_window

def create_addinfo_window(widmg, datmg, root):
    # Create a new top-level window
    metadata = datmg.json_data_dict['report_user_metadata']
    widmg.addaddinfo_window = tk.Toplevel(root)
    widmg.addaddinfo_window.title("Add Additional Inspection Type Info")
    widmg.addaddinfo_window.geometry("485x275")
    additional_info = metadata.get('inspection_type_additional')
    add_odid = metadata.get('inspection_type_addodid')
    if not add_odid:
        widmg.actual_odid_boolean = tk.BooleanVar()

    style = ttk.Style()
    style.configure('LargeFont.TCheckbutton', font=('Arial', 14, "bold"))
    
    actual_odid_check = ttk.Checkbutton(widmg.addaddinfo_window, text="Actual OD's, ID's, and Tong Space", variable=widmg.actual_odid_boolean, style='LargeFont.TCheckbutton')
    actual_odid_check.place(relx=0.5, rely=0.25, relwidth=0.95, relheight=0.15, anchor='center')


    entry_label = ttk.Label(widmg.addaddinfo_window, text="Additional Info", font=('Arial', 14, "bold"))
    entry_label.place(relx=0.50, rely=0.45, relwidth=0.30, relheight=0.15, anchor='center')

    # Create an Entry widget, populate it with the provided value, and make it read-only
    widmg.manual_addinfo_entry = ttk.Entry(widmg.addaddinfo_window, font=('Arial', 14))
    if additional_info:
        widmg.manual_addinfo_entry.insert(0, additional_info)
    widmg.manual_addinfo_entry.place(relx=0.5, rely=0.60, relwidth=0.95, relheight=0.15, anchor='center')

    # Create a button that for now does nothing when clicked
    # You will replace command=lambda: None with your actual function later
    confirm_button = ttk.Button(widmg.addaddinfo_window, text="Confirm", command=lambda: add_info_bridge(widmg, datmg, root))
    confirm_button.place(relx=0.33, rely=0.80, relwidth=0.25, relheight=0.15, anchor='center')

    cancel_button = ttk.Button(widmg.addaddinfo_window, text="Cancel", command=widmg.addaddinfo_window.destroy)
    cancel_button.place(relx=0.67, rely=0.80, relwidth=0.25, relheight=0.15, anchor='center')

def add_info_bridge(widmg, datmg, root):
    datmg.json_data_dict['report_user_metadata']['inspection_type_addodid'] = widmg.actual_odid_boolean.get()
    datmg.json_data_dict['report_user_metadata']['inspection_type_additional'] = widmg.manual_addinfo_entry.get()
    if datmg.filename is not None:
        datmg.save_dict_to_file()
    widmg.addaddinfo_window.destroy()

def add_grade_bridge(widmg, datmg, root):
    datmg.json_data_dict['report_user_metadata']['grade_info'] = widmg.manual_addgrade_entry.get()
    if datmg.filename is not None:
        datmg.save_dict_to_file()
    widmg.grade_info_tube.set(widmg.manual_addgrade_entry.get())
    widmg.addgradeinfo_window.destroy()



def submit_data_button(widmg, datmg, root, editing=False):
    page = widmg.main_metadata_widgets

    datmg.update_user_report_metadata()

    if datmg.filename is None:
        datmg.initialize_json(widmg)

    datmg.save_dict_to_file()

    # Place the static labels
    for key, widget_info in page['labels'].items():
        label_widget = widget_info['widget']
        placement_params = widget_info['params']
        label_widget.place(**placement_params)

    widmg.hide_widget_grouping(page['entries'])
    widmg.hide_widget_grouping(page['buttons'])
    widmg.hide_widget_grouping(page['dropdown_menus'])

    show_review_screen(widmg, datmg, root, editing=editing)


def show_review_screen(widmg, datmg, root, editing=False):
    page = widmg.main_metadata_widgets

    widmg.hide_widget_grouping(widmg.table_select_page_widgets['buttons'])
    widmg.hide_widget_grouping(widmg.table_select_page_widgets['labels'])
    widmg.place_back_widgets(page['labels'])

    combined_widgets = list(page['entries'].items()) + list(page['dropdown_menus'].items())

    # Replace entry fields with review_labels
    for key, widget_info in combined_widgets:
        entry_params = widget_info['params']
        user_input = datmg.json_data_dict['report_user_metadata'][key]
        review_label = ttk.Label(root, text=user_input, font=('Arial', 18, 'bold italic'))
        if key == "inspection_type_selection":
            widmg.store_and_place(page['review_widgets']['review_labels'], key, review_label, relx=0.25, rely=0.58, relwidth=0.73, relheight=0.1)
        else:
            widmg.store_and_place(page['review_widgets']['review_labels'], key, review_label, **entry_params)


    # Add "Proceed" and "Edit Data" buttons
    proceed_button = ttk.Button(root, text="Proceed", command=lambda: proceed_fn_button(widmg, datmg, root, editing=editing), style='Large.TButton')
    widmg.store_and_place(page['review_widgets']['review_buttons'], "proceed_button", proceed_button, relx=0.29, rely=0.87, relwidth=0.25, relheight=0.10)
    edit_data_button = ttk.Button(root, text="Edit Data", command=lambda: edit_data_fn_button(widmg, datmg, root), style='Large.TButton')
    widmg.store_and_place(page['review_widgets']['review_buttons'], "edit_button", edit_data_button, relx=0.03, rely=0.87, relwidth=0.25, relheight=0.10)

def edit_data_fn_button(widmg, datmg, root):
    page = widmg.main_metadata_widgets

    widmg.hide_widget_grouping(page['review_widgets']['review_buttons'])
    widmg.hide_widget_grouping(page['review_widgets']['review_labels'])

    widmg.place_back_widgets(page['buttons'])
    widmg.place_back_widgets(page['entries'])


    combined_widgets = list(page['entries'].items()) + list(page['dropdown_menus'].items())
    for key, widget_info in combined_widgets:
        widget = widget_info['widget']
        user_input = datmg.json_data_dict['report_user_metadata'].get(key, "")
        widget.delete(0, 'end')  # clear the current content
        widget.insert(0, user_input)  # insert the user input at the beginning of the entry field
        widget.place(**widget_info['params'])


def proceed_fn_button(widmg, datmg, root, editing=False):
    datmg.update_user_report_metadata()
    datmg.save_dict_to_file()

    widmg.hide_widget_grouping(widmg.main_metadata_widgets['review_widgets']['review_buttons'])
    widmg.hide_widget_grouping(widmg.main_metadata_widgets['review_widgets']['review_labels'])
    widmg.hide_widget_grouping(widmg.main_metadata_widgets['labels'])
    widmg.hide_widget_grouping(widmg.main_metadata_widgets['entries'])
    widmg.hide_widget_grouping(widmg.main_metadata_widgets['buttons'])
    widmg.hide_widget_grouping(widmg.main_metadata_widgets['dropdown_menus'])

    show_table_selection_screen(widmg, datmg, root, editing=editing)


def show_table_selection_screen(widmg, datmg, root, editing=False):
    widmg.current_joint_number.set('1')
    widmg.magni_header.set("")
    widmg.magni_content.set("")

    if datmg.json_data_dict['active_tab'] == 'Tubing Insp Report':
        widmg.tab_data_header.set(datmg.json_data_dict['report_user_metadata']['connection_type_selection'])

    widmg.hide_widget_grouping(widmg.main_table_display_widgets['back_buttons'])
    widmg.hide_widget_grouping(widmg.custom_column_selection_page_widgets['dropdown_menus'])
    widmg.hide_widget_grouping(widmg.custom_column_selection_page_widgets['labels'])
    widmg.hide_widget_grouping(widmg.custom_column_selection_page_widgets['buttons'])

    widmg.hide_widget_grouping(widmg.main_table_display_widgets['display_frames'])
    widmg.hide_widget_grouping(widmg.main_table_display_widgets['buttons'])
    
    if editing == True:
        display_main_report_screen(widmg, datmg, root, editing=editing)
        update_all_row_cells(widmg, datmg, root)
    else:
        active_tab = datmg.json_data_dict['active_tab']
        datmg.json_data_dict['report_data'][active_tab]['joint_data'] = {}
        widmg.prev_first_row_label_list = []
        widmg.prev_sec_row_label_list = []
        widmg.next_first_row_label_list = []
        widmg.next_second_row_label_list = []
        select_custom_columns_fn_button(widmg, datmg, root)





def select_custom_columns_fn_button(widmg, datmg, root):
    widmg.hide_widget_grouping(widmg.table_select_page_widgets['buttons'])
    widmg.hide_widget_grouping(widmg.table_select_page_widgets['labels'])
    add_dropdowns(widmg, datmg, root)


def determine_allowed_columns(widmg, datmg, root):
    branch = datmg.json_data_dict['branch']
    report_type = datmg.json_data_dict['report_type']
    report_style = datmg.json_data_dict['report_style'] if 'report_style' in datmg.json_data_dict else None
    active_tab = datmg.json_data_dict['active_tab']

    metadata = datmg.json_data_dict['report_user_metadata']
    if branch == "ND":
        cols = list(datmg.nd_column_types.keys())
        if report_type == 'Drill Pipe Inspection Report':
            if active_tab == 'Prop Drill Pipe Inp Report':
                return [key for key in cols if key not in ['DESCRIPTION', 'SUBS CONN/DATA', 'Visual OD']]
            elif active_tab == 'Prop HWDP Inp Report':
                return [key for key in cols if key not in ['UT', 'DESCRIPTION', 'SUBS CONN/DATA', 'Visual OD']]
            elif active_tab == 'Prop Subs Inp Report':
                return [key for key in cols if key not in ['UT', 'Visual OD']]
        else:
            return [key for key in cols if key not in ['DESCRIPTION', 'SUBS CONN/DATA', 'SERIAL', 'TS-BOX', 'TS-PIN', 'OD', 'ID']]
    elif branch == 'TX':
        cols = list(datmg.tx_column_types.keys())
        if report_style in ['Class 2 DBR', 'Class 2 NOT DBR']:
            if active_tab == 'Prop Drill Pipe Inp Report':
                return [key for key in cols if key not in ['DESCRIPTION', 'SUBS CONN/DATA', 'BORBAK', 'STRES REL GRV', 'C BORE', 'PIN NOSE DIA', 'SEAL WIDTH']]
            elif active_tab == 'Prop HWDP Inp Report':
                return [key for key in cols if key not in ['UT', 'DESCRIPTION', 'SUBS CONN/DATA', 'BORBAK', 'STRES REL GRV', 'C BORE', 'PIN NOSE DIA', 'SEAL WIDTH']]
            elif active_tab == 'Prop Subs Inp Report':
                return [key for key in cols if key not in ['UT', 'BORBAK', 'STRES REL GRV', 'C BORE', 'PIN NOSE DIA', 'SEAL WIDTH']]
        elif report_style == 'Full Dimensional':
            if active_tab == 'Prop Drill Pipe Inp Report':
                return [key for key in cols if key not in ['DESCRIPTION', 'SUBS CONN/DATA', 'BORBAK', 'STRES REL GRV']]
            elif active_tab == 'Prop HWDP Inp Report':
                return [key for key in cols if key not in ['UT', 'DESCRIPTION', 'SUBS CONN/DATA']]
            elif active_tab == 'Prop Subs Inp Report':
                return [key for key in cols if key not in ['UT']]
    else:
        return


def add_dropdowns(widmg, datmg, root, editing=False):
    active_tab = datmg.json_data_dict['active_tab']
    page = widmg.custom_column_selection_page_widgets
    options = determine_allowed_columns(widmg, datmg, root)
    
    if editing:
        current_selections = {f"combo_box{i}": datmg.json_data_dict['report_data'][active_tab]['users_column_select'][i-1] 
                              if i-1 < len(datmg.json_data_dict['report_data'][active_tab]['users_column_select']) else '' 
                              for i in range(1, 16)}  # Updated for 15 comboboxes
    else:
        current_selections = {f"combo_box{i}": '' for i in range(1, 16)}  # Updated for 15 comboboxes
    
    comboboxes = {}

    selecting_label = ttk.Label(root, text="Select Your Custom Column Types", font=('Arial', 28))
    widmg.store_and_place(page['labels'], "selecting_label", selecting_label, relx=0.10, rely=0.03, relwidth=0.80, relheight=0.17)

    def update_combobox_options(name):
        """Updates options for each combobox while keeping current selections visible."""
        excluded_options = set(current_selections.values()) - {''}
        for combobox_name, combobox in comboboxes.items():
            available_options = [option for option in options if option not in excluded_options]
            current_value = current_selections[combobox_name]
            combobox['values'] = available_options
            combobox.set(current_value)  # Keep the currently selected option visible

    def selection_handler(event, name):
        """Handles updating the current selection and refreshing ComboBox options."""
        current_selections[name] = comboboxes[name].get()
        update_combobox_options(name)

    for i in range(1, 16):  # Updated to iterate over 15 comboboxes
        combobox_name = f"combo_box{i}"
        combobox = ttk.Combobox(root, values=options, font=('Arial', 16))
        comboboxes[combobox_name] = combobox
        if editing:
            combobox.set(current_selections[combobox_name])
        
        # Calculate relx and rely for a 3x5 grid
        relx_value = 0.06 + 0.18 * ((i - 1) % 5)  # Adjusted relx for 5 in a row
        rely_value = 0.25 + 0.18 * ((i - 1) // 5)  # Adjusted rely for 3 rows
        widmg.store_and_place(page['dropdown_menus'], combobox_name, combobox, relx=relx_value, rely=rely_value, relwidth=0.16, relheight=0.12)  # Adjusted relwidth for narrower dropdowns

        # Bind the function to update selections and adjust options
        combobox.bind('<<ComboboxSelected>>', lambda event, name=combobox_name: selection_handler(event, name))

    update_combobox_options(None)
    
    back_button = ttk.Button(root, text="Back", command=lambda: dropdown_back_bridge(widmg, datmg, root, editing=editing), style='Large.TButton')
    widmg.store_and_place(page['buttons'], "back_button", back_button, relx=0.20, rely=0.80, relwidth=0.25, relheight=0.15)

    start_report_button = ttk.Button(root, text="Start Report", command=lambda: start_report_fn_button(widmg, datmg, root, editing=editing), style='Large.TButton')
    widmg.store_and_place(page['buttons'], "start_report_button", start_report_button, relx=0.55, rely=0.80, relwidth=0.25, relheight=0.15)


def dropdown_back_bridge(widmg, datmg, root, editing=False):
    if editing:
        page = widmg.custom_column_selection_page_widgets
        widmg.hide_widget_grouping(page['dropdown_menus'])
        widmg.hide_widget_grouping(page['buttons'])
        widmg.hide_widget_grouping(page['labels'])
        display_main_report_screen(widmg, datmg, root, editing=editing)
    else:
        cust_columns_to_review_screen(widmg, datmg, root)

def start_edited_col_report(widmg, datmg, root):
    active_tab = datmg.json_data_dict['active_tab']
    page = widmg.custom_column_selection_page_widgets
    combobox_list = page['dropdown_menus']
    
    new_selections = []
    for key, widget_info in combobox_list.items():
        widget = widget_info['widget']
        value = widget.get()
        if value != "":
            new_selections.append(value)


    # Ensure at least one option is always selected
    if not new_selections:
        messagebox.showwarning("Selection Error", "You must select at least one option.")
        return

    # Identify removed columns
    original_selections = datmg.json_data_dict['report_data'][active_tab]['users_column_select']
    removed_columns = [col for col in original_selections if col not in new_selections]

    if removed_columns:
        warning_message = f"Are you sure you want to proceed? Column data for {', '.join(removed_columns)} will be deleted."
        if not messagebox.askokcancel("Warning", warning_message):
            return

    # Update users_column_select
    datmg.json_data_dict['report_data'][active_tab]['users_column_select'] = new_selections
    datmg.save_dict_to_file()

    joint_data = datmg.json_data_dict['report_data'][active_tab]['joint_data']
    # Update report_data
    for joint, data in joint_data.items():
        updated_data = {key: value for key, value in data.items() if key in new_selections}
        datmg.json_data_dict['report_data'][active_tab]['joint_data'][joint] = updated_data

def start_report_fn_button(widmg, datmg, root, editing=False):
    if editing:
        start_edited_col_report(widmg, datmg, root)
        display_main_report_screen(widmg, datmg, root, editing=editing)
    else:
        add_column_selections_to_meta(widmg, datmg, root)
        display_main_report_screen(widmg, datmg, root)
    
def cust_columns_to_review_screen(widmg, datmg, root):
    page = widmg.custom_column_selection_page_widgets
    widmg.hide_widget_grouping(page['dropdown_menus'])
    widmg.hide_widget_grouping(page['buttons'])
    widmg.hide_widget_grouping(page['labels'])
    show_review_screen(widmg, datmg, root, editing=False)

def add_column_selections_to_meta(widmg, datmg, root):
    active_tab = datmg.json_data_dict['active_tab']
    combobox_list = widmg.custom_column_selection_page_widgets['dropdown_menus']
    values = []
    for key, widget_info in combobox_list.items():
        widget = widget_info['widget']
        value = widget.get()
        if value != "":
            values.append(value)

    datmg.json_data_dict['report_data'][active_tab]['users_column_select'] = values
    datmg.save_dict_to_file()



def back_bridge_step(widmg, datmg, root):
    page = widmg.main_table_display_widgets
    widmg.hide_widget_grouping(page['back_buttons'])
    widmg.hide_widget_grouping(page['display_frames'])
    widmg.hide_widget_grouping(page['buttons'])
    widmg.hide_widget_grouping(page['labels'])

    datmg = DataManager()
    widmg = WidgetManager(root)
    datmg.set_widget_manager(widmg)
    widmg.set_data_manager(datmg)
    root.bind("<Configure>", widmg.schedule_resize_fonts)
    
    branch_select_screen(widmg, datmg, root)


def back_bridge_shtblslscn(widmg, datmg, root):
    page = widmg.main_table_display_widgets
    widmg.hide_widget_grouping(page['buttons'])

    widmg.magni_header.set("Are You Sure?")
    widmg.magni_content.set("Continuing will clear your current data. (JSON File will be saved.)")

    # Create and store "Yes" and "Cancel" buttons
    yes_back_button = ttk.Button(root, text="Yes", command=lambda: back_bridge_step(widmg, datmg, root), style='Large.TButton')
    widmg.store_and_place(page['back_buttons'], "yes_back_button", yes_back_button, relx=0.05, rely=0.85, relwidth=0.15, relheight=0.1)

    cancel_back_button = ttk.Button(root, text="Cancel", command=lambda: cancel_back_action(widmg, datmg, root), style='Large.TButton')
    widmg.store_and_place(page['back_buttons'], "cancel_back_button", cancel_back_button, relx=0.80, rely=0.85, relwidth=0.15, relheight=0.1)


def cancel_back_action(widmg, datmg, root):
    page = widmg.main_table_display_widgets
    widmg.magni_header.set("")
    widmg.magni_content.set("")

    widmg.hide_widget_grouping(page['back_buttons'])
    widmg.place_back_widgets(page['buttons'])
    widmg.first_entry_widget.focus_set()


def display_main_report_screen(widmg, datmg, root, editing=False):
    widmg.hide_widget_grouping(widmg.table_select_page_widgets['buttons'])
    widmg.hide_widget_grouping(widmg.table_select_page_widgets['labels'])
    widmg.hide_widget_grouping(widmg.custom_column_selection_page_widgets['dropdown_menus'])
    widmg.hide_widget_grouping(widmg.custom_column_selection_page_widgets['labels'])
    widmg.hide_widget_grouping(widmg.custom_column_selection_page_widgets['buttons'])

    page = widmg.main_table_display_widgets

    create_header_and_first_row(widmg, datmg, root)
    create_all_four_rows(widmg, datmg, root, editing=editing)
    create_row_count_box(widmg, datmg, root)
    create_magnifier_window(widmg, datmg, root)
    create_microfier_windows(widmg, datmg, root)
    create_mini_meta_display(widmg, datmg, root)
    create_tab_data_box(widmg, datmg, root)
  
    # nav_increment_button = ttk.Button(root, text="Next Joint", command=lambda: increment_joint_number(widmg, datmg, root), style='Large.TButton')
    # widmg.store_and_place(page['buttons'], "nav_increment_button", nav_increment_button, relx=0.76, rely=0.55, relwidth=0.20, relheight=0.15)
    # nav_decrement_button = ttk.Button(root, text="Prev Joint", command=lambda: decrement_joint_number(widmg, datmg, root), style='Large.TButton')
    # widmg.store_and_place(page['buttons'], "nav_decrement_button", nav_decrement_button, relx=0.55, rely=0.55, relwidth=0.2, relheight=0.15)
    

    back_button = ttk.Button(root, text="Start Over", command=lambda: back_bridge_shtblslscn(widmg, datmg, root), style='Large.TButton')
    widmg.store_and_place(page['buttons'], "back_button", back_button, relx=0.02, rely=0.80, relwidth=0.20, relheight=0.08)

    finalize_button = ttk.Button(root, text="Finalize", command=lambda: create_confirmation_window(widmg, datmg, root), style='Large.TButton')
    widmg.store_and_place(page['buttons'], "finalize_button", finalize_button, relx=0.78, rely=0.80, relwidth=0.2, relheight=0.08)


def create_tab_data_box(widmg, datmg, root):
    page = widmg.main_table_display_widgets

    tab_data_box_frame = tk.Frame(root, bd=5, relief=tk.GROOVE)
    widmg.store_and_place(page['display_frames'], "tab_data_box_frame", tab_data_box_frame, relx=0.02, rely=0.02, relwidth=0.08, relheigh=0.08)
    tab_data_box_frame.bind("<MouseWheel>", lambda event: on_mouse_wheel(event, widmg, datmg, root))
    tab_data_box_frame.bind("<Up>", lambda event: decrement_joint_number(widmg, datmg, root))
    tab_data_box_frame.bind("<Down>", lambda event: increment_joint_number(widmg, datmg, root))

    tab_data_label = tk.Label(tab_data_box_frame, textvariable=widmg.tab_data_header, font=('Arial', 20, 'bold'))
    widmg.store_and_place(page['labels'], "tab_data_label", tab_data_label, font_changer=True, relx=0.01, rely=0.01, relwidth=0.99, relheight=0.99)
    tab_data_label.bind("<MouseWheel>", lambda event: on_mouse_wheel(event, widmg, datmg, root))
    tab_data_label.bind("<Up>", lambda event: decrement_joint_number(widmg, datmg, root))
    tab_data_label.bind("<Down>", lambda event: increment_joint_number(widmg, datmg, root))

def create_row_count_box(widmg, datmg, root):
    page = widmg.main_table_display_widgets

    row_count_box_frame = tk.Frame(root, bd=5, relief=tk.GROOVE)
    widmg.store_and_place(page['display_frames'], "row_count_box_frame", row_count_box_frame, relx=.02, rely=0.175, relwidth=0.13, relheigh=0.20)
    row_count_box_frame.bind("<MouseWheel>", lambda event: on_mouse_wheel(event, widmg, datmg, root))
    row_count_box_frame.bind("<Up>", lambda event: decrement_joint_number(widmg, datmg, root))
    row_count_box_frame.bind("<Down>", lambda event: increment_joint_number(widmg, datmg, root))


    joint_number_label = tk.Label(row_count_box_frame, text="JOINT #", font=('Arial', 12, 'bold italic'), padx=1, pady=1)
    widmg.store_and_place(page['labels'], "joint_number_label", joint_number_label, font_changer=True, relx=0.01, rely=0.01, relwidth=0.99, relheight=0.25)
    joint_number_label.bind("<MouseWheel>", lambda event: on_mouse_wheel(event, widmg, datmg, root))
    joint_number_label.bind("<Up>", lambda event: decrement_joint_number(widmg, datmg, root))
    joint_number_label.bind("<Down>", lambda event: increment_joint_number(widmg, datmg, root))
    current_joint_number_label = tk.Label(row_count_box_frame, textvariable=widmg.current_joint_number, font=('Arial', 40, 'bold'))
    widmg.store_and_place(page['labels'], "current_joint_number_label", current_joint_number_label, font_changer=True, relx=0.01, rely=0.26, relwidth=0.99, relheight=0.74)
    current_joint_number_label.bind("<MouseWheel>", lambda event: on_mouse_wheel(event, widmg, datmg, root))
    current_joint_number_label.bind("<Up>", lambda event: decrement_joint_number(widmg, datmg, root))
    current_joint_number_label.bind("<Down>", lambda event: increment_joint_number(widmg, datmg, root))

def create_header_and_first_row(widmg, datmg, root):
    page = widmg.main_table_display_widgets
    active_tab = datmg.json_data_dict['active_tab']
    joint_1_data = datmg.json_data_dict['report_data'][active_tab]['joint_data'].get('Joint_1', None)


    def adjust_row_height(event):
        # Calculate half the height of the frame
        new_height = event.height // 2
        # Adjust the minsize of the rows
        col_header_display_frame.grid_rowconfigure(0, minsize=new_height - 4)
        col_header_display_frame.grid_rowconfigure(1, minsize=new_height - 4)


    col_header_display_frame = tk.Frame(root, bd=4, relief=tk.GROOVE)
    widmg.store_and_place(page['display_frames'], "col_header_display_frame", col_header_display_frame, relx=0.15, rely=0.175, relwidth=0.83, relheight=0.20)
    col_header_display_frame.bind("<Configure>", adjust_row_height)
    col_header_display_frame.bind("<MouseWheel>", lambda event: on_mouse_wheel(event, widmg, datmg, root))
    col_header_display_frame.bind("<Up>", lambda event: decrement_joint_number(widmg, datmg, root))
    col_header_display_frame.bind("<Down>", lambda event: increment_joint_number(widmg, datmg, root))

    headers_list = datmg.json_data_dict['report_data'][active_tab]['users_column_select']
    total_headers = len(headers_list)
    grid_size = 15 // total_headers

    for idx, header in enumerate(headers_list):
        lbl = tk.Label(col_header_display_frame, text=header, bd=1, relief="solid", font=('Arial', 12, 'bold'), padx=1, pady=1, wraplength=((root_width * 0.83) / total_headers) - ((root_width * 0.83) / total_headers) * 0.04)
        widmg.store_and_grid(page['header_col_labels'], f"{idx}_{header}", lbl, font_changer=True, row=0, column=idx, sticky='nsew', padx=1, pady=1)
        lbl.bind("<MouseWheel>", lambda event: on_mouse_wheel(event, widmg, datmg, root))
        lbl.bind("<Up>", lambda event: decrement_joint_number(widmg, datmg, root))
        lbl.bind("<Down>", lambda event: increment_joint_number(widmg, datmg, root))
        col_header_display_frame.columnconfigure(idx, weight=grid_size)
    for idx, header in enumerate(headers_list):
        entry = tk.Entry(col_header_display_frame, bd=1, relief="solid", font=('Arial', 14, 'italic'))
        if joint_1_data and header in joint_1_data:
            entry.insert(0, joint_1_data[header])
        widmg.store_and_grid(page['col_entry_labels'], f"{idx}_{header}", entry, font_changer=True, row=1, column=idx, rowspan=2, sticky='nsew', padx=1, pady=1)
        entry.bind("<FocusIn>", lambda event: widmg.on_focus(event, root))  # Correct: Set focus info when the entry gets focus
        entry.bind("<FocusOut>", widmg.on_focus_out)
        entry.bind("<KeyRelease>", widmg.on_key_release)
        entry.bind("<Button-1>", lambda event: widmg.on_focus(event, root))  # Correct: Also set focus info when clicked
        entry.bind("<MouseWheel>", lambda event: on_mouse_wheel(event, widmg, datmg, root))
        entry.bind("<Up>", lambda event: decrement_joint_number(widmg, datmg, root))
        entry.bind("<Down>", lambda event: increment_joint_number(widmg, datmg, root))
        entry.bind("<Return>", lambda event: increment_joint_number(widmg, datmg, root))

        if idx == 0:
            widmg.first_entry_widget = entry
    widmg.first_entry_widget.focus_set()


def create_all_four_rows(widmg, datmg, root, editing=False):
    page = widmg.main_table_display_widgets
    active_tab = datmg.json_data_dict['active_tab']
    headers_list = datmg.json_data_dict['report_data'][active_tab]['users_column_select']
    total_headers = len(headers_list)

    def create_row_frame(rel_y, row_var_list, initial_value):
        def adjust_row_height(event):
            new_height = event.height
            row_frame.config(height=new_height)

        def adjust_label_sizes(event):
            new_width = event.width
            label_width = new_width // total_headers
            for idx, header in enumerate(headers_list):
                key = f"{header}_{idx}"
                if key in page['display_frames'][row_var_list]:
                    label_info = page['display_frames'][row_var_list][key]
                    label_widget = label_info['widget']
                    label_widget.config(width=label_width, wraplength=int(label_width * 0.98))

        row_frame = tk.Frame(root, bd=1, relief=tk.GROOVE)
        widmg.store_and_place(page['display_frames'], row_var_list, row_frame, relx=0.15, rely=rel_y, relwidth=0.83, relheight=0.07)
        row_frame.bind("<Configure>", adjust_row_height)
        row_frame.bind("<Configure>", adjust_label_sizes)
        row_frame.bind("<MouseWheel>", lambda event: on_mouse_wheel(event, widmg, datmg, root))
        row_frame.bind("<Up>", lambda event: decrement_joint_number(widmg, datmg, root))
        row_frame.bind("<Down>", lambda event: increment_joint_number(widmg, datmg, root))

        label_width = int((0.83 * root.winfo_width()) // total_headers)

        for idx, header in enumerate(headers_list):
            string_var = tk.StringVar(value=initial_value)
            getattr(widmg, row_var_list).append(string_var)

            row_label = tk.Label(row_frame, textvariable=string_var, bd=1, relief="solid", font=('Arial', 12, 'normal'))
            row_label.config(width=label_width, wraplength=int(label_width * 0.98))
            widmg.store_and_place(page['display_frames'][row_var_list], f"{header}_{idx}", row_label, font_changer=True, relx=idx/total_headers, rely=0, relwidth=1/total_headers, relheight=1)
            row_label.bind("<MouseWheel>", lambda event: on_mouse_wheel(event, widmg, datmg, root))
            row_label.bind("<Up>", lambda event: decrement_joint_number(widmg, datmg, root))
            row_label.bind("<Down>", lambda event: increment_joint_number(widmg, datmg, root))

    def update_row_frame(rel_y, row_var_list):
        row_frame_info = page['display_frames'][row_var_list]
        row_frame = row_frame_info['widget']
        row_frame.place(**row_frame_info['params'])

        label_width = int(0.83 * root.winfo_width() // total_headers)

        current_headers = {f"{header}_{idx}" for idx, header in enumerate(headers_list)}
        existing_headers = set(row_frame_info.keys()) - {'widget', 'params', 'visible'}

        for key in existing_headers - current_headers:
            widget_info = page['display_frames'][row_var_list].pop(key)
            widget_info['widget'].destroy()

            widget_dict_id = id(row_frame_info)
            if widget_dict_id in widmg.font_widgets and key in widmg.font_widgets[widget_dict_id]:
                del widmg.font_widgets[widget_dict_id][key]

            idx = int(key.split('_')[-1])
            if idx < len(getattr(widmg, row_var_list)):
                del getattr(widmg, row_var_list)[idx]

        new_string_var_list = []
        for idx, header in enumerate(headers_list):
            key = f"{header}_{idx}"
            if idx < len(getattr(widmg, row_var_list)):
                new_string_var_list.append(getattr(widmg, row_var_list)[idx])
            else:
                new_string_var_list.append(tk.StringVar())

        setattr(widmg, row_var_list, new_string_var_list)

        for idx, header in enumerate(headers_list):
            key = f"{header}_{idx}"
            string_var = getattr(widmg, row_var_list)[idx]

            if key in row_frame_info:
                row_label = page['display_frames'][row_var_list][key]['widget']
                row_label.config(textvariable=string_var, width=label_width, wraplength=int(label_width * 0.98))
                widmg.store_and_place(page['display_frames'][row_var_list], key, row_label, font_changer=True, relx=idx/total_headers, rely=0, relwidth=1/total_headers, relheight=1)
                row_label.bind("<MouseWheel>", lambda event: on_mouse_wheel(event, widmg, datmg, root))
                row_label.bind("<Up>", lambda event: decrement_joint_number(widmg, datmg, root))
                row_label.bind("<Down>", lambda event: increment_joint_number(widmg, datmg, root))
            else:
                row_label = tk.Label(row_frame, textvariable=string_var, bd=1, relief="solid", font=('Arial', 12, 'normal'))
                row_label.config(width=label_width, wraplength=int(label_width * 0.98))
                widmg.store_and_place(page['display_frames'][row_var_list], key, row_label, font_changer=True, relx=idx/total_headers, rely=0, relwidth=1/total_headers, relheight=1)
                row_label.bind("<MouseWheel>", lambda event: on_mouse_wheel(event, widmg, datmg, root))
                row_label.bind("<Up>", lambda event: decrement_joint_number(widmg, datmg, root))
                row_label.bind("<Down>", lambda event: increment_joint_number(widmg, datmg, root))


    if editing:
        update_row_frame(0.03, 'prev_sec_row_label_list')
        update_row_frame(0.10, 'prev_first_row_label_list')
        update_row_frame(0.375, 'next_first_row_label_list')
        update_row_frame(0.445, 'next_second_row_label_list')
    else:
        create_row_frame(0.03, 'prev_sec_row_label_list', "-")
        create_row_frame(0.10, 'prev_first_row_label_list', "-")
        create_row_frame(0.375, 'next_first_row_label_list', "")
        create_row_frame(0.445, 'next_second_row_label_list', "")

def create_magnifier_window(widmg, datmg, root):
    page = widmg.main_table_display_widgets
    magnifier_disp_frame = tk.Frame(root, bd=5, relief=tk.GROOVE)
    widmg.store_and_place(page['display_frames'], "magnifier_disp_frame", magnifier_disp_frame, relx=0.06, rely=0.52, relwidth=0.90, relheight=0.24)
    magnifier_disp_frame.bind("<MouseWheel>", lambda event: on_mouse_wheel(event, widmg, datmg, root))

    magni_header_label = tk.Label(magnifier_disp_frame, textvariable=widmg.magni_header, font=("Arial", 45, 'bold'), wraplength=388)
    widmg.store_and_place(page['labels'], "magni_header_label", magni_header_label, font_changer=True, relx=0.01, rely=0.01, relwidth=0.40, relheight=0.95, anchor='nw')
    magni_header_label.bind("<MouseWheel>", lambda event: on_mouse_wheel(event, widmg, datmg, root))

    # Create the green border frame for the magni_content_label
    magni_content_border_frame = tk.Frame(magnifier_disp_frame, background='green', bd=0)
    widmg.store_and_place(page['display_frames'], "magni_content_border_frame", magni_content_border_frame, relx=0.42, rely=0.01, relwidth=0.58, relheight=0.95, anchor='nw')

    # Place the actual magni_content_label inside this green border frame
    magni_content_label = tk.Label(magni_content_border_frame, textvariable=widmg.magni_content, font=("Helvetica", 38, 'italic'), wraplength=552)
    widmg.store_and_place(page['labels'], "magni_content_label", magni_content_label, font_changer=True, relx=0.01, rely=0.01, relwidth=0.98, relheight=0.98, anchor='nw')
    magni_content_label.bind("<MouseWheel>", lambda event: on_mouse_wheel(event, widmg, datmg, root))


def adjust_magnifier_fonts(widmg, root):
    def adjust_font(label, text_variable, base_font_family, base_font_size, base_font_weight):
        text = text_variable.get()
        label_height = int(label.winfo_height())
        label_width = int(label.winfo_width())
        wrap_length = int(label_width * 0.98)

        # Create a test label with similar configuration
        test_label = tk.Label(label.master, text=text, bd=1, relief="solid", font=(base_font_family, base_font_size, base_font_weight))
        test_label.config(wraplength=wrap_length)
        test_label.update_idletasks()

        # Adjust font size to fit text within the label
        while (test_label.winfo_reqwidth() > wrap_length or test_label.winfo_reqheight() > label_height) and base_font_size > 5:
            base_font_size -= 1
            test_label.config(font=(base_font_family, base_font_size, base_font_weight))
            test_label.update_idletasks()

        # Destroy the test label after adjustments
        test_label.destroy()

        # Apply the adjusted font size to the actual label
        label.config(font=(base_font_family, base_font_size, base_font_weight))

    # Adjust font for magni_header_label
    magni_header_label = widmg.main_table_display_widgets['labels']['magni_header_label']['widget']
    adjust_font(magni_header_label, widmg.magni_header, "Arial", 45, 'bold')

    # Adjust font for magni_content_label
    magni_content_label = widmg.main_table_display_widgets['labels']['magni_content_label']['widget']
    adjust_font(magni_content_label, widmg.magni_content, "Helvetica", 38, 'italic')



def create_mini_meta_display(widmg, datmg, root):
    page = widmg.main_table_display_widgets
    mini_meta_disp_frame = tk.Frame(root, bd=5, relief=tk.GROOVE)
    widmg.store_and_place(page['display_frames'], "mini_meta_disp_frame", mini_meta_disp_frame, relx=0.5, rely=0.86, relwidth=0.55, relheight=0.20, anchor='center')
    mini_meta_disp_frame.bind("<MouseWheel>", lambda event: on_mouse_wheel(event, widmg, datmg, root))
    mini_meta_disp_frame.bind("<Up>", lambda event: decrement_joint_number(widmg, datmg, root))
    mini_meta_disp_frame.bind("<Down>", lambda event: increment_joint_number(widmg, datmg, root))

    # Metadata and labels
    metadata = datmg.json_data_dict['report_user_metadata']
    header_labels = {
        'operator_entry': 'Operator: ',
        'date_entry': 'Date: ',
        'contractor_entry': 'Contractor: ',
        'invoice_entry': 'Invoice: ',
        'location_entry': 'Location: ',
        'inspected_by_entry': 'Inspected By: ',
        'inspection_type_selection': 'Inspection Type: ',
        'connection_size_selection': 'Connection Size: ',
        'connection_type_selection': 'Connection Type: '
    }

    # Manually calculated positions
    label_positions = [
        ('operator_entry', 0.0, 0.0, 0.17),  # (key, x position header, y position, width)
        ('date_entry', 0.5, 0.0, 0.20),
        ('contractor_entry', 0.0, 0.15, 0.17),
        ('invoice_entry', 0.5, 0.15, 0.20),
        ('location_entry', 0.0, 0.30, 0.17),
        ('inspected_by_entry', 0.5, 0.30, 0.20),
        ('inspection_type_selection', 0.0, 0.45, 0.235),
        ('connection_size_selection', 0.0, 0.85, 0.24),
        ('connection_type_selection', 0.5, 0.85, 0.24)
    ]

    for key, relx, rely, relwidth in label_positions:
        if key == 'inspection_type_selection':
            # Create and store metadata labels
            header_text = header_labels[key]
            header_label = tk.Label(mini_meta_disp_frame, text=header_text, bd=1, relief="solid", font=('Arial', 12, 'bold'), anchor='w')
            widmg.store_and_place(page['labels'], f"{key}_header_label", header_label, font_changer=True, relx=relx, rely=rely, relwidth=relwidth, relheight=0.4, anchor='nw')
            
            # Construct the value text
            value_text = metadata.get(key, '')
            if metadata.get('inspection_type_addodid', False):
                value_text += " Actual OD's, ID's, and Tong Space"
            if 'inspection_type_additional' in metadata:
                value_text += f" {metadata['inspection_type_additional']}"

            value_label = tk.Label(mini_meta_disp_frame, text=value_text, bd=1, relief="solid", font=('Helvetica', 11, 'italic'), anchor='center', wraplength=425)
            widmg.store_and_place(page['labels'], f"{key}_value_label", value_label, font_changer=True, relx=relx + relwidth, rely=rely, relwidth=float(1 - relwidth), relheight=0.4)
        elif key == 'connection_size_selection':
            header_text = header_labels[key]
            header_label = tk.Label(mini_meta_disp_frame, text=header_text, bd=1, relief="solid", font=('Arial', 12, 'bold'), anchor='w')
            widmg.store_and_place(page['labels'], f"{key}_header_label", header_label, font_changer=True, relx=relx, rely=rely, relwidth=relwidth, relheight=0.15, anchor='nw')
            value_to_get = metadata.get(key, '') + " " + metadata.get('grade_info', '')
            value_label = tk.Label(mini_meta_disp_frame, text=value_to_get, bd=1, relief="solid", font=('Helvetica', 11, 'italic'), anchor='w')
            widmg.store_and_place(page['labels'], f"{key}_value_label", value_label, font_changer=True, relx=relx + relwidth, rely=rely, relwidth=float(0.5 - relwidth), relheight=0.15)

        else:
            header_text = header_labels[key]
            header_label = tk.Label(mini_meta_disp_frame, text=header_text, bd=1, relief="solid", font=('Arial', 12, 'bold'), anchor='w')
            widmg.store_and_place(page['labels'], f"{key}_header_label", header_label, font_changer=True, relx=relx, rely=rely, relwidth=relwidth, relheight=0.15, anchor='nw')
            value_label = tk.Label(mini_meta_disp_frame, text=metadata.get(key, ''), bd=1, relief="solid", font=('Helvetica', 11, 'italic'), anchor='w')
            widmg.store_and_place(page['labels'], f"{key}_value_label", value_label, font_changer=True, relx=relx + relwidth, rely=rely, relwidth=float(0.5 - relwidth), relheight=0.15)

    edit_metadata_btn = ttk.Button(root, text="Edit Metadata", command=lambda: edit_metadata_action_bridge(widmg, datmg, root), style='Small.TButton')
    widmg.store_and_place(page['buttons'], "edit_metadata_btn", edit_metadata_btn, relx=0.05, rely=0.90, relwidth=0.15, relheight=0.05)

    edit_columns_btn = ttk.Button(root, text="Edit Columns", command=lambda: edit_columns_bridge(widmg, datmg, root), style='Small.TButton')
    widmg.store_and_place(page['buttons'], "edit_columns_btn", edit_columns_btn, relx=0.81, rely=0.90, relwidth=0.15, relheight=0.05)



def edit_columns_bridge(widmg, datmg, root):
    page = widmg.main_table_display_widgets
    widmg.hide_widget_grouping(page['buttons'])
    widmg.hide_widget_grouping(page['labels'])
    widmg.hide_widget_grouping(page['display_frames'])

    add_dropdowns(widmg, datmg, root, editing=True)

def edit_metadata_action_bridge(widmg, datmg, root):
    page = widmg.main_table_display_widgets

    widmg.hide_widget_grouping(page['buttons'])
    widmg.hide_widget_grouping(page['labels'])
    widmg.hide_widget_grouping(page['display_frames'])

    if 'grade_info' in datmg.json_data_dict['report_user_metadata']:
        widmg.grade_info_tube.set(datmg.json_data_dict['report_user_metadata']['grade_info'])
    
    create_report_metadata_input_widgets(widmg, datmg, root, editing=True)

def update_microfier_windows(widmg, datmg, root):
    current_value = int(widmg.current_joint_number.get())

    widmg.micro_plus_one_stvar.set(f"{current_value + 1}")
    widmg.micro_plus_two_stvar.set(f"{current_value + 2}")

    if current_value == 1:
        widmg.micro_negative_two_stvar.set("-")
        widmg.micro_negative_one_stvar.set("-")

    elif current_value == 2:
        widmg.micro_negative_two_stvar.set("-")
        widmg.micro_negative_one_stvar.set(f"{current_value - 1}")
    else:
        widmg.micro_negative_two_stvar.set(f"{current_value - 2}")
        widmg.micro_negative_one_stvar.set(f"{current_value - 1}")


def create_microfier_windows(widmg, datmg, root):
    page = widmg.main_table_display_widgets
    microfier_negative_two_frame = tk.Frame(root, bd=1, relief=tk.GROOVE)
    widmg.store_and_place(page['display_frames'], "microfier_negative_two_frame", microfier_negative_two_frame, relx=0.11, rely=0.03, relwidth=0.04, relheight=0.07)
    micro_neg_two_label = tk.Label(microfier_negative_two_frame, textvariable=widmg.micro_negative_two_stvar, bd=1, font=('Arial', 12, 'bold'))
    widmg.store_and_place(page['labels'], "micro_neg_two_label", micro_neg_two_label, relx=0.01, rely=0.01)

    microfier_negative_one_frame = tk.Frame(root, bd=1, relief=tk.GROOVE)
    widmg.store_and_place(page['display_frames'], "microfier_negative_one_frame", microfier_negative_one_frame, relx=0.11, rely=0.10, relwidth=0.04, relheight=0.07)
    micro_neg_one_label = tk.Label(microfier_negative_one_frame, textvariable=widmg.micro_negative_one_stvar, bd=1, font=('Arial', 12, 'bold'))
    widmg.store_and_place(page['labels'], "micro_neg_one_label", micro_neg_one_label, relx=0.01, rely=0.01)

    microfier_plus_one_frame = tk.Frame(root, bd=1, relief=tk.GROOVE)
    widmg.store_and_place(page['display_frames'], "microfier_plus_one_frame", microfier_plus_one_frame, relx=0.11, rely=0.375, relwidth=0.04, relheight=0.07)
    micro_plus_one_label = tk.Label(microfier_plus_one_frame, textvariable=widmg.micro_plus_one_stvar, bd=1, font=('Arial', 12, 'bold'))
    widmg.store_and_place(page['labels'], "micro_plus_one_label", micro_plus_one_label, relx=0.01, rely=0.01)

    microfier_plus_two_frame = tk.Frame(root, bd=1, relief=tk.GROOVE)
    widmg.store_and_place(page['display_frames'], "microfier_plus_two_frame", microfier_plus_two_frame, relx=0.11, rely=0.445, relwidth=0.04, relheight=0.07)
    micro_plus_two_label = tk.Label(microfier_plus_two_frame, textvariable=widmg.micro_plus_two_stvar, bd=1, font=('Arial', 12, 'bold'))
    widmg.store_and_place(page['labels'], "micro_plus_two_label", micro_plus_two_label, relx=0.01, rely=0.01)


def on_mouse_wheel(event, widmg, datmg, root):
    if event.delta > 0:
        decrement_joint_number(widmg, datmg, root)  # Scroll up to decrement
    else:
        increment_joint_number(widmg, datmg, root)  # Scroll down to increment

        
def increment_joint_number(widmg, datmg, root):
    current_value = int(widmg.current_joint_number.get())
    save_current_row_data(widmg, datmg, root)
    if current_value < 1000:
        widmg.current_joint_number.set(str(current_value + 1))
    update_microfier_windows(widmg, datmg, root)
    load_row_data(widmg, datmg, root)
    datmg.save_dict_to_file()
    update_all_row_cells(widmg, datmg, root)
    adjust_magnifier_fonts(widmg, root)
    widmg.update_magnifier(root)



def decrement_joint_number(widmg, datmg, root):
    current_value = int(widmg.current_joint_number.get())
    save_current_row_data(widmg, datmg, root)
    if current_value > 1:
        widmg.current_joint_number.set(str(current_value -1))
    load_row_data(widmg, datmg, root)
    update_microfier_windows(widmg, datmg, root)
    datmg.save_dict_to_file()
    update_all_row_cells(widmg, datmg, root)
    adjust_magnifier_fonts(widmg, root)
    widmg.update_magnifier(root)




### DONT KNOW IF THIS IS A DUPLICATE IN ERROR OR IF THIS SOMEHOW HAS AN ACTUAL PURPOSE

# def update_all_row_cells(widmg, datmg, root, editing=False):
#     report_type = datmg.json_data_dict['report_user_metadata']['report_type']
#     col_selects = datmg.json_data_dict['report_user_metadata']['users_column_select']
#     current_value = int(widmg.current_joint_number.get())
#     headers_list = datmg.json_data_dict['report_user_metadata']['users_column_select']
#     total_headers = len(headers_list)

#     def adjust_font_to_fit_label(up_lab_widget, fontchange, text):
#         base_font_family, base_font_size, base_font_weight = fontchange
#         label_width = int((0.83 * root.winfo_width()) // total_headers)
#         wrap_length = label_width * 0.98
#         label_height = int(0.07 * root.winfo_height())

#         # Create a test label with similar configuration
#         test_label = tk.Label(up_lab_widget.master, text=text, bd=1, relief="solid", font=(base_font_family, base_font_size, base_font_weight))
#         test_label.config(wraplength=wrap_length, width=label_width)
#         test_label.update_idletasks()

#         while (test_label.winfo_reqwidth() > wrap_length or test_label.winfo_reqheight() > label_height) and base_font_size > 5:
#             base_font_size -= 1
#             test_label.config(font=(base_font_family, base_font_size, base_font_weight))
#             test_label.update_idletasks()

#         if base_font_size == 5 and (test_label.winfo_reqwidth() > wrap_length or test_label.winfo_reqheight() > label_height):
#             test_label.destroy()
#             return (base_font_family, base_font_size, base_font_weight), text[:10] + "..."
#         else:
#             test_label.destroy()
#             return (base_font_family, base_font_size, base_font_weight), text



def update_all_row_cells(widmg, datmg, root, editing=False):
    report_type = datmg.json_data_dict['report_type']
    active_tab = datmg.json_data_dict['active_tab']
    col_selects = datmg.json_data_dict['report_data'][active_tab]['users_column_select']
    headers_list = datmg.json_data_dict['report_data'][active_tab]['users_column_select']
    total_headers = len(headers_list)
    current_value = int(widmg.current_joint_number.get())
    joint_data = datmg.json_data_dict['report_data'][active_tab]['joint_data']

    def adjust_font_to_fit_label(up_lab_widget, fontchange, text):
        base_font_family, base_font_size, base_font_weight = fontchange
        label_width = int((0.83 * root.winfo_width()) // total_headers)
        wrap_length = int(label_width * 0.98)
        label_height = int(0.07 * root.winfo_height())

        # Create a test label with similar configuration
        test_label = tk.Label(up_lab_widget.master, text=text, bd=1, relief="solid", font=(base_font_family, base_font_size, base_font_weight))
        test_label.config(wraplength=wrap_length)
        test_label.update_idletasks()

        # Adjust font size to fit text within the label
        while (test_label.winfo_reqwidth() > wrap_length or test_label.winfo_reqheight() > label_height) and base_font_size > 6:
            base_font_size -= 1
            test_label.config(font=(base_font_family, base_font_size, base_font_weight))
            test_label.update_idletasks()

        # Destroy the test label after adjustments
        test_label.destroy()

        # Return the adjusted font without truncating the text
        return (base_font_family, base_font_size, base_font_weight), text

    def update_cells(joint_key_offset, row_var_list, set_default=False):
        active_tab = datmg.json_data_dict['active_tab']
        joint_key = f"Joint_{current_value + joint_key_offset}"
        cells = getattr(widmg, row_var_list)
        row_dict = widmg.main_table_display_widgets['display_frames'][row_var_list]
        frame_key = f"{row_var_list}"

        if set_default:
            for idx, var in enumerate(cells):
                cells[idx].set("-")
            for idx, header in enumerate(col_selects):
                label_key = f"{header}_{idx}"
                up_lab_widget = widmg.get_label_widget(widmg, frame_key, label_key)
                default_font = ('TKDefaultFont', 12, 'normal')
                up_lab_widget.config(bg='systemButtonFace', fg='black', font=default_font)
                widget_dict_id = id(row_dict)
                if widget_dict_id in widmg.font_widgets and label_key in widmg.font_widgets[widget_dict_id]:
                    widmg.font_widgets[widget_dict_id][label_key]['initial_font'] = default_font
        else:
            for idx, header in enumerate(col_selects):
                label_key = f"{header}_{idx}"
                if joint_key in joint_data and header in joint_data[joint_key]:
                    header_value = datmg.json_data_dict['report_data'][active_tab]['joint_data'][joint_key][header]
                    cells[idx].set(header_value)



                    if header == 'UT':
                        txt_color, bg_color, fontchange = validate_ut(header_value, header, datmg)
                        up_lab_widget = widmg.get_label_widget(widmg, frame_key, label_key)
                        fontchange, adjusted_text = adjust_font_to_fit_label(up_lab_widget, fontchange, header_value)
                        cells[idx].set(adjusted_text)
                        up_lab_widget.config(bg=bg_color, fg=txt_color, font=fontchange)
                        widget_dict_id = id(row_dict)
                        if widget_dict_id in widmg.font_widgets and label_key in widmg.font_widgets[widget_dict_id]:
                            widmg.font_widgets[widget_dict_id][label_key]['initial_font'] = fontchange

                    elif header == 'PIN':
                        if report_type == "Drill Pipe Inspection Report":
                            txt_color, bg_color, fontchange = validate_reface(header_value, header, datmg)
                            up_lab_widget = widmg.get_label_widget(widmg, frame_key, label_key)
                            fontchange, adjusted_text = adjust_font_to_fit_label(up_lab_widget, fontchange, header_value)
                            cells[idx].set(adjusted_text)
                            up_lab_widget.config(bg=bg_color, fg=txt_color, font=fontchange)
                            widget_dict_id = id(row_dict)
                            if widget_dict_id in widmg.font_widgets and label_key in widmg.font_widgets[widget_dict_id]:
                                widmg.font_widgets[widget_dict_id][label_key]['initial_font'] = fontchange
                        elif report_type == "Tubing/Casing Report":
                            up_lab_widget = widmg.get_label_widget(widmg, frame_key, label_key)
                            default_font = ('TKDefaultFont', 12, 'normal')
                            fontchange, adjusted_text = adjust_font_to_fit_label(up_lab_widget, default_font, header_value)
                            cells[idx].set(adjusted_text)
                            up_lab_widget.config(bg='systemButtonFace', fg='black', font=fontchange)
                            widget_dict_id = id(row_dict)
                            if widget_dict_id in widmg.font_widgets and label_key in widmg.font_widgets[widget_dict_id]:
                                widmg.font_widgets[widget_dict_id][label_key]['initial_font'] = fontchange

                    elif header == 'BOX':
                        if report_type == "Drill Pipe Inspection Report":
                            txt_color, bg_color, fontchange = validate_reface(header_value, header, datmg)
                            up_lab_widget = widmg.get_label_widget(widmg, frame_key, label_key)
                            fontchange, adjusted_text = adjust_font_to_fit_label(up_lab_widget, fontchange, header_value)
                            cells[idx].set(adjusted_text)
                            up_lab_widget.config(bg=bg_color, fg=txt_color, font=fontchange)
                            widget_dict_id = id(row_dict)
                            if widget_dict_id in widmg.font_widgets and label_key in widmg.font_widgets[widget_dict_id]:
                                widmg.font_widgets[widget_dict_id][label_key]['initial_font'] = fontchange
                        elif report_type == "Tubing/Casing Report":
                            up_lab_widget = widmg.get_label_widget(widmg, frame_key, label_key)
                            default_font = ('TKDefaultFont', 12, 'normal')
                            fontchange, adjusted_text = adjust_font_to_fit_label(up_lab_widget, default_font, header_value)
                            cells[idx].set(adjusted_text)
                            up_lab_widget.config(bg='systemButtonFace', fg='black', font=fontchange)
                            widget_dict_id = id(row_dict)
                            if widget_dict_id in widmg.font_widgets and label_key in widmg.font_widgets[widget_dict_id]:
                                widmg.font_widgets[widget_dict_id][label_key]['initial_font'] = fontchange

                    elif header == 'TUBE':
                        up_lab_widget = widmg.get_label_widget(widmg, frame_key, label_key)
                        default_font = ('TKDefaultFont', 12, 'normal')
                        fontchange, adjusted_text = adjust_font_to_fit_label(up_lab_widget, default_font, header_value)
                        cells[idx].set(adjusted_text)
                        up_lab_widget.config(bg='systemButtonFace', fg='black', font=fontchange)
                        widget_dict_id = id(row_dict)
                        if widget_dict_id in widmg.font_widgets and label_key in widmg.font_widgets[widget_dict_id]:
                            widmg.font_widgets[widget_dict_id][label_key]['initial_font'] = fontchange

                    elif header == 'SERIAL':
                        up_lab_widget = widmg.get_label_widget(widmg, frame_key, label_key)
                        default_font = ('TKDefaultFont', 12, 'normal')
                        fontchange, adjusted_text = adjust_font_to_fit_label(up_lab_widget, default_font, header_value)
                        cells[idx].set(adjusted_text)
                        up_lab_widget.config(bg='systemButtonFace', fg='black', font=fontchange)
                        widget_dict_id = id(row_dict)
                        if widget_dict_id in widmg.font_widgets and label_key in widmg.font_widgets[widget_dict_id]:
                            widmg.font_widgets[widget_dict_id][label_key]['initial_font'] = fontchange
                    else:
                        up_lab_widget = widmg.get_label_widget(widmg, frame_key, label_key)
                        default_font = ('TKDefaultFont', 12, 'normal')
                        fontchange, adjusted_text = adjust_font_to_fit_label(up_lab_widget, default_font, header_value)
                        cells[idx].set(adjusted_text)
                        up_lab_widget.config(bg='systemButtonFace', fg='black', font=fontchange)
                        widget_dict_id = id(row_dict)
                        if widget_dict_id in widmg.font_widgets and label_key in widmg.font_widgets[widget_dict_id]:
                            widmg.font_widgets[widget_dict_id][label_key]['initial_font'] = fontchange

    if current_value == 1:
        update_cells(-1, 'prev_first_row_label_list', set_default=True)
    else:
        update_cells(-1, 'prev_first_row_label_list')

    if current_value <= 2:
        update_cells(-2, 'prev_sec_row_label_list', set_default=True)
    else:
        update_cells(-2, 'prev_sec_row_label_list')

    update_cells(1, 'next_first_row_label_list')
    update_cells(2, 'next_second_row_label_list')






def validate_ut(header_value, header, datmg):
    active_tab = datmg.json_data_dict['active_tab']
    connection_size = datmg.json_data_dict['report_user_metadata']['connection_size_selection']
    conn_type_select = datmg.json_data_dict['report_user_metadata']['connection_type_selection'] 
    report_type = datmg.json_data_dict['report_type']
    if report_type == 'Drill Pipe Inspection Report':
        nominal_wall = 1000 * (datmg.dp_conn_size_nom_rel_dict[connection_size])
    elif report_type == 'Tubing/Casing Report':
        nominal_wall = 1000 * (datmg.tube_conn_size_nom_rel_dict[connection_size][conn_type_select])

    if header_value != '':
        if len(header_value) != 3 or not header_value.isdigit() or (float(header_value) / nominal_wall) >= 1.15:
            return datmg.color_code_dict["Invalid Syntax"]

        elif header_value.isdigit() and len(header_value) == 3:
            perc_nom_wall = float(header_value) / nominal_wall
            if perc_nom_wall < 1.10 and perc_nom_wall > 0.80:
                return datmg.color_code_dict["Good Nominal Wall"]
            elif perc_nom_wall <= 0.80 and perc_nom_wall > 0.75:
                return datmg.color_code_dict["Class 2"]
            elif perc_nom_wall <= 0.75 and perc_nom_wall > 0.70:
                return datmg.color_code_dict["Class 3"]
            elif perc_nom_wall <= 0.70:
                return datmg.color_code_dict["Invalid Validation"]
            else:
                return datmg.color_code_dict["Default"]
        else:
            return datmg.color_code_dict["Default"]
    return datmg.color_code_dict["Default"]



def validate_reface(header_value, header, datmg):
    header_value = re.sub(r'(\d+)\s(\d+/\d+)', r'\1_\2', header_value)
    data_chunks = header_value.split()
    n = len(data_chunks)
    reface_kws = ["R1", "R2", "R3", "R4", "R"]

    conn_type = datmg.json_data_dict['report_user_metadata']['connection_type_selection']

    boxcl_min_bef = str(datmg.dp_conn_type_vals_dict[conn_type][4]).split(".")[0]
    boxcl_max_bef = str(datmg.dp_conn_type_vals_dict[conn_type][5]).split(".")[0]
    pincl_min_bef = str(datmg.dp_conn_type_vals_dict[conn_type][6]).split(".")[0]
    pincl_max_bef = str(datmg.dp_conn_type_vals_dict[conn_type][7]).split(".")[0]

    minbox_reface, maxbox_reface = datmg.dp_conn_type_vals_dict[conn_type][4:6]
    minpin_reface, maxpin_reface = datmg.dp_conn_type_vals_dict[conn_type][6:8]

    if conn_type != "NC-50":
        for chunk in data_chunks:
            if chunk in reface_kws:
                rf_index = data_chunks.index(chunk)
                if rf_index < n - 1:
                    nxt_idx = data_chunks[rf_index + 1]
                    if nxt_idx.isdigit() and len(nxt_idx) == 6:
                        if header == 'BOX':
                            rbcl = float(f"{boxcl_min_bef}.{nxt_idx[:3]}")
                            racl = float(f"{boxcl_max_bef}.{nxt_idx[3:]}") 
                            if (rbcl < minbox_reface or rbcl > maxbox_reface) or \
                            (racl < minbox_reface or racl > maxbox_reface):
                                return datmg.color_code_dict["Invalid Validation"] 
                        if header == 'PIN': 
                            rbcl = float(f"{pincl_min_bef}.{nxt_idx[:3]}")
                            racl = float(f"{pincl_max_bef}.{nxt_idx[3:]}")  
                            if (rbcl < minpin_reface or rbcl > maxpin_reface) or \
                            (racl < minpin_reface or racl > maxpin_reface):
                                return datmg.color_code_dict["Invalid Validation"]
                    elif nxt_idx.isdigit() and (len(nxt_idx) > 6 or len(nxt_idx) < 6):
                        return datmg.color_code_dict["Invalid Syntax"]
                        
    return datmg.color_code_dict["Default"]



def save_current_row_data(widmg, datmg, root):
    active_tab = datmg.json_data_dict['active_tab']
    col_selects = datmg.json_data_dict['report_data'][active_tab]['users_column_select']
    col_head_frame = widmg.main_table_display_widgets['display_frames']['col_header_display_frame']['widget']
    current_value = int(widmg.current_joint_number.get())
    joint_key = f"Joint_{current_value}"

    entry_data = {}
    for idx, header in enumerate(col_selects):
        entry_widget = col_head_frame.grid_slaves(row=1, column=idx)[0]
        entry_data[header] = entry_widget.get() 


    datmg.json_data_dict['report_data'][active_tab]['joint_data'][joint_key] = entry_data

def load_row_data(widmg, datmg, root):
    active_tab = datmg.json_data_dict['active_tab']
    col_selects = datmg.json_data_dict['report_data'][active_tab]['users_column_select']

    col_head_frame = widmg.main_table_display_widgets['display_frames']['col_header_display_frame']['widget']
    current_value = int(widmg.current_joint_number.get())
    joint_key = f"Joint_{current_value}"

    joint_data = datmg.json_data_dict['report_data'][active_tab]['joint_data']
    for idx, header in enumerate(col_selects):
        entry_widget = col_head_frame.grid_slaves(row=1, column=idx)[0]
        entry_widget.delete(0, tk.END)
        if joint_key in joint_data and header in joint_data[joint_key]:
            entry_widget.insert(0, datmg.json_data_dict['report_data'][active_tab]['joint_data'][joint_key][header])


def update_report_data(widmg, datmg, root):
    active_tab = datmg.json_data_dict['active_tab']
    joint_data = datmg.json_data_dict['report_data'][active_tab]['joint_data']

    max_joint_number = int(widmg.update_rows_entry_widget.get())
    keys_to_delete = [key for key in joint_data.keys() if int(key.split('_')[1]) > max_joint_number]
    for key in keys_to_delete:
        del joint_data[key]
    datmg.save_dict_to_file()

def create_confirmation_window(widmg, datmg, root):
    if hasattr(widmg, 'confirmation_window') and widmg.confirmation_window.winfo_exists():
        # If the window already exists, bring it to the front
        widmg.confirmation_window.lift()
        return

    # Create a new top-level window
    widmg.confirmation_window = tk.Toplevel(root)
    widmg.confirmation_window.title("Confirm Row Count")
    
    # Set a reasonable size for the window
    widmg.confirmation_window.geometry("485x275")

    counted_joints = find_highest_joint_num_for_export(widmg, datmg, root)

    label_widget = ttk.Label(widmg.confirmation_window, text="Confirm Row Count for Writing to Excel", font=('Arial', 12, "bold"))
    label_widget.place(relx=0.5, rely=0.15, relwidth=0.98, relheight=0.15, anchor='center')
    
    # Create an Entry widget, populate it with the provided value, and make it read-only
    widmg.update_rows_entry_widget = ttk.Entry(widmg.confirmation_window, font=('Arial', 16))
    widmg.update_rows_entry_widget.insert(0, counted_joints)
    widmg.update_rows_entry_widget.place(relx=0.5, rely=0.31, relwidth=0.50, relheight=0.12, anchor='center')
    
    # Create a button that for now does nothing when clicked
    # You will replace command=lambda: None with your actual function later
    confirm_button = ttk.Button(widmg.confirmation_window, text="Confirm", command=lambda: confirm_joints_button(widmg, datmg, root, widmg.confirmation_window))
    confirm_button.place(relx=0.33, rely=0.65, relwidth=0.25, relheight=0.15, anchor='center')

    cancel_button = ttk.Button(widmg.confirmation_window, text="Cancel", command=widmg.confirmation_window.destroy)
    cancel_button.place(relx=0.67, rely=0.65, relwidth=0.25, relheight=0.15, anchor='center')

    # Ensure the window is properly cleaned up when closed
    widmg.confirmation_window.protocol("WM_DELETE_WINDOW", lambda: on_confirmation_window_close(widmg))

def on_confirmation_window_close(widmg):
    widmg.confirmation_window.destroy()
    del widmg.confirmation_window


def check_and_update_dpnd_complete(datmg, report_type):
    datmg.excel_files_tct[report_type]['Completed?'] = 'Yes'

    # Now check the "Included?" and corresponding "Complete?" values to determine completeness
    report_complete = True  # Assume complete until proven otherwise

    # Iterate over each report type and check if conditions are met
    for report_type, values in datmg.excel_files_tct.items():
        if report_type != "REPORT_COMPLETE":  # Skip the "REPORT_COMPLETE" entry
            included = values["Included?"]
            completed = values["Completed?"]

            if included == "Yes" and completed != "Yes":  # If Included is Yes but Complete is not
                report_complete = False
                break  # No need to check further if one condition fails

    # Update the "REPORT_COMPLETE" status based on the check
    datmg.excel_files_tct["REPORT_COMPLETE"] = "Complete" if report_complete else "Incomplete"


def move_related_pdfs(excel_filename, pdf_abs_location, incomplete_reports_folder, folder_selected):
    # Extract Excel joint count
    excel_basename = os.path.basename(excel_filename)
    excel_joint_str = excel_basename.split('_')[-1]
    excel_total_count = int(excel_joint_str.replace('JTS.xlsx', ''))

    # Extract fed PDF joint count
    pdf_basename = os.path.basename(pdf_abs_location)
    fed_joint_str = pdf_basename.split('_')[-2]
    fed_joint_count = int(''.join([char for char in fed_joint_str if char.isdigit()])) 


    # Find PDFs in incomplete_reports_folder
    prelim_matching_pdfs = []
    allowed_variations = ['DP', 'PDPIR', 'HWDP', 'SUBS']
    for file in os.listdir(incomplete_reports_folder):
        if file.endswith('PDF-COPY.pdf'):
            parts_excel = excel_basename.split('_')[:-1]  # Ignore the joint count in Excel filename
            parts_pdf = file.split('_')[:-2]  # Ignore the joint count in the PDF filename
            if parts_excel[:2] == parts_pdf[:2] and parts_excel[3:] == parts_pdf[3:]:
                # Handle comparison for the third element (index 2) in both lists
                excel_third = parts_excel[2].split()  # Split by space
                pdf_third = parts_pdf[2].split()  # Split by space

                # Ensure the 'size' and 'Inch' portions match
                if excel_third[0] == pdf_third[0] and excel_third[1] == pdf_third[1]:
                    # Check if the variable portion ('DP', 'PDPIR', etc.) is one of the allowed variations
                    if excel_third[2] in allowed_variations and pdf_third[2] in allowed_variations:
                        # If everything matches, add to the preliminary list
                        pdf_joint_count_str = file.split('_')[-2]
                        pdf_joint_count = int(''.join([char for char in pdf_joint_count_str if char.isdigit()])) 
                        prelim_matching_pdfs.append((file, pdf_joint_count))

    # Verify joint counts
    matched_pdfs = []
    current_joint_sum = fed_joint_count
    for pdf_file, pdf_joint_count in prelim_matching_pdfs:
        current_joint_sum += pdf_joint_count
        matched_pdfs.append(pdf_file)
        if current_joint_sum == excel_total_count:
            break

    # If matched, move the PDFs
    if current_joint_sum == excel_total_count:
        for pdf_file in matched_pdfs:
            src = os.path.join(incomplete_reports_folder, pdf_file)
            dest = os.path.join(folder_selected, pdf_file)

            # Manually copy the file to new location
            with open(src, 'rb') as file:
                data = file.read()
            with open(dest, 'wb') as file:
                file.write(data)

            # Delete the original after copying
            os.remove(src)

    else:
        print("No matching PDFs found that meet the criteria.")




def delete_excel_from_incomplete_reports(datmg, excel_file, incomplete_reports_folder):
    excel_filename = os.path.basename(excel_file)
    old_excel_location = os.path.join(incomplete_reports_folder, excel_filename)
    if os.path.exists(old_excel_location):
        os.remove(old_excel_location)
    if os.path.exists(datmg.json_tct_filepath):
        os.remove(datmg.json_tct_filepath)


def update_new_report_metadata(datmg, widmg, typrep):
    # Step 1: Clear the 'report_data' section but keep the key
    widmg.tab_data_header.set(typrep)

    # Step 2: Update 'active_tab' based on the value of typrep
    if typrep == 'PDPIR':
        datmg.json_data_dict['active_tab'] = 'Prop Drill Pipe Inp Report'
    elif typrep == 'HWDP':
        datmg.json_data_dict['active_tab'] = 'Prop HWDP Inp Report'
    elif typrep == 'SUBS':
        datmg.json_data_dict['active_tab'] = 'Prop Subs Inp Report'
    else:
        raise ValueError("Invalid typrep value. Expected 'PDPIR', 'HWDP', or 'SUBS'.")


def hide_main_report_scrn_dp_tab_new_json(datmg, widmg, root):
    page = widmg.main_table_display_widgets

    widmg.hide_widget_grouping(page['buttons'])
    widmg.hide_widget_grouping(page['labels'])
    widmg.hide_widget_grouping(page['display_frames'])

    create_report_metadata_input_widgets(widmg, datmg, root, editing=False)

def open_continue_dp_report_window(widmg, datmg, root, workbook, excel_filepath, undo_filepath, undo_jsontct_filepath, pdf_filename, pdf_to_delete, report_type, incomplete_reports_folder, summary_data):
    # Load the Excel workbook and access the "Data Sheet" tab

    # Create a new top-level window (this is the pop-up)
    new_window = tk.Toplevel(root)
    new_window.title("Select Active Tabs")
    new_window.geometry("400x300")  # Adjust the window size to be larger for better spacing

    # Apply the 'arc' theme (if needed)
    style = ttk.Style(new_window)
    style.theme_use('arc')

    # Create a label at the top and center it
    label = ttk.Label(new_window, text="Select How To Continue Completing The Report", font=('Arial', 12))
    label.place(relx=0.5, rely=0.1, anchor="n")  # Center it horizontally and provide space above

    # Frame to hold the dynamic buttons and center them
    button_frame = ttk.Frame(new_window)
    button_frame.place(relx=0.5, rely=0.3, relwidth=0.98, relheight=0.33, anchor="n")  # Center the frame, with 80% width

    # List to store the dynamic buttons
    buttons = []

    # Helper function to create dynamic buttons based on the logic
    def create_button_if_needed(report_type):
        # Retrieve "Included?" and "Complete?" values from the dictionary
        included = datmg.excel_files_tct[report_type]["Included?"]
        completed = datmg.excel_files_tct[report_type]["Completed?"]
        
        # Check the logic for "Included?" and "Complete?"
        if included == 'Yes' and completed != 'Yes':
            button = ttk.Button(
                button_frame, text=f"Continue with {report_type}", width=25, 
                command=lambda: on_button_click(datmg, widmg, root, workbook, excel_filepath, report_type, new_window, pdf_to_delete, pdf_filename))
            buttons.append(button)  # Add the button to the list

    # Logic to determine which buttons to display
    create_button_if_needed("PDPIR")  # Checks "PDPIR"
    create_button_if_needed("HWDP")   # Checks "HWDP"
    create_button_if_needed("SUBS")   # Checks "SUBS"

    # Manually position the buttons using .place()
    if len(buttons) == 1:
        buttons[0].place(relx=0.5, rely=0.25, relheight=0.5, anchor="n", relwidth=0.6)  # Center the single button
    elif len(buttons) == 2:
        buttons[0].place(relx=0.3, rely=0.25, relheight=0.5, anchor="n", relwidth=0.4)
        buttons[1].place(relx=0.7, rely=0.25, relheight=0.5, anchor="n", relwidth=0.4)
    elif len(buttons) == 3:
        buttons[0].place(relx=0.2, rely=0.25, relheight=0.5, anchor="n", relwidth=0.25)
        buttons[1].place(relx=0.5, rely=0.25, relheight=0.5, anchor="n", relwidth=0.25)
        buttons[2].place(relx=0.8, rely=0.25, relheight=0.5, anchor="n", relwidth=0.25)

    def on_button_click(datmg, widmg, root, workbook, excel_filepath, report_type, new_window, pdf_to_delete, pdf_filename):
        datmg.editing_spec_tab = False
        if datmg.xel_file_path is None:
            datmg.xel_file_path = datmg.new_excel_fp

        datmg.save_tab_status_table_to_json(datmg.json_tct_filepath)
        
        excel_file = os.path.join(incomplete_reports_folder, os.path.basename(excel_filepath))
        workbook.save(excel_file)

        pdf_abs_location = os.path.join(incomplete_reports_folder, pdf_filename)
        generate_pdf_copy(summary_data, pdf_abs_location, widmg, datmg, root)
        
        if pdf_to_delete is not None:
            if os.path.exists(pdf_to_delete):
                os.remove(pdf_to_delete)
        

        update_new_report_metadata(datmg, widmg, report_type)
        new_window.destroy()
        hide_main_report_scrn_dp_tab_new_json(datmg, widmg, root)

    def close_to_edit_button_action(datmg, widmg, root, new_window, undo_filepath, undo_jsontct_filepath):
        datmg.excel_files_tct[report_type]["Completed?"] = "No"

        if datmg.xel_file_path is not None:
            os.rename(datmg.xel_file_path, undo_filepath)
            os.rename(datmg.json_tct_filepath, undo_jsontct_filepath)
            datmg.xel_file_path = undo_filepath
            datmg.json_tct_filepath = undo_jsontct_filepath
            datmg.save_tab_status_table_to_json(datmg.json_tct_filepath)

        # Close both windows
        new_window.destroy()

    def close_to_save_button_action(datmg, widmg, root, excel_filepath, new_window, pdf_to_delete):
        excel_filepath = os.path.join(incomplete_reports_folder, os.path.basename(excel_filepath))
        workbook.save(excel_filepath)
        datmg.save_tab_status_table_to_json(datmg.json_tct_filepath)

        pdf_abs_location = os.path.join(incomplete_reports_folder, pdf_filename)
        generate_pdf_copy(summary_data, pdf_abs_location, widmg, datmg, root)

        if pdf_to_delete is not None:
            if os.path.exists(pdf_to_delete):
                os.remove(pdf_to_delete)
        
        new_window.destroy()
        datmg.editing_spec_tab = False
        datmg.xel_file_path = None
        back_bridge_step(widmg, datmg, root)
        
    new_window.protocol("WM_DELETE_WINDOW", lambda: close_to_edit_button_action(datmg, widmg, root, new_window, undo_filepath, undo_jsontct_filepath))

    # Add a close button to close the pop-up window, place it near the bottom
    close_to_save_button = ttk.Button(new_window, text=f"Save {widmg.tab_data_header.get()} Report and Go To HOME Screen", command=lambda: close_to_save_button_action(datmg, widmg, root, excel_filepath, new_window, pdf_to_delete))    
    close_to_save_button.place(relx=0.5, rely=0.635, relwidth=0.75, relheight=0.15, anchor="n")  # Center it at the bottom

    # Add a close button to close the pop-up window, place it near the bottom
    close_to_edit_button = ttk.Button(new_window, text=f"Continue Editing {widmg.tab_data_header.get()} Report", command=lambda: close_to_edit_button_action(datmg, widmg, root, new_window, undo_filepath, undo_jsontct_filepath))    
    close_to_edit_button.place(relx=0.5, rely=0.80, relwidth=0.75, relheight=0.15, anchor="n")  # Center it at the bottom




def confirm_joints_button(widmg, datmg, root, confirmation_window):
    update_report_data(widmg, datmg, root)
    metadata = datmg.json_data_dict['report_user_metadata']
    report_data = datmg.json_data_dict["report_data"]

    # Folder location for incomplete reports
    incomplete_reports_folder = os.path.join(os.getcwd(), 'incomplete_reports')

    # Check if the folder exists, if not, create it
    if not os.path.exists(incomplete_reports_folder):
        os.makedirs(incomplete_reports_folder)

    datmg.keyword_tally_dict["Keyword Tallies"] = {}
    datmg.keyword_tally_dict["Joint Tallies"] = {}

    def write_report_data_all_nd_pdpir(datmg, report_data, workbook, active_tab):
        sheet = workbook[active_tab]
        for joint_num, joint_values in report_data.items():
            row_num = 9 + int(joint_num.split('_')[-1]) - 1
            process_for_write_report_nd_pdpir(joint_values, row_num, sheet, datmg)

    if metadata['branch'] == "ND":
        if metadata['report_type'] == 'Drill Pipe Inspection Report':
            workbook, undo_filepath, undo_jsontct_filepath, pdf_to_delete, pdffilename = create_modify_excel_document_nd_pdpir(incomplete_reports_folder, widmg, datmg, root)
            excel_filepath = datmg.xel_file_path if datmg.xel_file_path is not None else datmg.new_excel_fp
           
            if metadata['active_tab'] == 'Prop Drill Pipe Inp Report':
                report_type = "PDPIR"
                write_report_data_all_nd_pdpir(datmg, report_data, workbook, metadata['active_tab'])
                tot_joints = len(report_data)
                summary_data = generate_summary_entry_nd_dp(tot_joints, datmg.keyword_tally_dict, datmg)
                check_and_update_dpnd_complete(datmg, report_type)

                if datmg.excel_files_tct["REPORT_COMPLETE"] == "Incomplete":
                    open_continue_dp_report_window(widmg, datmg, root, workbook, excel_filepath, undo_filepath, undo_jsontct_filepath, pdffilename, pdf_to_delete, report_type, incomplete_reports_folder, summary_data)
                else:

                    folder_selected = filedialog.askdirectory(title='Select Folder to save Report Files')
                    if not folder_selected:
                        if datmg.xel_file_path is not None:
                            os.rename(datmg.xel_file_path, undo_filepath)
                            os.rename(datmg.json_tct_filepath, undo_jsontct_filepath)
                            datmg.xel_file_path = undo_filepath
                            datmg.json_tct_filepath = undo_jsontct_filepath
                            datmg.save_tab_status_table_to_json(datmg.json_tct_filepath)
                        messagebox.showerror("Error", "Please select folder to store excel report.")
                        confirmation_window.destroy()
                        return
                    excel_filepath = os.path.join(folder_selected, os.path.basename(excel_filepath))
                    pdf_abs_location = os.path.join(folder_selected, pdffilename)
                    generate_pdf_copy(summary_data, pdf_abs_location, widmg, datmg, root)
                    move_related_pdfs(excel_filepath, pdf_abs_location, incomplete_reports_folder, folder_selected)
                    delete_excel_from_incomplete_reports(datmg, excel_filepath, incomplete_reports_folder)
                    workbook.save(excel_filepath)
                    messagebox.showinfo("Success", f"{excel_filepath} Saved Successfully to {folder_selected}" )
                    datmg.editing_spec_tab = False
                    back_bridge_step(widmg, datmg, root)
                # write_summary_notes_nd_dp()


            elif metadata['active_tab'] == 'Prop HWDP Inp Report':
                report_type = "HWDP"
                write_report_data_all_nd_pdpir(datmg, report_data, workbook, metadata['active_tab'])
                tot_joints = len(report_data)
                summary_data = generate_summary_entry_nd_dp(tot_joints, datmg.keyword_tally_dict, datmg)
                check_and_update_dpnd_complete(datmg, report_type)


                if datmg.excel_files_tct["REPORT_COMPLETE"] == "Incomplete":
                    open_continue_dp_report_window(widmg, datmg, root, workbook, excel_filepath, undo_filepath, undo_jsontct_filepath, pdffilename, pdf_to_delete, report_type, incomplete_reports_folder, summary_data)
                else:
                    excel_filepath = datmg.xel_file_path if datmg.xel_file_path is not None else datmg.new_excel_fp
                    folder_selected = filedialog.askdirectory(title='Select Folder to save Report Files')
                    if not folder_selected:
                        if datmg.xel_file_path is not None:
                            os.rename(datmg.xel_file_path, undo_filepath)
                            os.rename(datmg.json_tct_filepath, undo_jsontct_filepath)
                            datmg.xel_file_path = undo_filepath
                            datmg.json_tct_filepath = undo_jsontct_filepath
                            datmg.save_tab_status_table_to_json(datmg.json_tct_filepath)
                        messagebox.showerror("Error", "Please select folder to store excel report.")
                        confirmation_window.destroy()
                        return
                    excel_filepath = os.path.join(folder_selected, os.path.basename(excel_filepath))
                    pdf_abs_location = os.path.join(folder_selected, pdffilename)
                    generate_pdf_copy(summary_data, pdf_abs_location, widmg, datmg, root)
                    move_related_pdfs(excel_filepath, pdf_abs_location, incomplete_reports_folder, folder_selected)
                    delete_excel_from_incomplete_reports(datmg, excel_filepath, incomplete_reports_folder)
                    workbook.save(excel_filepath)
                    messagebox.showinfo("Success", f"{excel_filepath} Saved Successfully to {folder_selected}" )
                    datmg.editing_spec_tab = False
                    back_bridge_step(widmg, datmg, root)

            elif metadata['active_tab'] == 'Prop Subs Inp Report':
                report_type = "SUBS"
                write_report_data_all_nd_pdpir(datmg, report_data, workbook, metadata['active_tab'])
                tot_joints = len(report_data)
                summary_data = generate_summary_entry_nd_dp(tot_joints, datmg.keyword_tally_dict, datmg)
                check_and_update_dpnd_complete(datmg, report_type)

                if datmg.excel_files_tct["REPORT_COMPLETE"] == "Incomplete":
                    open_continue_dp_report_window(widmg, datmg, root, workbook, excel_filepath, undo_filepath, undo_jsontct_filepath, pdffilename, pdf_to_delete, report_type, incomplete_reports_folder, summary_data)
                else:
                    excel_filepath = datmg.xel_file_path if datmg.xel_file_path is not None else datmg.new_excel_fp
                    folder_selected = filedialog.askdirectory(title='Select Folder to save Report Files')
                    if not folder_selected:
                        if datmg.xel_file_path is not None:
                            os.rename(datmg.xel_file_path, undo_filepath)
                            os.rename(datmg.json_tct_filepath, undo_jsontct_filepath)
                            datmg.xel_file_path = undo_filepath
                            datmg.json_tct_filepath = undo_jsontct_filepath
                            datmg.save_tab_status_table_to_json(datmg.json_tct_filepath)
                        messagebox.showerror("Error", "Please select folder to store excel report.")
                        confirmation_window.destroy()
                        return
                    excel_filepath = os.path.join(folder_selected, os.path.basename(excel_filepath))
                    pdf_abs_location = os.path.join(folder_selected, pdffilename)
                    generate_pdf_copy(summary_data, pdf_abs_location, widmg, datmg, root)
                    move_related_pdfs(excel_filepath, pdf_abs_location, incomplete_reports_folder, folder_selected)
                    delete_excel_from_incomplete_reports(datmg, excel_filepath, incomplete_reports_folder)
                    workbook.save(excel_filepath)
                    messagebox.showinfo("Success", f"{excel_filepath} Saved Successfully to {folder_selected}" )
                    datmg.editing_spec_tab = False
                    back_bridge_step(widmg, datmg, root)

        elif metadata['report_type'] == 'Tubing/Casing Report':
            pdf_abs_location = create_modify_excel_document_nd_tubing(folder_selected, widmg, datmg, root)
            excel_file = datmg.xel_file_path if datmg.xel_file_path is not None else datmg.new_excel_fp
            workbook = openpyxl.load_workbook(excel_file)
            excel_file = os.path.join(folder_selected, os.path.basename(excel_file))
            sheet = workbook['Tubing Insp Report']
            for joint_num, joint_values in report_data.items():
                row_num = 10 + int(joint_num.split('_')[-1]) - 1
                process_for_write_report_nd_tubing(joint_values, row_num, sheet, datmg)
            summary_data = generate_summary_entry_ndtube(datmg.keyword_tally_dict, datmg)
            generate_pdf_copy(summary_data, pdf_abs_location, widmg, datmg, root)
            write_summary_notes_nd_tubing(datmg, workbook)
            workbook.save(excel_file)
            messagebox.showinfo("Success", f"Excel File Created/Updated Successfully: {excel_file}" )

    elif metadata['branch'] == "TX":
        pdf_abs_location = create_modify_excel_document_tx_multi(incomplete_reports_folder, widmg, datmg, root)
        excel_file = datmg.xel_file_path if datmg.xel_file_path is not None else datmg.new_excel_fp
        workbook = openpyxl.load_workbook(excel_file)
        excel_file = os.path.join(incomplete_reports_folder, os.path.basename(excel_file))

        if metadata['active_tab'] == 'Prop Drill Pipe Inp Report':
            sheet = workbook['Prop Drill Pipe Inp Report']
            tot_joints = len(report_data)
            for joint_num, joint_values in report_data.items():
                row_num = 9 + int(joint_num.split('_')[-1]) - 1
                if datmg.json_data_dict['report_user_metadata']['report_style'] in ['Class 2 DBR', 'Class 2 NOT DBR']:
                    process_for_write_report_tx_pdpir_cl2dbr(joint_values, row_num, sheet, datmg)
                else:
                    process_for_write_report_tx_pdpir_fd(joint_values, row_num, sheet, datmg)
            if datmg.json_data_dict['report_user_metadata']['report_style'] in ['Class 2 DBR', 'Class 2 NOT DBR']:
                summary_data = generate_summary_entry_tx_cl2dbr(tot_joints, datmg.keyword_tally_dict, datmg)
                generate_pdf_copy(summary_data, pdf_abs_location, widmg, datmg, root)
            else:
                summary_data = generate_summary_entry_tx_fd(tot_joints, datmg.keyword_tally_dict, datmg)
                generate_pdf_copy(summary_data, pdf_abs_location, widmg, datmg, root)

            workbook.save(excel_file)
            messagebox.showinfo("Success", f"Excel File Created/Updated Successfully: {excel_file}" )

        elif metadata['active_tab'] == 'Prop HWDP Inp Report':
            sheet = workbook['Prop HWDP Inp Report']
            tot_joints = len(report_data)
            for joint_num, joint_values in report_data.items():
                row_num = 9 + int(joint_num.split('_')[-1]) - 1
                if datmg.json_data_dict['report_user_metadata']['report_style'] in ['Class 2 DBR', 'Class 2 NOT DBR']:
                    process_for_write_report_tx_pdpir_cl2dbr(joint_values, row_num, sheet, datmg)
                else:
                    process_for_write_report_tx_pdpir_fd(joint_values, row_num, sheet, datmg)
            if datmg.json_data_dict['report_user_metadata']['report_style'] in ['Class 2 DBR', 'Class 2 NOT DBR']:
                summary_data = generate_summary_entry_tx_cl2dbr(tot_joints, datmg.keyword_tally_dict, datmg)
                generate_pdf_copy(summary_data, pdf_abs_location, widmg, datmg, root)
            else:
                summary_data = generate_summary_entry_tx_fd(tot_joints, datmg.keyword_tally_dict, datmg)
                generate_pdf_copy(summary_data, pdf_abs_location, widmg, datmg, root)

            workbook.save(excel_file)
            messagebox.showinfo("Success", f"Excel File Created/Updated Successfully: {excel_file}" )

        elif metadata['active_tab'] == 'Prop Subs Inp Report':
            sheet = workbook['Prop Subs Inp Report']
            tot_joints = len(report_data)
            for joint_num, joint_values in report_data.items():
                row_num = 9 + int(joint_num.split('_')[-1]) - 1
                if datmg.json_data_dict['report_user_metadata']['report_style'] in ['Class 2 DBR', 'Class 2 NOT DBR']:
                    process_for_write_report_tx_pdpir_cl2dbr(joint_values, row_num, sheet, datmg)
                else:
                    process_for_write_report_tx_pdpir_fd(joint_values, row_num, sheet, datmg)
            if datmg.json_data_dict['report_user_metadata']['report_style'] in ['Class 2 DBR', 'Class 2 NOT DBR']:
                summary_data = generate_summary_entry_tx_cl2dbr(tot_joints, datmg.keyword_tally_dict, datmg)
                generate_pdf_copy(summary_data, pdf_abs_location, widmg, datmg, root)
            else:
                summary_data = generate_summary_entry_tx_fd(tot_joints, datmg.keyword_tally_dict, datmg)
                generate_pdf_copy(summary_data, pdf_abs_location, widmg, datmg, root)

            workbook.save(excel_file)
            messagebox.showinfo("Success", f"Excel File Created/Updated Successfully: {excel_file}" )


    elif metadata['branch'] == "WY":
        pass
        
    confirmation_window.destroy()

def find_highest_joint_num_for_export(widmg, datmg, root):
    active_tab = datmg.json_data_dict['active_tab']
    joint_data = datmg.json_data_dict['report_data'][active_tab]['joint_data']
    joint_numbers = []

    for key in joint_data.keys():
        match = re.match(r'Joint_(\d+)', key)
        if match:
            joint_number = int(match.group(1))
            joint_numbers.append(joint_number)
    if joint_numbers:
        highest_joint_number = max(joint_numbers)
        return highest_joint_number
    else:
        return 0



def create_modify_excel_document_nd_tubing(folder_selected, widmg, datmg, root):
    if datmg.xel_file_path is None:
        wb = openpyxl.load_workbook('DATE_INV_PRODUCER_2.375_2.875_Tubing_WELL_Grade_JTSJts.xlsx')
    else:
        wb = openpyxl.load_workbook(datmg.xel_file_path)

    metadata = datmg.json_data_dict['report_user_metadata']
    if metadata['connection_type_selection'] in ["PH6", "FATBOY-PH6", "CS8", "AOH", "BTC", "TWCC"]:
        del wb['Sum EUE Tubing']
    elif metadata['connection_type_selection'] == "EUE":
        del wb['Sum PH6 Tubing']

    data_sheet = wb['Data Sheet']
    data_mappings = {
        'B1': metadata['operator_entry'],
        'B2': metadata['contractor_entry'],
        'B6': metadata['location_entry'],
        'B3': metadata['date_entry'],
        'B4': metadata['invoice_entry'],
        'B11': metadata['connection_size_selection'],
        'B12': metadata['connection_type_selection'],
        'B13': metadata['grade_info'] if 'grade_info' in metadata else "",
        'B7': metadata['inspected_by_entry'],
        'B9': metadata['inspection_type_selection'] + (", " + metadata['inspection_type_additional'] if 'inspection_type_additional' in metadata else "")
    }

    if metadata['connection_size_selection'] == "2 7/8\"":
        file_con_sizesel = '2.875'
    elif metadata['connection_size_selection'] == "2 3/8\"":
        file_con_sizesel = '2.375'
    else:
        file_con_sizesel = metadata['connection_size_selection'].strip('"')

    for cell, value in data_mappings.items():
        data_sheet[cell] = value
    data_sheet.sheet_state = 'hidden'


    total_joint_count = int(widmg.update_rows_entry_widget.get())
    total_joints = total_joint_count + 10
    prop_sheet = wb['Tubing Insp Report']
    rows_to_delete = 709 - total_joints + 1
    if rows_to_delete > 0:
        prop_sheet.delete_rows(total_joints, rows_to_delete)
    
    sheet_index = wb.sheetnames.index('Tubing Insp Report')       
    wb.active = sheet_index

    if 'grade_info' in metadata and metadata['grade_info'] != "":
        grade_info = f"{metadata['grade_info']}_"
    else:
        grade_info = f"{metadata['grade_info']}_" if 'grade_info' in metadata else ""

    if datmg.xel_file_path is not None:
        datmg.xel_file_path = os.path.join(folder_selected, datmg.xel_file_path)
        wb.save(datmg.xel_file_path)
    else:
        datmg.new_excel_fp = os.path.join(folder_selected, f"{metadata['date_entry'].replace('/', '.')}_INV{metadata['invoice_entry']}_{metadata['operator_entry']}_{file_con_sizesel}_{grade_info}Tubing_{metadata['contractor_entry']}_{metadata['connection_type_selection']}_{total_joint_count}JTS.xlsx")
        wb.save(datmg.new_excel_fp) 
                
    pdffilename = f"{metadata['date_entry'].replace('/', '.')}_INV{metadata['invoice_entry']}_{metadata['operator_entry']}_{file_con_sizesel}_{grade_info}Tubing_{metadata['contractor_entry']}_{metadata['connection_type_selection']}_{total_joint_count}JTS_PDF-COPY.pdf"
    pdf_abs_location = os.path.join(folder_selected, pdffilename)

    return pdf_abs_location



def create_modify_excel_document_nd_pdpir(folder_selected, widmg, datmg, root):
    # Assuming you have the path to your Excel file in datmg.xel_file_path
    metadata = datmg.json_data_dict['report_user_metadata']
    pdf_to_delete = None
    undo_filepath = None 
    undo_jsontct_filepath = None 

    if datmg.xel_file_path is None:
        wb = openpyxl.load_workbook('DATE_INV_Inch DP Inspection_OPERATOR_CONTRACTOR.xlsx')
        datmg.excel_files_tct = {
            "PDPIR": {"Included?": "Yes", "Completed?": "-"},
            "HWDP": {"Included?": "Yes", "Completed?": "-"},
            "SUBS": {"Included?": "Yes", "Completed?": "-"},
            "REPORT_COMPLETE": "Incomplete"
        }
    else:
        wb = openpyxl.load_workbook(datmg.xel_file_path)
        undo_filepath = datmg.xel_file_path
        undo_jsontct_filepath = datmg.json_tct_filepath




    if 'add_pdpir_tab' in metadata and metadata['add_pdpir_tab'] == False:
        if 'Prop Drill Pipe Inp Report' in wb.sheetnames:
            del wb['Prop Drill Pipe Inp Report']
        if 'Sum Drill Pipe' in wb.sheetnames:
            del wb['Sum Drill Pipe']
        datmg.excel_files_tct['PDPIR']['Included?'] = 'No'

    if 'add_hwdp_tab' in metadata and metadata['add_hwdp_tab'] == False:
        if 'Prop HWDP Inp Report' in wb.sheetnames:
            del wb['Prop HWDP Inp Report']
        if 'Sum HWDP' in wb.sheetnames:
            del wb['Sum HWDP']
        datmg.excel_files_tct['HWDP']['Included?'] = 'No'

    if 'add_subs_tab' in metadata and metadata['add_subs_tab'] == False:
        if 'Prop Subs Inp Report' in wb.sheetnames:
            del wb['Prop Subs Inp Report']
        if 'Sum Subs' in wb.sheetnames:
            del wb['Sum Subs']
        datmg.excel_files_tct['SUBS']['Included?'] = 'No'



    # Handle data writing to the 'Data Sheet'
    data_sheet = wb['Data Sheet']
    data_mappings = {
        'B2': metadata['operator_entry'],
        'B3': metadata['contractor_entry'],
        'B12': metadata['location_entry'],
        'B4': metadata['date_entry'],
        'B5': metadata['invoice_entry'],
        'B8': float(metadata['connection_size_selection']),
        'B9': metadata['connection_type_selection'],
        'B6': metadata['inspected_by_entry'],
        'B10': metadata['grade_info'] if 'grade_info' in metadata else ""
    }

    for cell, value in data_mappings.items():
        data_sheet[cell] = value
    data_sheet.sheet_state = 'hidden'

    # Handle inspection_type data
    if metadata['active_tab'] == 'Prop Drill Pipe Inp Report':
        target_tab = 'Sum Drill Pipe'
    elif metadata['active_tab'] == 'Prop HWDP Inp Report':
        target_tab = 'Sum HWDP'
    elif metadata['active_tab'] == 'Prop Subs Inp Report':
        target_tab = 'Sum Subs'

    inspection_type = metadata['inspection_type_selection']
    additional_1 = ", Actual OD's, ID's, and Tong Space" if 'inspection_type_addodid' in metadata and metadata['inspection_type_addodid'] == True else None
    additional_2 = metadata['inspection_type_additional'] if 'inspection_type_additional' in metadata and metadata['inspection_type_additional'] is not None else None

    target_sheet = wb[target_tab]
    target_sheet['M3'] = inspection_type
    if additional_1 is not None:
        target_sheet['M5'] = additional_1
    if additional_2 is not None:
        target_sheet['M7'] = additional_2

    total_joint_count = int(widmg.update_rows_entry_widget.get())
    total_joints = total_joint_count + 9

    if 'Prop Drill Pipe Inp Report' in wb.sheetnames:
        typrep = 'DP'
    elif 'Prop HWDP Inp Report' in wb.sheetnames:
        if metadata['active_tab'] == 'Prop HWDP Inp Report':
            typrep = 'HWDP'
        else:
            typrep = 'DP'
    elif 'Prop Subs Inp Report' in wb.sheetnames:
        if metadata['active_tab'] == 'Prop Subs Inp Report':
            typrep = 'SUBS'
        else:
            typrep = 'HWDP'
    else:
        typrep = 'DP'

    print(f"metadata['active_tab'] = {metadata['active_tab']} right before the sheet_index variable is chosen")
    if metadata['active_tab'] == 'Prop HWDP Inp Report':
        pdf_typrep = 'HWDP'
        sheet_index = wb.sheetnames.index('Prop HWDP Inp Report')
    elif metadata['active_tab'] == 'Prop Subs Inp Report':
        pdf_typrep = 'SUBS'
        sheet_index = wb.sheetnames.index('Prop Subs Inp Report')
    elif metadata['active_tab'] == 'Prop Drill Pipe Inp Report':
        pdf_typrep = 'PDPIR'
        sheet_index = wb.sheetnames.index('Prop Drill Pipe Inp Report')

    if datmg.editing_spec_tab:
        wb = switcharoo_tabs(wb, pdf_typrep)


    if metadata['active_tab'] == 'Prop Drill Pipe Inp Report':
        prop_sheet = wb['Prop Drill Pipe Inp Report']
        rows_to_delete = 1008 - total_joints + 1
    elif metadata['active_tab'] == 'Prop HWDP Inp Report':
        prop_sheet = wb['Prop HWDP Inp Report']
        rows_to_delete = 608 - total_joints + 1
    elif metadata['active_tab'] == 'Prop Subs Inp Report':
        prop_sheet = wb['Prop Subs Inp Report']
        rows_to_delete = 158 - total_joints + 1

    if rows_to_delete > 0:
        prop_sheet.delete_rows(total_joints, rows_to_delete)

    wb.active = sheet_index

    # Update the filename with the new total_joint_count
    if datmg.xel_file_path is not None:
        # Extract previous count from the filename
        filename = os.path.basename(datmg.xel_file_path)
        json_tct_filename = os.path.basename(datmg.json_tct_filepath)
        if datmg.editing_spec_tab:
            filename, pdf_to_delete = update_new_edit_count_and_get_pdf_filename(pdf_typrep, filename)
            json_tct_filename = filename.replace(".xlsx", "_tct.json")
            pdf_to_delete = os.path.join(folder_selected, pdf_to_delete)


        match = re.search(r'(\d+)JTS', filename)
        tct_match = re.search(r'(\d+)JTS', json_tct_filename)
        if match:
            previous_count = int(match.group(1))
        else:
            previous_count = 0
        if tct_match:
            previous_count_tct = int(tct_match.group(1))
        else:
            previous_count_tct = 0

        # Calculate the new count
        new_count = previous_count + total_joint_count
        new_tct_count = previous_count_tct + total_joint_count

        # Replace the count in the filename
        new_filename = re.sub(r'(\d+)JTS', f'{new_count}JTS', filename)
        new_jsontct_filename = re.sub(r'(\d+)JTS', f'{new_tct_count}JTS', json_tct_filename)
        new_file_path = os.path.join(folder_selected, new_filename)
        new_jsontct_file_path = os.path.join(folder_selected, new_jsontct_filename)

        # Rename the existing file to the new filename
        os.rename(datmg.xel_file_path, new_file_path)
        os.rename(datmg.json_tct_filepath, new_jsontct_file_path)
        datmg.xel_file_path = new_file_path
        datmg.json_tct_filepath = new_jsontct_file_path


    else:
        base_full_path = os.path.join(folder_selected, f"{metadata['date_entry'].replace('/', '.')}_INV{metadata['invoice_entry']}_{metadata['connection_size_selection']} Inch {typrep} Inspection Report_{metadata['operator_entry']}_{metadata['contractor_entry']}_{total_joint_count}JTS")
        datmg.new_excel_fp = f"{base_full_path}.xlsx"
        datmg.json_tct_filepath = f"{base_full_path}_tct.json"
 

    pdffilename = f"{metadata['date_entry'].replace('/', '.')}_INV{metadata['invoice_entry']}_{metadata['connection_size_selection']} Inch {pdf_typrep} Inspection Report_{metadata['operator_entry']}_{metadata['contractor_entry']}_{total_joint_count}JTS_PDF-COPY.pdf"


    return wb, undo_filepath, undo_jsontct_filepath, pdf_to_delete, pdffilename





def create_modify_excel_document_tx_multi(folder_selected, widmg, datmg, root):
    # Assuming you have the path to your Excel file in datmg.xel_file_path
    temp_sel = datmg.json_data_dict['report_user_metadata']['report_style']
    if temp_sel == 'Full Dimensional':
        spreadsheet_name = 'DATE_INV_Inch DP Inspection_OPERATOR_CONTRACTOR_FullDimensional_LnkFix.xlsx'
    elif temp_sel == 'Class 2 DBR':
        spreadsheet_name = 'DATE_INV_Inch DP Inspection_OPERATOR_CONTRACTOR_CL2DBR_Test.xlsx'
    elif temp_sel == 'Class 2 NOT DBR':
        spreadsheet_name = 'DATE_INV_Inch DP Inspection_OPERATOR_CONTRACTOR_CL2NOTDBR_Test.xlsx'


    if datmg.xel_file_path is None:
        wb = openpyxl.load_workbook(spreadsheet_name)
    else:
        wb = openpyxl.load_workbook(datmg.xel_file_path)

    metadata = datmg.json_data_dict['report_user_metadata']

    if 'add_pdpir_tab' in metadata and metadata['add_pdpir_tab'] == False:
        del wb['Prop Drill Pipe Inp Report']
        del wb['Summary Drill Pipe']
    if 'add_hwdp_tab' in metadata and metadata['add_hwdp_tab'] == False:
        del wb['Prop HWDP Inp Report']
        del wb['Summary HWDP']
    if 'add_subs_tab' in metadata and metadata['add_subs_tab'] == False:
        del wb['Prop Subs Inp Report']
        del wb['Summary Sub']

    # Handle data writing to the 'Data Sheet'
    data_sheet = wb['DATA SHEET']
    data_mappings = {
        'B2': metadata['operator_entry'],
        'B3': metadata['contractor_entry'],
        'B4': metadata['location_entry'],
        'B5': metadata['date_entry'],
        'B6': metadata['invoice_entry'],
        'B9': float(metadata['connection_size_selection']),
        'B10': metadata['connection_type_selection'],
        'B7': metadata['inspected_by_entry'],
        'B11': metadata['grade_info'] if 'grade_info' in metadata else "",

    }

    for cell, value in data_mappings.items():
        data_sheet[cell] = value


    inspection_type = metadata['inspection_type_selection']
    additional_1 = ", Actual OD's, ID's, and Tong Space" if 'inspection_type_addodid' in metadata and metadata['inspection_type_addodid'] == True else None
    additional_2 = metadata['inspection_type_additional'] if 'inspection_type_additional' in metadata and metadata['inspection_type_additional'] is not None else None

    if metadata['active_tab'] == 'Prop Drill Pipe Inp Report':
        data_sheet['E3'] = inspection_type 
        data_sheet['E5'] = additional_1
        data_sheet['E7'] = additional_2
    elif metadata['active_tab'] == 'Prop HWDP Inp Report':
        data_sheet['E10'] = inspection_type 
        data_sheet['E12'] = additional_1
        data_sheet['E14'] = additional_2
    elif metadata['active_tab'] == 'Prop Subs Inp Report':
        data_sheet['E17'] = inspection_type 
        data_sheet['E19'] = additional_1
        data_sheet['E21'] = additional_2

    data_sheet.sheet_state = 'hidden'

    add_hidden_sheets = [
    'Drill Size Info', 'Drill Pipe Info', 'DP Count', 'Data Sheet DP', 
    'Data Sheet HWDP', 'HWDP Count', 'Data Sheet Subs', 'Sub Count',
    'Contractor-Operator', 'Shop List', 'PATHFINDER LOCATIONS'
    ]

    for sheet in add_hidden_sheets:
        if sheet in wb.sheetnames:
            to_hide = wb[sheet]
            to_hide.sheet_state = 'hidden'

    total_joint_count = int(widmg.update_rows_entry_widget.get())
    total_joints = total_joint_count + 9

    if metadata['active_tab'] == 'Prop Drill Pipe Inp Report':
        prop_sheet = wb['Prop Drill Pipe Inp Report']
        rows_to_delete = 1008 - total_joints + 1
    elif metadata['active_tab'] == 'Prop HWDP Inp Report':
        prop_sheet = wb['Prop HWDP Inp Report']
        rows_to_delete = 608 - total_joints + 1
    elif metadata['active_tab'] == 'Prop Subs Inp Report':
        prop_sheet = wb['Prop Subs Inp Report']
        rows_to_delete = 608 - total_joints + 1

    if rows_to_delete > 0:
        prop_sheet.delete_rows(total_joints, rows_to_delete)

    if 'Prop Drill Pipe Inp Report' in wb.sheetnames:
        typrep = 'DP'
        sheet_index = wb.sheetnames.index('Prop Drill Pipe Inp Report')
    elif 'Prop HWDP Inp Report' in wb.sheetnames:
        if metadata['active_tab'] == 'Prop HWDP Inp Report':
            typrep = 'HWDP'
            sheet_index = wb.sheetnames.index('Prop HWDP Inp Report')
        else:
            typrep = 'DP'
            sheet_index = wb.sheetnames.index('Prop Drill Pipe Inp Report')
    elif 'Prop Subs Inp Report' in wb.sheetnames:
        if metadata['active_tab'] == 'Prop Subs Inp Report':
            typrep = 'SUBS'
            sheet_index = wb.sheetnames.index('Prop Subs Inp Report')
        else:
            typrep = 'HWDP'
            sheet_index = wb.sheetnames.index('Prop HWDP Inp Report')
    else:
        typrep = 'DP'
        sheet_index = wb.sheetnames.index('Prop Drill Pipe Inp Report')


    if metadata['active_tab'] == 'Prop HWDP Inp Report':
        pdf_typrep = 'HWDP'
    elif metadata['active_tab'] == 'Prop Subs Inp Report':
        pdf_typrep = 'SUBS'
    else:
        pdf_typrep = 'DP'

    wb.active = sheet_index

    # Update the filename with the new total_joint_count
    if datmg.xel_file_path is not None:
        # Extract previous count from the filename
        filename = os.path.basename(datmg.xel_file_path)
        match = re.search(r'(\d+)JTS', filename)
        if match:
            previous_count = int(match.group(1))
        else:
            previous_count = 0

        # Calculate the new count
        new_count = previous_count + total_joint_count

        # Replace the count in the filename
        new_filename = re.sub(r'(\d+)JTS', f'{new_count}JTS', filename)
        new_file_path = os.path.join(folder_selected, new_filename)

        # Rename the existing file to the new filename
        os.rename(datmg.xel_file_path, new_file_path)
        datmg.xel_file_path = new_file_path
        wb.save(datmg.xel_file_path)
    else:
        datmg.new_excel_fp = os.path.join(folder_selected, f"{metadata['date_entry'].replace('/', '.')}_INV{metadata['invoice_entry']}_{metadata['connection_size_selection']} Inch {typrep} Inspection Report_{metadata['operator_entry']}_{metadata['contractor_entry']}_{total_joint_count}JTS.xlsx")
        wb.save(datmg.new_excel_fp)

    pdffilename = f"{metadata['date_entry'].replace('/', '.')}_INV{metadata['invoice_entry']}_{metadata['connection_size_selection']} Inch {pdf_typrep} Inspection Report_{metadata['operator_entry']}_{metadata['contractor_entry']}_{total_joint_count}JTS_PDF-COPY.pdf"
    pdf_abs_location = os.path.join(folder_selected, pdffilename)
    return pdf_abs_location

def process_for_write_report_tx_pdpir_fd(joint_values, row_num, sheet, datmg):
    reptyp = datmg.json_data_dict['report_user_metadata']['active_tab']
    taldict = datmg.keyword_tally_dict["Keyword Tallies"]
    joint_dict = datmg.keyword_tally_dict["Joint Tallies"]

    def update_keyword_tally(tally_dict, keyword):
        if keyword not in tally_dict:
            tally_dict[keyword] = 0
        tally_dict[keyword] += 1

    col_chooser = {
        "Not Recognized Keywords": {
            "BOX": {"Prop Drill Pipe Inp Report": "DO" , "Prop HWDP Inp Report": "DH" , "Prop Subs Inp Report": "DH" },
            "PIN": {"Prop Drill Pipe Inp Report": "DQ" , "Prop HWDP Inp Report": "DJ" , "Prop Subs Inp Report": "DJ" },
            "TUBE": {"Prop Drill Pipe Inp Report": "DP" , "Prop HWDP Inp Report": "DI" , "Prop Subs Inp Report": "DI" }
        },
        "Tong Space": {
            "BOX": {"Prop Drill Pipe Inp Report": "L" , "Prop HWDP Inp Report": "J" , "Prop Subs Inp Report": "J" },
            "PIN": {"Prop Drill Pipe Inp Report": "M", "Prop HWDP Inp Report": "K", "Prop Subs Inp Report": "K" },
            "TS-BOX": {"Prop Drill Pipe Inp Report": "L" , "Prop HWDP Inp Report": "J" , "Prop Subs Inp Report": "J" },
            "TS-PIN": {"Prop Drill Pipe Inp Report": "M", "Prop HWDP Inp Report": "K", "Prop Subs Inp Report": "K" }
        },
        "Bevel Diameter": {
            "BD-BOX": {"Prop Drill Pipe Inp Report": "J", "Prop HWDP Inp Report": "L", "Prop Subs Inp Report": "L"},
            "BD-PIN": {"Prop Drill Pipe Inp Report": "K", "Prop HWDP Inp Report": "M", "Prop Subs Inp Report": "M"},
            "BOX": {"Prop Drill Pipe Inp Report": "J", "Prop HWDP Inp Report": "L", "Prop Subs Inp Report": "L"},
            "PIN": {"Prop Drill Pipe Inp Report": "K", "Prop HWDP Inp Report": "M", "Prop Subs Inp Report": "M"},
        },
        "Tool Joint": {
            "BOX": {"Prop Drill Pipe Inp Report": "H", "Prop HWDP Inp Report": "G", "Prop Subs Inp Report": "G" },
            "PIN": {"Prop Drill Pipe Inp Report": "I", "Prop HWDP Inp Report": "H", "Prop Subs Inp Report": "H"},
            "ID": {"Prop Drill Pipe Inp Report": "I", "Prop HWDP Inp Report": "H", "Prop Subs Inp Report": "H"},
            "OD": {"Prop Drill Pipe Inp Report": "H", "Prop HWDP Inp Report": "G", "Prop Subs Inp Report": "G" },
            "TUBE": {"Prop Drill Pipe Inp Report": "H", "Prop HWDP Inp Report": "G", "Prop Subs Inp Report": "G" }

        },
        "Stress Relief Groove Diameter": {
            "STRES REL GRV": {"Prop HWDP Inp Report": "P", "Prop Subs Inp Report": "P"}
        },
        "Stress Relief Groove Length": {
            "STRES REL GRV": {"Prop HWDP Inp Report": "Q", "Prop Subs Inp Report": "Q"}
        },

        "C Bore Depth": {
            "C BORE": {"Prop Drill Pipe Inp Report": "N", "Prop HWDP Inp Report": "R", "Prop Subs Inp Report": "R"}
        },
        "C Bore Diameter": {
            "C BORE": {"Prop Drill Pipe Inp Report": "O", "Prop HWDP Inp Report": "S", "Prop Subs Inp Report": "S"}
        },
        "Boreback Dia": {
            "BORBAK": {"Prop HWDP Inp Report": "N", "Prop Subs Inp Report": "N"}
        },
        "Boreback Length": {
            "BORBAK": {"Prop HWDP Inp Report": "O", "Prop Subs Inp Report": "O"}
        },
        "Seal Width": {
            "SEAL WIDTH": {"Prop Drill Pipe Inp Report": "P", "Prop HWDP Inp Report": "T", "Prop Subs Inp Report": "T"}
        },
        "Pin Nose Diameter": {
            "PIN NOSE DIA": {"Prop Drill Pipe Inp Report": "Q", "Prop HWDP Inp Report": "U", "Prop Subs Inp Report": "U"}
        },

        "Serial #": {
            "SERIAL": {"Prop Drill Pipe Inp Report": "C" , "Prop HWDP Inp Report": "D" , "Prop Subs Inp Report": "D"},
        },
        "Critical Lengths": {
            "BOX": {"Prop Drill Pipe Inp Report": "DI", "Prop HWDP Inp Report": "DB"},
            "PIN": {"Prop Drill Pipe Inp Report": "DL", "Prop HWDP Inp Report": "DE"}
        },

        "R Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "V" , "Prop HWDP Inp Report": "Z" , "Prop Subs Inp Report": "Z" },
            "PIN": {"Prop Drill Pipe Inp Report": "X" , "Prop HWDP Inp Report": "AB" , "Prop Subs Inp Report": "AB" }
        },
        "R Type": {
            "BOX": {"Prop Drill Pipe Inp Report": "W" , "Prop HWDP Inp Report": "AA" , "Prop Subs Inp Report": "AA" },
            "PIN": {"Prop Drill Pipe Inp Report": "Y" , "Prop HWDP Inp Report": "AC" , "Prop Subs Inp Report": "AC" }
        },
        "R Value": {
            "BOX": {"Prop Drill Pipe Inp Report": "DH" , "Prop HWDP Inp Report": "DA" , "Prop Subs Inp Report": "DA" },
            "PIN": {"Prop Drill Pipe Inp Report": "DK" , "Prop HWDP Inp Report": "DD" , "Prop Subs Inp Report": "DD" }
        },

        "UT Value": {
            "UT": {"Prop Drill Pipe Inp Report": "DN" , "Prop HWDP Inp Report": "DG" , "Prop Subs Inp Report": "DG" },
        },
        "DS Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "Z" , "Prop HWDP Inp Report": "AD" , "Prop Subs Inp Report": "AD" },
            "PIN": {"Prop Drill Pipe Inp Report": "AA" , "Prop HWDP Inp Report": "AE" , "Prop Subs Inp Report": "AE" }
        },
        "DT Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "AB" , "Prop HWDP Inp Report": "AF" , "Prop Subs Inp Report": "AF" },
            "PIN": {"Prop Drill Pipe Inp Report": "AC" , "Prop HWDP Inp Report": "AG" , "Prop Subs Inp Report": "AG" }
        },
        "DTS Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "AD" , "Prop HWDP Inp Report": "AH" , "Prop Subs Inp Report": "AH" },
            "PIN": {"Prop Drill Pipe Inp Report": "AE" , "Prop HWDP Inp Report": "AI" , "Prop Subs Inp Report": "AI" }
        },
        "PIT Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "AF" },
            "PIN": {"Prop Drill Pipe Inp Report": "AG" }
        },
        "OR Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "AH" , "Prop HWDP Inp Report": "AJ" , "Prop Subs Inp Report": "AJ" },
            "PIN": {"Prop Drill Pipe Inp Report": "AI" , "Prop HWDP Inp Report": "AK" , "Prop Subs Inp Report": "AK" }
        },
        "BNT Tally": {
            "TUBE": {"Prop Drill Pipe Inp Report": "AJ" , "Prop HWDP Inp Report": "AL" , "Prop Subs Inp Report": "AL"}
        },
        "Other Damages 1": {
            "TUBE": {"Prop Drill Pipe Inp Report": "AK", "Prop HWDP Inp Report": "AM", "Prop Subs Inp Report": "AM"},
            "BOX": {"Prop Drill Pipe Inp Report": "AK", "Prop HWDP Inp Report": "AM", "Prop Subs Inp Report": "AM"},
            "PIN": {"Prop Drill Pipe Inp Report": "AK", "Prop HWDP Inp Report": "AM", "Prop Subs Inp Report": "AM"}
        },
        "Other Damages 2": {
            "TUBE": {"Prop Drill Pipe Inp Report": "AL", "Prop HWDP Inp Report": "AN", "Prop Subs Inp Report": "AN"},
            "BOX": {"Prop Drill Pipe Inp Report": "AL", "Prop HWDP Inp Report": "AN", "Prop Subs Inp Report": "AN"},
            "PIN": {"Prop Drill Pipe Inp Report": "AL", "Prop HWDP Inp Report": "AN", "Prop Subs Inp Report": "AN"}
        },
        "Other Damages 3": {
            "TUBE": {"Prop Drill Pipe Inp Report": "AM", "Prop HWDP Inp Report": "AO", "Prop Subs Inp Report": "AO"},
            "BOX": {"Prop Drill Pipe Inp Report": "AM", "Prop HWDP Inp Report": "AO", "Prop Subs Inp Report": "AO"},
            "PIN": {"Prop Drill Pipe Inp Report": "AM", "Prop HWDP Inp Report": "AO", "Prop Subs Inp Report": "AO"}
        },
        "Other Damages 4": {
            "TUBE": {"Prop Drill Pipe Inp Report": "AN"},
            "BOX": {"Prop Drill Pipe Inp Report": "AN"},
            "PIN": {"Prop Drill Pipe Inp Report": "AN"}
        },
        "DHB Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "AO" , "Prop HWDP Inp Report": "AP" , "Prop Subs Inp Report": "AP" },
            "PIN": {"Prop Drill Pipe Inp Report": "AP" , "Prop HWDP Inp Report": "AQ" , "Prop Subs Inp Report": "AQ" }
        },
        "HB Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "AQ" , "Prop HWDP Inp Report": "AR" , "Prop Subs Inp Report": "AR" },
            "PIN": {"Prop Drill Pipe Inp Report": "AT" , "Prop HWDP Inp Report": "AU" , "Prop Subs Inp Report": "AU" }
        },

        "HBCP Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "AR" , "Prop HWDP Inp Report": "AS" , "Prop Subs Inp Report": "AS" },
            "PIN": {"Prop Drill Pipe Inp Report": "AS" , "Prop HWDP Inp Report": "AT" , "Prop Subs Inp Report": "AT" }
        },
        "DBRHB Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "AU" , "Prop HWDP Inp Report": "AV" , "Prop Subs Inp Report": "AV" },
            "PIN": {"Prop Drill Pipe Inp Report": "AV" , "Prop HWDP Inp Report": "AW" , "Prop Subs Inp Report": "AW" }
        },
        "MW Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "AW" , "Prop HWDP Inp Report": "AX" , "Prop Subs Inp Report": "AX" },
            "PIN": {"Prop Drill Pipe Inp Report": "AW" , "Prop HWDP Inp Report": "AX" , "Prop Subs Inp Report": "AX" },
            "TUBE": {"Prop Drill Pipe Inp Report": "AW" , "Prop HWDP Inp Report": "AX" , "Prop Subs Inp Report": "AX" }
        },
        "MT Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "AX" , "Prop HWDP Inp Report": "AY" , "Prop Subs Inp Report": "AY" },
            "PIN": {"Prop Drill Pipe Inp Report": "AY" , "Prop HWDP Inp Report": "AZ" , "Prop Subs Inp Report": "AZ" },
        },
        "MS Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "AZ" , "Prop HWDP Inp Report": "BA" , "Prop Subs Inp Report": "BA" },
            "PIN": {"Prop Drill Pipe Inp Report": "BA" , "Prop HWDP Inp Report": "BB" , "Prop Subs Inp Report": "BB" },
        },
        "MOD Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "BB" , "Prop HWDP Inp Report": "BC" , "Prop Subs Inp Report": "BC" },
            "PIN": {"Prop Drill Pipe Inp Report": "BB" , "Prop HWDP Inp Report": "BC" , "Prop Subs Inp Report": "BC" },
            "TUBE": {"Prop Drill Pipe Inp Report": "BB" , "Prop HWDP Inp Report": "BC" , "Prop Subs Inp Report": "BC" },
        },
        "DAM Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "BC" , "Prop HWDP Inp Report": "BD" , "Prop Subs Inp Report": "BD" },
            "PIN": {"Prop Drill Pipe Inp Report": "BC" , "Prop HWDP Inp Report": "BD" , "Prop Subs Inp Report": "BD" },
            "TUBE": {"Prop Drill Pipe Inp Report": "BC" , "Prop HWDP Inp Report": "BD" , "Prop Subs Inp Report": "BD" },
        },

        "EMI Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "BD" , "Prop HWDP Inp Report": "BE" , "Prop Subs Inp Report": "BE" },
            "PIN": {"Prop Drill Pipe Inp Report": "BD" , "Prop HWDP Inp Report": "BE" , "Prop Subs Inp Report": "BE" },
            "TUBE": {"Prop Drill Pipe Inp Report": "BD" , "Prop HWDP Inp Report": "BE" , "Prop Subs Inp Report": "BE" },
        },

        "OTHER Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "BE" , "Prop HWDP Inp Report": "BF" , "Prop Subs Inp Report": "BF" },
            "PIN": {"Prop Drill Pipe Inp Report": "BE" , "Prop HWDP Inp Report": "BF" , "Prop Subs Inp Report": "BF" },
            "TUBE": {"Prop Drill Pipe Inp Report": "BE" , "Prop HWDP Inp Report": "BF" , "Prop Subs Inp Report": "BF" },
        },
        "TR Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "BF" , "Prop HWDP Inp Report": "BG" , "Prop Subs Inp Report": "BG" },
            "PIN": {"Prop Drill Pipe Inp Report": "BG" , "Prop HWDP Inp Report": "BH" , "Prop Subs Inp Report": "BH" },
        },
        "BVR Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "BH" , "Prop HWDP Inp Report": "BI" , "Prop Subs Inp Report": "BI" },
            "PIN": {"Prop Drill Pipe Inp Report": "BI" , "Prop HWDP Inp Report": "BJ" , "Prop Subs Inp Report": "BJ" },
        },
        "SUBS CONN/DATA": {
            "SUBS CONN/DATA": {"Prop Subs Inp Report": "V"}
        },
        "COMMENTS": {
            "COMMENTS": {"Prop Drill Pipe Inp Report": "DR", "Prop HWDP Inp Report": "DK", "Prop Subs Inp Report": "DK" }
        }
    }

    def process_tube_col(col_name, col_value, row_num, sheet):
        keywords = col_value.split()
        n = len(keywords)
        for i in range(n):
            keyword = keywords[i]
            if keyword == "MW":
                sheet[f"{col_chooser['MW Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, keyword)
            elif keyword == "MOD":
                sheet[f"{col_chooser['MOD Tally'][col_name][reptyp]}{row_num}"] = "X"
                process_mod_keyword(col_name, keywords, keyword, row_num, sheet)
                update_keyword_tally(taldict, keyword)
            elif keyword == "DAM":
                sheet[f"{col_chooser['DAM Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, keyword)
            elif keyword == "EMI":
                sheet[f"{col_chooser['EMI Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, keyword)
            elif keyword == "OTHER":
                sheet[f"{col_chooser['OTHER Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, keyword)
            elif keyword == "BNT" or keyword == "BENT":
                sheet[f"{col_chooser['BNT Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, "BNT")
            elif keyword == "ODAM1":
                sheet[f"{col_chooser['Other Damages 1'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "ODAM2":
                sheet[f"{col_chooser['Other Damages 2'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "ODAM3":
                sheet[f"{col_chooser['Other Damages 3'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "ODAM4":
                if reptyp != 'Prop Drill Pipe Inp Report':
                    sheet[f"{col_chooser['Not Recognized Keywords'][col_name][reptyp]}{row_num}"] = keyword
                else:
                    sheet[f"{col_chooser['Other Damages 4'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif not re.match(r'^(\d+(\.\d+)?|\d+|\d+\s\d+/\d+|\d+/\d+)$', keyword.replace('_', ' ')):
                if i + 1 < n:  
                    next_keyword = keywords[i + 1].replace('_', ' ')
                    if re.match(r'^(\d+(\.\d+)?|\d+|\d+\s\d+/\d+|\d+/\d+)$', next_keyword):
                        sheet[f"{col_chooser['Not Recognized Keywords'][col_name][reptyp]}{row_num}"] = f"{keyword} {next_keyword}"
                        continue  
                sheet[f"{col_chooser['Not Recognized Keywords'][col_name][reptyp]}{row_num}"] = keyword



    def process_box_pin_col(col_name, col_value, row_num, sheet):
        keywords = col_value.split()
        n = len(keywords)
        for i in range(n):
            keyword = keywords[i]
            if len(keyword) == 3 and keyword.isdigit():
                sheet[f"{col_chooser['Critical Lengths'][col_name][reptyp]}{row_num}"] = keyword
            elif keyword == "TS":
                process_ts_keyword(col_name, keywords, keyword, row_num, sheet)
            elif keyword == "TJ":
                process_tj_keyword(col_name, keywords, keyword, row_num, sheet)
            elif keyword == "BD":
                process_bd_keyword(col_name, keywords, keyword, row_num, sheet)
            elif keyword == "MT":
                sheet[f"{col_chooser['MT Tally'][col_name][reptyp]}{row_num}"] = "X"
                process_mtong_keyword(col_name, keywords, keyword, row_num, sheet)
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "MS":
                sheet[f"{col_chooser['MS Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "DS":
                sheet[f"{col_chooser['DS Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "DT":
                sheet[f"{col_chooser['DT Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "DTS":
                sheet[f"{col_chooser['DTS Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "PIT":
                sheet[f"{col_chooser['PIT Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "OR":
                sheet[f"{col_chooser['OR Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "DHB":
                sheet[f"{col_chooser['DHB Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "HB":
                sheet[f"{col_chooser['HB Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "DBRHB":
                sheet[f"{col_chooser['DBRHB Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "HBCP":
                sheet[f"{col_chooser['HBCP Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "MOD":
                sheet[f"{col_chooser['MOD Tally'][col_name][reptyp]}{row_num}"] = "X"
                process_mod_keyword(col_name, keywords, keyword, row_num, sheet)
                update_keyword_tally(taldict, keyword)
            elif keyword == "R1":
                sheet[f"{col_chooser['R Tally'][col_name][reptyp]}{row_num}"] = "X"
                sheet[f"{col_chooser['R Type'][col_name][reptyp]}{row_num}"] = 1
                process_r_keyword(col_name, keywords, row_num, sheet)
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "R2":
                sheet[f"{col_chooser['R Tally'][col_name][reptyp]}{row_num}"] = "X"
                sheet[f"{col_chooser['R Type'][col_name][reptyp]}{row_num}"] = 2
                process_r_keyword(col_name, keywords, row_num, sheet)
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "R3":
                sheet[f"{col_chooser['R Tally'][col_name][reptyp]}{row_num}"] = "X"
                sheet[f"{col_chooser['R Type'][col_name][reptyp]}{row_num}"] = 3
                process_r_keyword(col_name, keywords, row_num, sheet)
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "R4":
                sheet[f"{col_chooser['R Tally'][col_name][reptyp]}{row_num}"] = "X"
                sheet[f"{col_chooser['R Type'][col_name][reptyp]}{row_num}"] = 4
                process_r_keyword(col_name, keywords, row_num, sheet)
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "ODAM":
                if reptyp == "Prop Drill Pipe Inp Report":
                    sheet[f"{col_chooser['BOX/PIN Other Damages Tally'][col_name][reptyp]}{row_num}"] = "X"
                else:
                    sheet[f"{col_chooser['Not Recognized Keywords'][col_name][reptyp]}{row_num}"] = keyword
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "ODAM1":
                sheet[f"{col_chooser['Other Damages 1'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}")
            elif keyword == "ODAM2":
                sheet[f"{col_chooser['Other Damages 2'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}")
            elif keyword == "ODAM3":
                sheet[f"{col_chooser['Other Damages 3'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}")
            elif keyword == "ODAM4":
                if reptyp != 'Prop Drill Pipe Inp Report':
                    sheet[f"{col_chooser['Not Recognized Keywords'][col_name][reptyp]}{row_num}"] = keyword
                else:
                    sheet[f"{col_chooser['Other Damages 4'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}")
            elif keyword == "TR":
                sheet[f"{col_chooser['TR Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "BVR":
                sheet[f"{col_chooser['BVR Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif not re.match(r'^(\d+(\.\d+)?|\d+|\d+\s\d+/\d+|\d+/\d+)$', keyword.replace('_', ' ')):
                if i + 1 < n:  
                    next_keyword = keywords[i + 1].replace('_', ' ')
                    if re.match(r'^(\d+(\.\d+)?|\d+|\d+\s\d+/\d+|\d+/\d+)$', next_keyword):
                        sheet[f"{col_chooser['Not Recognized Keywords'][col_name][reptyp]}{row_num}"] = f"{keyword} {next_keyword}"
                        continue  
                sheet[f"{col_chooser['Not Recognized Keywords'][col_name][reptyp]}{row_num}"] = keyword




    def process_ts_keyword(col_name, keywords, keyword, row_num, sheet):
        keywords = col_value.split()
        n = len(keywords)
        ts_index = keywords.index(keyword)
        if ts_index < n - 1:
            fraction_dec_pat = re.compile(r'((\d+_)?\d+/\d+|\d+(\.\d{3})?)')
            val_match = keywords[ts_index + 1]
            if fraction_dec_pat.match(val_match):
                sheet[f"{col_chooser['Tong Space'][col_name][reptyp]}{row_num}"] = val_match.replace('_', ' ')
    def process_tj_keyword(col_name, keywords, keyword, row_num, sheet):
        keywords = col_value.split()
        n = len(keywords)
        tj_index = keywords.index(keyword)
        if tj_index < n - 1:
            fraction_dec_pat = re.compile(r'((\d+_)?\d+/\d+|\d+(\.\d{3})?)')
            val_match = keywords[tj_index + 1]
            if fraction_dec_pat.match(val_match):
                sheet[f"{col_chooser['Tool Joint'][col_name][reptyp]}{row_num}"] = val_match.replace('_', ' ')
    def process_bd_keyword(col_name, keywords, keyword, row_num, sheet):
        keywords = col_value.split()
        n = len(keywords)
        bd_index = keywords.index(keyword)
        if bd_index < n - 1:
            fraction_dec_pat = re.compile(r'((\d+_)?\d+/\d+|\d+(\.\d{3})?)')
            val_match = keywords[bd_index + 1]
            if fraction_dec_pat.match(val_match):
                sheet[f"{col_chooser['Bevel Diameter'][col_name][reptyp]}{row_num}"] = val_match.replace('_', ' ')



    def process_stres_rel_grv_col(col_name, col_value, row_num, sheet):
        keywords = col_value.split()
        n = len(keywords)
        for i in range(n):
            keyword = keywords[i]
            if keyword in ["D", "DIA", "DIAMETER"]:
                process_srg_dia_keyword(col_name, keywords, keyword, row_num, sheet)
            elif keyword in ["L", "LEN", "LENGTH"]:
                process_srg_len_keyword(col_name, keywords, keyword, row_num, sheet)
    def process_srg_dia_keyword(col_name, keywords, keyword, row_num, sheet):
        dia_index = keywords.index(keyword)
        if dia_index < len(keywords) - 1:
            fraction_dec_pat = re.compile(r'((\d+_)?\d+/\d+|\d+(\.\d{3})?)')
            val_match = keywords[dia_index + 1]
            if fraction_dec_pat.match(val_match):
                sheet[f"{col_chooser['Stress Relief Groove Diameter'][col_name][reptyp]}{row_num}"] = val_match.replace('_', ' ')
    def process_srg_len_keyword(col_name, keywords, keyword, row_num, sheet):
        len_index = keywords.index(keyword)
        if len_index < len(keywords) - 1:
            fraction_dec_pat = re.compile(r'((\d+_)?\d+/\d+|\d+(\.\d{3})?)')
            val_match = keywords[len_index + 1]
            if fraction_dec_pat.match(val_match):
                sheet[f"{col_chooser['Stress Relief Groove Length'][col_name][reptyp]}{row_num}"] = val_match.replace('_', ' ')




    def process_borbak_col(col_name, col_value, row_num, sheet):
        keywords = col_value.split()
        n = len(keywords)
        for i in range(n):
            keyword = keywords[i]
            if keyword in ["D", "DIA"]:
                process_borbak_dia_keyword(col_name, keywords, keyword, row_num, sheet)
            elif keyword in ["L", "LEN"]:
                process_borbak_len_keyword(col_name, keywords, keyword, row_num, sheet)
    def process_borbak_dia_keyword(col_name, keywords, keyword, row_num, sheet):
        dia_index = keywords.index(keyword)
        if dia_index < len(keywords) - 1:
            fraction_dec_pat = re.compile(r'((\d+_)?\d+/\d+|\d+(\.\d{3})?)')
            val_match = keywords[dia_index + 1]
            if fraction_dec_pat.match(val_match):
                sheet[f"{col_chooser['Boreback Dia'][col_name][reptyp]}{row_num}"] = val_match.replace('_', ' ')
    def process_borbak_len_keyword(col_name, keywords, keyword, row_num, sheet):
        len_index = keywords.index(keyword)
        if len_index < len(keywords) - 1:
            fraction_dec_pat = re.compile(r'((\d+_)?\d+/\d+|\d+(\.\d{3})?)')
            val_match = keywords[len_index + 1]
            if fraction_dec_pat.match(val_match):
                sheet[f"{col_chooser['Boreback Length'][col_name][reptyp]}{row_num}"] = val_match.replace('_', ' ')



    def process_cbore_col(col_name, col_value, row_num, sheet):
        keywords = col_value.split()
        n = len(keywords)
        for i in range(n):
            keyword = keywords[i]
            if keyword in ["DEP", "DEPTH", "DE", "DPTH"]:
                process_cbore_depth_keyword(col_name, keywords, keyword, row_num, sheet)
            elif keyword in ["DIA", "DI", "DIAMETER", "DIAMTR"]:
                process_cbore_dia_keyword(col_name, keywords, keyword, row_num, sheet)
    def process_cbore_dia_keyword(col_name, keywords, keyword, row_num, sheet):
        dia_index = keywords.index(keyword)
        if dia_index < len(keywords) - 1:
            fraction_dec_pat = re.compile(r'((\d+_)?\d+/\d+|\d+(\.\d{3})?)')
            val_match = keywords[dia_index + 1]
            if fraction_dec_pat.match(val_match):
                sheet[f"{col_chooser['C Bore Diameter'][col_name][reptyp]}{row_num}"] = val_match.replace('_', ' ')
    def process_cbore_depth_keyword(col_name, keywords, keyword, row_num, sheet):
        depth_index = keywords.index(keyword)
        if depth_index < len(keywords) - 1:
            fraction_dec_pat = re.compile(r'((\d+_)?\d+/\d+|\d+(\.\d{3})?)')
            val_match = keywords[depth_index + 1]
            if fraction_dec_pat.match(val_match):
                sheet[f"{col_chooser['C Bore Depth'][col_name][reptyp]}{row_num}"] = val_match.replace('_', ' ')



    def process_short_keywords(col_name, keywords, keyword, row_num, sheet):
        sb_index = keywords.index(keyword) if keyword == "SB" else None
        sp_index = keywords.index(keyword) if keyword == "SP" else None 
        short_index = sb_index if sb_index is not None else sp_index
        if short_index < len(keywords) -1:
            fraction_dec_pat = re.compile(r'((\d+_)?\d+/\d+|\d+(\.\d{3})?)')
            val_match = keywords[short_index + 1]
            if fraction_dec_pat.match(val_match):
                sheet[f"{col_chooser['Tool Joint'][col_name][reptyp]}{row_num}"] = val_match.replace('_', ' ') 

    def process_long_keywords(col_name, keywords, keyword, row_num, sheet):
        lp_index = keywords.index(keyword) if keyword == "LP" else None  
        lb_index = keywords.index(keyword) if keyword == "LB" else None
        long_index = lp_index if lp_index is not None else lb_index
        if long_index < len(keywords) - 1:
            fraction_dec_pat = re.compile(r'((\d+_)?\d+/\d+|\d+(\.\d{3})?)')
            val_match = keywords[long_index + 1]
            if fraction_dec_pat.match(val_match):
                sheet[f"{col_chooser['Tool Joint'][col_name][reptyp]}{row_num}"] = val_match.replace('_', ' ') 


    def process_mtong_keyword(col_name, keywords, keyword, row_num, sheet):
        mt_index = keywords.index(keyword)
        if mt_index < len(keywords) - 1:
            fraction_dec_pat = re.compile(r'((\d+_)?\d+/\d+|\d+(\.\d{3})?)')
            val_match = keywords[mt_index + 1]
            if fraction_dec_pat.match(val_match):
                sheet[f"{col_chooser['Tong Space'][col_name][reptyp]}{row_num}"] = val_match.replace('_', ' ')


    def process_mod_keyword(col_name, keywords, keyword, row_num, sheet):
        mod_index = keywords.index(keyword)
        if mod_index < len(keywords) - 1:
            fraction_dec_pat = re.compile(r'((\d+_)?\d+/\d+|\d+(\.\d{3})?)')
            val_match = keywords[mod_index + 1]
            if fraction_dec_pat.match(val_match):
                sheet[f"{col_chooser['Tool Joint']['BOX'][reptyp]}{row_num}"]  = val_match.replace('_', ' ')

    def process_r_keyword(col_name, keywords, row_num, sheet):
        for keyword in keywords:
            if keyword in ["R1", "R2", "R3", "R4"]:
                r_index = keywords.index(keyword)
        if r_index < len(keywords) - 1 and keywords[r_index + 1].isdigit() and len(keywords[r_index + 1]) == 6:
            sheet[f"{col_chooser['R Value'][col_name][reptyp]}{row_num}"] = keywords[r_index + 1]




    for col_name, col_value in joint_values.items():
        col_value = re.sub(r'(\d+)\s(\d+/\d+)', r'\1_\2', col_value)
        if col_name in ["BOX", "PIN"]:
            process_box_pin_col(col_name, col_value, row_num, sheet)
        elif col_name == "UT":
            sheet[f"{col_chooser['UT Value'][col_name][reptyp]}{row_num}"] = col_value
        elif col_name == "SERIAL":
            sheet[f"{col_chooser['Serial #'][col_name][reptyp]}{row_num}"]  = col_value
        elif col_name == "TUBE":
            process_tube_col(col_name, col_value, row_num, sheet)
        elif col_name == "BORBAK":
            process_borbak_col(col_name, col_value, row_num, sheet)
        elif col_name == "C BORE":
            process_cbore_col(col_name, col_value, row_num, sheet)
        elif col_name == "SEAL WIDTH":
            sheet[f"{col_chooser['Seal Width'][col_name][reptyp]}{row_num}"] = col_value.replace('_', ' ')
        elif col_name == "PIN NOSE DIA":
            sheet[f"{col_chooser['Pin Nose Diameter'][col_name][reptyp]}{row_num}"] = col_value.replace('_', ' ')
        elif col_name == "STRES REL GRV":
            process_stres_rel_grv_col(col_name, col_value, row_num, sheet)
        elif col_name == "BD-BOX":
            if col_value != "":
                sheet[f"{col_chooser['Bevel Diameter'][col_name][reptyp]}{row_num}"] = col_value.replace('_', ' ')
        elif col_name == "BD-PIN":
            if col_value != "":
                sheet[f"{col_chooser['Bevel Diameter'][col_name][reptyp]}{row_num}"] = col_value.replace('_', ' ')
        elif col_name == "ID":
            if col_value != "":
                sheet[f"{col_chooser['Tool Joint'][col_name][reptyp]}{row_num}"] = col_value.replace('_', ' ')
        elif col_name == "OD":
            if col_value != "":
                sheet[f"{col_chooser['Tool Joint'][col_name][reptyp]}{row_num}"] = col_value.replace('_', ' ')
        elif col_name == "TS-PIN":
            if col_value != "":
                sheet[f"{col_chooser['Tong Space'][col_name][reptyp]}{row_num}"] = col_value.replace('_', ' ')
        elif col_name == "TS-BOX":
            if col_value != "":
                sheet[f"{col_chooser['Tong Space'][col_name][reptyp]}{row_num}"] = col_value.replace('_', ' ')
        elif col_name == "DESCRIPTION":
            sheet[f"C{row_num}"] = col_value
        elif col_name == "SUBS CONN/DATA":
            sheet[f"{col_chooser['SUBS CONN/DATA'][col_name][reptyp]}{row_num}"] = col_value
        elif col_name == "COMMENTS":
            sheet[f"{col_chooser['COMMENTS'][col_name][reptyp]}{row_num}"] = col_value



def process_for_write_report_tx_pdpir_cl2dbr(joint_values, row_num, sheet, datmg):
    reptyp = datmg.json_data_dict['report_user_metadata']['active_tab']
    taldict = datmg.keyword_tally_dict["Keyword Tallies"]
    joint_dict = datmg.keyword_tally_dict["Joint Tallies"]


    def update_keyword_tally(tally_dict, keyword):
        if keyword not in tally_dict:
            tally_dict[keyword] = 0
        tally_dict[keyword] += 1

    col_chooser = {
        "SUBS CONN/DATA": {
            "SUBS CONN/DATA": {"Prop Subs Inp Report": "N"}
        },
        "COMMENTS": {
            "COMMENTS": {"Prop Drill Pipe Inp Report": "DD", "Prop HWDP Inp Report": "DD", "Prop Subs Inp Report": "DD" }
        },
        "Serial #": {
            "SERIAL": {"Prop Drill Pipe Inp Report": "C" , "Prop HWDP Inp Report": "D" , "Prop Subs Inp Report": "D"},
        },
        "UT Value": {
            "UT": {"Prop Drill Pipe Inp Report": "CZ" , "Prop HWDP Inp Report": "CZ" , "Prop Subs Inp Report": "CZ" },
        },
        "Tool Joint": {
            "BOX": {"Prop Drill Pipe Inp Report": "H", "Prop HWDP Inp Report": "G", "Prop Subs Inp Report": "G" },
            "PIN": {"Prop Drill Pipe Inp Report": "I", "Prop HWDP Inp Report": "H", "Prop Subs Inp Report": "H"},
            "ID": {"Prop Drill Pipe Inp Report": "I", "Prop HWDP Inp Report": "H", "Prop Subs Inp Report": "H"},
            "OD": {"Prop Drill Pipe Inp Report": "H", "Prop HWDP Inp Report": "G", "Prop Subs Inp Report": "G" },
            "TUBE": {"Prop Drill Pipe Inp Report": "H", "Prop HWDP Inp Report": "G", "Prop Subs Inp Report": "G" }

        },
        "Bevel Diameter": {
            "BD-BOX": {"Prop Drill Pipe Inp Report": "J", "Prop HWDP Inp Report": "L", "Prop Subs Inp Report": "L"},
            "BD-PIN": {"Prop Drill Pipe Inp Report": "K", "Prop HWDP Inp Report": "M", "Prop Subs Inp Report": "M"},
            "BOX": {"Prop Drill Pipe Inp Report": "J", "Prop HWDP Inp Report": "L", "Prop Subs Inp Report": "L"},
            "PIN": {"Prop Drill Pipe Inp Report": "K", "Prop HWDP Inp Report": "M", "Prop Subs Inp Report": "M"},
        },

        "Tong Space": {
            "BOX": {"Prop Drill Pipe Inp Report": "L" , "Prop HWDP Inp Report": "J" , "Prop Subs Inp Report": "J" },
            "PIN": {"Prop Drill Pipe Inp Report": "M", "Prop HWDP Inp Report": "K", "Prop Subs Inp Report": "K" },
            "TS-BOX": {"Prop Drill Pipe Inp Report": "L" , "Prop HWDP Inp Report": "J" , "Prop Subs Inp Report": "J" },
            "TS-PIN": {"Prop Drill Pipe Inp Report": "M", "Prop HWDP Inp Report": "K", "Prop Subs Inp Report": "K" }
        },
        "Critical Lengths": {
            "BOX": {"Prop Drill Pipe Inp Report": "CU", "Prop HWDP Inp Report": "CU"},
            "PIN": {"Prop Drill Pipe Inp Report": "CX", "Prop HWDP Inp Report": "CX"}
        },
        "R Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "R" , "Prop HWDP Inp Report": "R" , "Prop Subs Inp Report": "R" },
            "PIN": {"Prop Drill Pipe Inp Report": "T" , "Prop HWDP Inp Report": "T" , "Prop Subs Inp Report": "T" }
        },
        "R Type": {
            "BOX": {"Prop Drill Pipe Inp Report": "S" , "Prop HWDP Inp Report": "S" , "Prop Subs Inp Report": "S" },
            "PIN": {"Prop Drill Pipe Inp Report": "U" , "Prop HWDP Inp Report": "U" , "Prop Subs Inp Report": "U" }
        },
        "R Value": {
            "BOX": {"Prop Drill Pipe Inp Report": "CT" , "Prop HWDP Inp Report": "CT" , "Prop Subs Inp Report": "CT" },
            "PIN": {"Prop Drill Pipe Inp Report": "CW" , "Prop HWDP Inp Report": "CW" , "Prop Subs Inp Report": "CW" }
        },
        "DS Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "V" , "Prop HWDP Inp Report": "V" , "Prop Subs Inp Report": "V" },
            "PIN": {"Prop Drill Pipe Inp Report": "W" , "Prop HWDP Inp Report": "W" , "Prop Subs Inp Report": "W" }
        },
        "DT Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "X" , "Prop HWDP Inp Report": "X" , "Prop Subs Inp Report": "X" },
            "PIN": {"Prop Drill Pipe Inp Report": "Y" , "Prop HWDP Inp Report": "Y" , "Prop Subs Inp Report": "Y" }
        },
        "DTS Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "Z"},
            "PIN": {"Prop Drill Pipe Inp Report": "AA"}
        },
        "PIT Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "AB", "Prop HWDP Inp Report": "AE", "Prop Subs Inp Report": "AE"},
            "PIN": {"Prop Drill Pipe Inp Report": "AC", "Prop HWDP Inp Report": "AF", "Prop Subs Inp Report": "AF"}
        },
        "OR Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "AD" , "Prop HWDP Inp Report": "AC" , "Prop Subs Inp Report": "AC" },
            "PIN": {"Prop Drill Pipe Inp Report": "AE" , "Prop HWDP Inp Report": "AD" , "Prop Subs Inp Report": "AD" }
        },
        "BNT Tally": {
            "TUBE": {"Prop Drill Pipe Inp Report": "AF" , "Prop HWDP Inp Report": "AB" , "Prop Subs Inp Report": "AB"}
        },
        "SW Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "AG", "Prop HWDP Inp Report": "Z", "Prop Subs Inp Report": "Z"},
            "TUBE": {"Prop Drill Pipe Inp Report": "AG", "Prop HWDP Inp Report": "Z", "Prop Subs Inp Report": "Z"},
            "PIN": {"Prop Drill Pipe Inp Report": "AG", "Prop HWDP Inp Report": "Z", "Prop Subs Inp Report": "Z"}
        },
        "SP Tally": {
            "PIN": {"Prop Drill Pipe Inp Report": "AH", "Prop HWDP Inp Report": "AA", "Prop Subs Inp Report": "AA"},
            "BOX": {"Prop Drill Pipe Inp Report": "AH", "Prop HWDP Inp Report": "AA", "Prop Subs Inp Report": "AA"}
        },
        "SB Tally": {
            "PIN": {"Prop Drill Pipe Inp Report": "AI"},
            "BOX": {"Prop Drill Pipe Inp Report": "AI"}
        },
        "Other Damages Tally": {
            "BOX": {"Prop HWDP Inp Report": "AG" , "Prop Subs Inp Report": "AG" },
            "PIN": {"Prop HWDP Inp Report": "AG" , "Prop Subs Inp Report": "AG" },
            "TUBE": {"Prop HWDP Inp Report": "AG" , "Prop Subs Inp Report": "AG" }
        },
        "OTHER DBR Tally": {
            "BOX": {"Prop HWDP Inp Report": "AY" , "Prop Subs Inp Report": "AY"},
            "PIN": {"Prop HWDP Inp Report": "AY" , "Prop Subs Inp Report": "AY"},
            "TUBE": {"Prop HWDP Inp Report": "AY" , "Prop Subs Inp Report": "AY"},
        },
        "UG Tally": {
            "BOX": {"Prop HWDP Inp Report": "AV" , "Prop Subs Inp Report": "AV"},
            "PIN": {"Prop HWDP Inp Report": "AV" , "Prop Subs Inp Report": "AV"},
            "TUBE": {"Prop HWDP Inp Report": "AV" , "Prop Subs Inp Report": "AV"},
        },
        "DHB Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "AJ" , "Prop HWDP Inp Report": "AI" , "Prop Subs Inp Report": "AI" },
            "PIN": {"Prop Drill Pipe Inp Report": "AK" , "Prop HWDP Inp Report": "AJ" , "Prop Subs Inp Report": "AJ" },
            "TUBE": {"Prop HWDP Inp Report": "AH", "Prop Subs Inp Report": "AH" } #FOR 'CWP'
        },
        "DHBCWP Tally": {
            "BOX": {"Prop HWDP Inp Report": "AH" , "Prop Subs Inp Report": "AH" },
            "PIN": {"Prop HWDP Inp Report": "AH" , "Prop Subs Inp Report": "AH" }
        },
        "HBCP Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "AM" , "Prop HWDP Inp Report": "AL" , "Prop Subs Inp Report": "AL" },
            "PIN": {"Prop Drill Pipe Inp Report": "AN" , "Prop HWDP Inp Report": "AM" , "Prop Subs Inp Report": "AM" }
        },
        "HB Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "AL" , "Prop HWDP Inp Report": "AK" , "Prop Subs Inp Report": "AK" },
            "PIN": {"Prop Drill Pipe Inp Report": "AO" , "Prop HWDP Inp Report": "AN" , "Prop Subs Inp Report": "AN" }
        },
        "DBRHB Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "AP" , "Prop HWDP Inp Report": "AO" , "Prop Subs Inp Report": "AO" },
            "PIN": {"Prop Drill Pipe Inp Report": "AQ" , "Prop HWDP Inp Report": "AP" , "Prop Subs Inp Report": "AP" }
        },
        "MW Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "AR" , "Prop HWDP Inp Report": "AQ" , "Prop Subs Inp Report": "AQ" },
            "PIN": {"Prop Drill Pipe Inp Report": "AR" , "Prop HWDP Inp Report": "AQ" , "Prop Subs Inp Report": "AQ" },
            "TUBE": {"Prop Drill Pipe Inp Report": "AR" , "Prop HWDP Inp Report": "AQ" , "Prop Subs Inp Report": "AQ" }
        },
        "MT Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "AS" , "Prop HWDP Inp Report": "AR" , "Prop Subs Inp Report": "AR" },
            "PIN": {"Prop Drill Pipe Inp Report": "AT" , "Prop HWDP Inp Report": "AS" , "Prop Subs Inp Report": "AS" },
        },
        "MBD Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "AU", "Prop HWDP Inp Report": "AT", "Prop Subs Inp Report": "AT"},
            "PIN": {"Prop Drill Pipe Inp Report": "AV", "Prop HWDP Inp Report": "AU", "Prop Subs Inp Report": "AU"}
        },
        "CRACKED Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "AW", "Prop Subs Inp Report": "AX"},
            "PIN": {"Prop Drill Pipe Inp Report": "AW", "Prop Subs Inp Report": "AX"},
            "TUBE": {"Prop Drill Pipe Inp Report": "AW", "Prop Subs Inp Report": "AX"},
        },
        "DAM Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "AX" , "Prop HWDP Inp Report": "AW" , "Prop Subs Inp Report": "AW" },
            "PIN": {"Prop Drill Pipe Inp Report": "AX" , "Prop HWDP Inp Report": "AW" , "Prop Subs Inp Report": "AW" },
            "TUBE": {"Prop Drill Pipe Inp Report": "AX" , "Prop HWDP Inp Report": "AW" , "Prop Subs Inp Report": "AW" },
        },
        "CLASS2 Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "AY"},
            "PIN": {"Prop Drill Pipe Inp Report": "AY"},
            "TUBE": {"Prop Drill Pipe Inp Report": "AY"},
        },
        "EMI Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "AZ" , "Prop HWDP Inp Report": "AX"},
            "PIN": {"Prop Drill Pipe Inp Report": "AZ" , "Prop HWDP Inp Report": "AX"},
            "TUBE": {"Prop Drill Pipe Inp Report": "AZ" , "Prop HWDP Inp Report": "AX"},
        },
        "Not Recognized Keywords": {
            "BOX": {"Prop Drill Pipe Inp Report": "DA" , "Prop HWDP Inp Report": "DA" , "Prop Subs Inp Report": "DA" },
            "PIN": {"Prop Drill Pipe Inp Report": "DC" , "Prop HWDP Inp Report": "DC" , "Prop Subs Inp Report": "DC" },
            "TUBE": {"Prop Drill Pipe Inp Report": "DB" , "Prop HWDP Inp Report": "DB" , "Prop Subs Inp Report": "DB" }
        },
    }


    
    def process_tube_col(col_name, col_value, row_num, sheet):
        keywords = col_value.split()
        n = len(keywords)
        for i in range(n):
            keyword = keywords[i]
            if keyword == "MW":
                sheet[f"{col_chooser['MW Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, keyword)
            elif keyword == "DAM":
                sheet[f"{col_chooser['DAM Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, keyword)
            elif keyword == "EMI":
                sheet[f"{col_chooser['EMI Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, keyword)
            elif keyword == "BNT" or keyword == "BENT":
                sheet[f"{col_chooser['BNT Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, "BNT")
            elif keyword == "CRK":
                sheet[f"{col_chooser['CRACKED Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, keyword)
            elif keyword == "CL2":
                sheet[f"{col_chooser['CLASS2 Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, keyword)
            elif keyword == "UG":
                sheet[f"{col_chooser['UG Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, keyword)
            elif keyword == "OTHER":
                sheet[f"{col_chooser['OTHER DBR Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, keyword)
            elif keyword == "DHB":
                sheet[f"{col_chooser['DHB Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, keyword)
            elif not re.match(r'^(\d+(\.\d+)?|\d+|\d+\s\d+/\d+|\d+/\d+)$', keyword.replace('_', ' ')):
                if i + 1 < n:  
                    next_keyword = keywords[i + 1].replace('_', ' ')
                    if re.match(r'^(\d+(\.\d+)?|\d+|\d+\s\d+/\d+|\d+/\d+)$', next_keyword):
                        sheet[f"{col_chooser['Not Recognized Keywords'][col_name][reptyp]}{row_num}"] = f"{keyword} {next_keyword}"
                        continue  
                sheet[f"{col_chooser['Not Recognized Keywords'][col_name][reptyp]}{row_num}"] = keyword



            if len(keyword) == 3 and keyword.isdigit():
                sheet[f"{col_chooser['Critical Lengths'][col_name][reptyp]}{row_num}"] = keyword




    def process_box_pin_col(col_name, col_value, row_num, sheet):
        keywords = col_value.split()
        n = len(keywords)
        for i in range(n):
            keyword = keywords[i]
            if len(keyword) == 3 and keyword.isdigit():
                sheet[f"{col_chooser['Critical Lengths'][col_name][reptyp]}{row_num}"] = keyword
            elif keyword == "TS":
                process_ts_keyword(col_name, keywords, keyword, row_num, sheet)
            elif keyword == "TJ":
                process_tj_keyword(col_name, keywords, keyword, row_num, sheet)
            elif keyword == "BD":
                process_bd_keyword(col_name, keywords, keyword, row_num, sheet)
            elif keyword == "MT":
                sheet[f"{col_chooser['MT Tally'][col_name][reptyp]}{row_num}"] = "X"
                process_mtong_keyword(col_name, keywords, keyword, row_num, sheet)
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "MBD":
                sheet[f"{col_chooser['MBD Tally'][col_name][reptyp]}{row_num}"] = "X"
                process_mbd_keyword(col_name, keywords, keyword, row_num, sheet)
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "DS":
                sheet[f"{col_chooser['DS Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "DT":
                sheet[f"{col_chooser['DT Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "DTS":
                sheet[f"{col_chooser['DTS Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "PIT":
                sheet[f"{col_chooser['PIT Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "OR":
                sheet[f"{col_chooser['OR Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "DHB":
                sheet[f"{col_chooser['DHB Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "HB":
                sheet[f"{col_chooser['HB Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "DBRHB":
                sheet[f"{col_chooser['DBRHB Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "HBCP":
                sheet[f"{col_chooser['HBCP Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "R1":
                sheet[f"{col_chooser['R Tally'][col_name][reptyp]}{row_num}"] = "X"
                sheet[f"{col_chooser['R Type'][col_name][reptyp]}{row_num}"] = 1
                process_r_keyword(col_name, keywords, row_num, sheet)
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "R2":
                sheet[f"{col_chooser['R Tally'][col_name][reptyp]}{row_num}"] = "X"
                sheet[f"{col_chooser['R Type'][col_name][reptyp]}{row_num}"] = 2
                process_r_keyword(col_name, keywords, row_num, sheet)
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "R3":
                sheet[f"{col_chooser['R Tally'][col_name][reptyp]}{row_num}"] = "X"
                sheet[f"{col_chooser['R Type'][col_name][reptyp]}{row_num}"] = 3
                process_r_keyword(col_name, keywords, row_num, sheet)
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "R4":
                sheet[f"{col_chooser['R Tally'][col_name][reptyp]}{row_num}"] = "X"
                sheet[f"{col_chooser['R Type'][col_name][reptyp]}{row_num}"] = 4
                process_r_keyword(col_name, keywords, row_num, sheet)
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "SW":
                sheet[f"{col_chooser['SW Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, keyword)
            elif keyword == "SB":
                sheet[f"{col_chooser['SB Tally'][col_name][reptyp]}{row_num}"] = "X"
                process_short_keywords(col_name, keywords, keyword, row_num, sheet)
                update_keyword_tally(taldict, keyword)
            elif keyword == "SP":
                sheet[f"{col_chooser['SP Tally'][col_name][reptyp]}{row_num}"] = "X"
                process_short_keywords(col_name, keywords, keyword, row_num, sheet)
                update_keyword_tally(taldict, keyword)
            elif keyword == "ODAM":
                sheet[f"{col_chooser['Other Damages Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, keyword)
            elif not re.match(r'^(\d+(\.\d+)?|\d+|\d+\s\d+/\d+|\d+/\d+)$', keyword.replace('_', ' ')):
                if i + 1 < n:  
                    next_keyword = keywords[i + 1].replace('_', ' ')
                    if re.match(r'^(\d+(\.\d+)?|\d+|\d+\s\d+/\d+|\d+/\d+)$', next_keyword):
                        sheet[f"{col_chooser['Not Recognized Keywords'][col_name][reptyp]}{row_num}"] = f"{keyword} {next_keyword}"
                        continue  
                sheet[f"{col_chooser['Not Recognized Keywords'][col_name][reptyp]}{row_num}"] = keyword



    def process_ts_keyword(col_name, keywords, keyword, row_num, sheet):
        keywords = col_value.split()
        n = len(keywords)
        ts_index = keywords.index(keyword)
        if ts_index < n - 1:
            fraction_dec_pat = re.compile(r'((\d+_)?\d+/\d+|\d+(\.\d{3})?)')
            val_match = keywords[ts_index + 1]
            if fraction_dec_pat.match(val_match):
                sheet[f"{col_chooser['Tong Space'][col_name][reptyp]}{row_num}"] = val_match.replace('_', ' ')
    def process_tj_keyword(col_name, keywords, keyword, row_num, sheet):
        keywords = col_value.split()
        n = len(keywords)
        tj_index = keywords.index(keyword)
        if tj_index < n - 1:
            fraction_dec_pat = re.compile(r'((\d+_)?\d+/\d+|\d+(\.\d{3})?)')
            val_match = keywords[tj_index + 1]
            if fraction_dec_pat.match(val_match):
                sheet[f"{col_chooser['Tool Joint'][col_name][reptyp]}{row_num}"] = val_match.replace('_', ' ')
    def process_bd_keyword(col_name, keywords, keyword, row_num, sheet):
        keywords = col_value.split()
        n = len(keywords)
        bd_index = keywords.index(keyword)
        if bd_index < n - 1:
            fraction_dec_pat = re.compile(r'((\d+_)?\d+/\d+|\d+(\.\d{3})?)')
            val_match = keywords[bd_index + 1]
            if fraction_dec_pat.match(val_match):
                sheet[f"{col_chooser['Bevel Diameter'][col_name][reptyp]}{row_num}"] = val_match.replace('_', ' ')


    def process_short_keywords(col_name, keywords, keyword, row_num, sheet):
        sb_index = keywords.index(keyword) if keyword == "SB" else None
        sp_index = keywords.index(keyword) if keyword == "SP" else None 
        short_index = sb_index if sb_index is not None else sp_index
        if short_index < len(keywords) -1:
            fraction_dec_pat = re.compile(r'((\d+_)?\d+/\d+|\d+(\.\d{3})?)')
            val_match = keywords[short_index + 1]
            if fraction_dec_pat.match(val_match):
                if keyword == "SB":
                    sheet[f"{col_chooser['Tool Joint']['BOX'][reptyp]}{row_num}"] = val_match.replace('_', ' ') 
                elif keyword == "SP":
                    sheet[f"{col_chooser['Tool Joint']['PIN'][reptyp]}{row_num}"] = val_match.replace('_', ' ') 

    def process_long_keywords(col_name, keywords, keyword, row_num, sheet):
        lp_index = keywords.index(keyword) if keyword == "LP" else None  
        lb_index = keywords.index(keyword) if keyword == "LB" else None
        long_index = lp_index if lp_index is not None else lb_index
        if long_index < len(keywords) - 1:
            fraction_dec_pat = re.compile(r'((\d+_)?\d+/\d+|\d+(\.\d{3})?)')
            val_match = keywords[long_index + 1]
            if fraction_dec_pat.match(val_match):
                sheet[f"{col_chooser['Tool Joint'][col_name][reptyp]}{row_num}"] = val_match.replace('_', ' ') 


    def process_mtong_keyword(col_name, keywords, keyword, row_num, sheet):
        mt_index = keywords.index(keyword)
        if mt_index < len(keywords) - 1:
            fraction_dec_pat = re.compile(r'((\d+_)?\d+/\d+|\d+(\.\d{3})?)')
            val_match = keywords[mt_index + 1]
            if fraction_dec_pat.match(val_match):
                sheet[f"{col_chooser['Tong Space'][col_name][reptyp]}{row_num}"] = val_match.replace('_', ' ')
    def process_mbd_keyword(col_name, keywords, keyword, row_num, sheet):
        mbd_index = keywords.index(keyword)
        if mbd_index < len(keywords) - 1:
            fraction_dec_pat = re.compile(r'((\d+_)?\d+/\d+|\d+(\.\d{3})?)')
            val_match = keywords[mbd_index + 1]
            if fraction_dec_pat.match(val_match):
                sheet[f"{col_chooser['Bevel Diameter'][col_name][reptyp]}{row_num}"]  = val_match.replace('_', ' ')
    def process_r_keyword(col_name, keywords, row_num, sheet):
        for keyword in keywords:
            if keyword in ["R1", "R2", "R3", "R4"]:
                r_index = keywords.index(keyword)
        if r_index < len(keywords) - 1 and keywords[r_index + 1].isdigit() and len(keywords[r_index + 1]) == 6:
            sheet[f"{col_chooser['R Value'][col_name][reptyp]}{row_num}"] = keywords[r_index + 1]

    for col_name, col_value in joint_values.items():
        col_value = re.sub(r'(\d+)\s(\d+/\d+)', r'\1_\2', col_value)
        if col_name in ["BOX", "PIN"]:
            process_box_pin_col(col_name, col_value, row_num, sheet)
        elif col_name == "UT":
            sheet[f"{col_chooser['UT Value'][col_name][reptyp]}{row_num}"] = col_value
        elif col_name == "SERIAL":
            sheet[f"{col_chooser['Serial #'][col_name][reptyp]}{row_num}"]  = col_value
        elif col_name == "TUBE":
            process_tube_col(col_name, col_value, row_num, sheet)
        elif col_name == "BD-BOX":
            if col_value != "":
                sheet[f"{col_chooser['Bevel Diameter'][col_name][reptyp]}{row_num}"] = col_value.replace('_', ' ')
        elif col_name == "BD-PIN":
            if col_value != "":
                sheet[f"{col_chooser['Bevel Diameter'][col_name][reptyp]}{row_num}"] = col_value.replace('_', ' ')
        elif col_name == "ID":
            if col_value != "":
                sheet[f"{col_chooser['Tool Joint'][col_name][reptyp]}{row_num}"] = col_value.replace('_', ' ')
        elif col_name == "OD":
            if col_value != "":
                sheet[f"{col_chooser['Tool Joint'][col_name][reptyp]}{row_num}"] = col_value.replace('_', ' ')
        elif col_name == "TS-PIN":
            if col_value != "":
                sheet[f"{col_chooser['Tong Space'][col_name][reptyp]}{row_num}"] = col_value.replace('_', ' ')
        elif col_name == "TS-BOX":
            if col_value != "":
                sheet[f"{col_chooser['Tong Space'][col_name][reptyp]}{row_num}"] = col_value.replace('_', ' ')
        elif col_name == "DESCRIPTION":
            sheet[f"C{row_num}"] = col_value
        elif col_name == "SUBS CONN/DATA":
            sheet[f"{col_chooser['SUBS CONN/DATA'][col_name][reptyp]}{row_num}"] = col_value
        elif col_name == "COMMENTS":
            sheet[f"{col_chooser['COMMENTS'][col_name][reptyp]}{row_num}"] = col_value

def process_for_write_report_nd_pdpir(joint_values, row_num, sheet, datmg):
    reptyp = datmg.json_data_dict['report_user_metadata']['active_tab']
    taldict = datmg.keyword_tally_dict["Keyword Tallies"]
    joint_dict = datmg.keyword_tally_dict["Joint Tallies"]

    def update_keyword_tally(tally_dict, keyword):
        if keyword not in tally_dict:
            tally_dict[keyword] = 0
        tally_dict[keyword] += 1

    col_chooser = {
        "Tong Space": {
            "BOX": {"Prop Drill Pipe Inp Report": "L" , "Prop HWDP Inp Report": "J" , "Prop Subs Inp Report": "J" },
            "PIN": {"Prop Drill Pipe Inp Report": "M", "Prop HWDP Inp Report": "K", "Prop Subs Inp Report": "K" },
            "TS-BOX": {"Prop Drill Pipe Inp Report": "L" , "Prop HWDP Inp Report": "J" , "Prop Subs Inp Report": "J" },
            "TS-PIN": {"Prop Drill Pipe Inp Report": "M", "Prop HWDP Inp Report": "K", "Prop Subs Inp Report": "K" }
        },
        "Bevel Diameter": {
            "BOX": {"Prop Drill Pipe Inp Report": "J", "Prop HWDP Inp Report": "L", "Prop Subs Inp Report": "L"},
            "PIN": {"Prop Drill Pipe Inp Report": "K", "Prop HWDP Inp Report": "M", "Prop Subs Inp Report": "M"},
        },
        "Tool Joint": {
            "BOX": {"Prop Drill Pipe Inp Report": "H", "Prop HWDP Inp Report": "G", "Prop Subs Inp Report": "G" },
            "PIN": {"Prop Drill Pipe Inp Report": "I", "Prop HWDP Inp Report": "H", "Prop Subs Inp Report": "H"},
            "ID": {"Prop Drill Pipe Inp Report": "I", "Prop HWDP Inp Report": "H", "Prop Subs Inp Report": "H"},
            "OD": {"Prop Drill Pipe Inp Report": "H", "Prop HWDP Inp Report": "G", "Prop Subs Inp Report": "G" },
            "TUBE": {"Prop Drill Pipe Inp Report": "H", "Prop HWDP Inp Report": "G", "Prop Subs Inp Report": "G" }
        },
        "Serial #": {
            "SERIAL": {"Prop Drill Pipe Inp Report": "C" , "Prop HWDP Inp Report": "D" , "Prop Subs Inp Report": "D"},
        },
        "Critical Lengths": {
            "BOX": {"Prop Drill Pipe Inp Report": "BO", "Prop HWDP Inp Report": "BO", "Prop Subs Inp Report": "BO" },
            "PIN": {"Prop Drill Pipe Inp Report": "BR", "Prop HWDP Inp Report": "BR", "Prop Subs Inp Report": "BR" }
        },
        "R Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "R" , "Prop HWDP Inp Report": "R" , "Prop Subs Inp Report": "O" },
            "PIN": {"Prop Drill Pipe Inp Report": "S" , "Prop HWDP Inp Report": "S" , "Prop Subs Inp Report": "P" }
        },
        "R Value": {
            "BOX": {"Prop Drill Pipe Inp Report": "BN" , "Prop HWDP Inp Report": "BN" , "Prop Subs Inp Report": "BN" },
            "PIN": {"Prop Drill Pipe Inp Report": "BQ" , "Prop HWDP Inp Report": "BQ" , "Prop Subs Inp Report": "BQ" }
        },
        "UT Value": {
            "UT": {"Prop Drill Pipe Inp Report": "BT" , "Prop HWDP Inp Report": "BT" , "Prop Subs Inp Report": "BT" },
        },
        "DS Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "T" , "Prop HWDP Inp Report": "T" , "Prop Subs Inp Report": "Q" },
            "PIN": {"Prop Drill Pipe Inp Report": "U" , "Prop HWDP Inp Report": "U" , "Prop Subs Inp Report": "R" }
        },
        "DT Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "V" , "Prop HWDP Inp Report": "V" , "Prop Subs Inp Report": "S" },
            "PIN": {"Prop Drill Pipe Inp Report": "W" , "Prop HWDP Inp Report": "W" , "Prop Subs Inp Report": "T" }
        },
        "SB/SP Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "X" , "Prop HWDP Inp Report": "X" , "Prop Subs Inp Report": "U" },
            "PIN": {"Prop Drill Pipe Inp Report": "X" , "Prop HWDP Inp Report": "X" , "Prop Subs Inp Report": "U" }
        },
        "LB/LP Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "Y" , "Prop HWDP Inp Report": "Y" , "Prop Subs Inp Report": "V" },
            "PIN": {"Prop Drill Pipe Inp Report": "Y" , "Prop HWDP Inp Report": "Y" , "Prop Subs Inp Report": "V" }
        },
        "OR Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "Z" , "Prop HWDP Inp Report": "Z" , "Prop Subs Inp Report": "W" },
            "PIN": {"Prop Drill Pipe Inp Report": "AA" , "Prop HWDP Inp Report": "AA" , "Prop Subs Inp Report": "X" }
        },
        "BNT Tally": {
            "TUBE": {"Prop Drill Pipe Inp Report": "AB" , "Prop HWDP Inp Report": "AB" , "Prop Subs Inp Report": "Y"}
        },
        "Other Damages Tally": {
            "TUBE": {"Prop Drill Pipe Inp Report": "AC" , "Prop HWDP Inp Report": "AC" , "Prop Subs Inp Report": "AB" },
            "BOX": {"Prop Drill Pipe Inp Report": "AC" , "Prop HWDP Inp Report": "AC" , "Prop Subs Inp Report": "AC" },
            "PIN": {"Prop Drill Pipe Inp Report": "AD" , "Prop HWDP Inp Report": "AD" , "Prop Subs Inp Report": "AC" }
        },
        "BOX/PIN Other Damages Tally": {
            "TUBE": {"Prop Drill Pipe Inp Report": "AD" , "Prop HWDP Inp Report": "AD" , "Prop Subs Inp Report": "AC" },
            "BOX": {"Prop Drill Pipe Inp Report": "AE" , "Prop HWDP Inp Report": "AE" , "Prop Subs Inp Report": "Z" },
            "PIN": {"Prop Drill Pipe Inp Report": "AF" , "Prop HWDP Inp Report": "AF" , "Prop Subs Inp Report": "AA" }
        },
        "DHB Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "AG" , "Prop HWDP Inp Report": "AG" , "Prop Subs Inp Report": "AD" },
            "PIN": {"Prop Drill Pipe Inp Report": "AH" , "Prop HWDP Inp Report": "AH" , "Prop Subs Inp Report": "AE" }
        },
        "HB Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "AI" , "Prop HWDP Inp Report": "AI" , "Prop Subs Inp Report": "AF" },
            "PIN": {"Prop Drill Pipe Inp Report": "AL" , "Prop HWDP Inp Report": "AL" , "Prop Subs Inp Report": "AI" }
        },
        "HBCP Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "AJ" , "Prop HWDP Inp Report": "AJ" , "Prop Subs Inp Report": "AG" },
            "PIN": {"Prop Drill Pipe Inp Report": "AK" , "Prop HWDP Inp Report": "AK" , "Prop Subs Inp Report": "AH" }
        },
        "DBRHB Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "AM" , "Prop HWDP Inp Report": "AM" , "Prop Subs Inp Report": "AJ" },
            "PIN": {"Prop Drill Pipe Inp Report": "AN" , "Prop HWDP Inp Report": "AN" , "Prop Subs Inp Report": "AK" }
        },
        "MW Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "AO" , "Prop HWDP Inp Report": "AO" , "Prop Subs Inp Report": "AL" },
            "PIN": {"Prop Drill Pipe Inp Report": "AO" , "Prop HWDP Inp Report": "AO" , "Prop Subs Inp Report": "AL" },
            "TUBE": {"Prop Drill Pipe Inp Report": "AO" , "Prop HWDP Inp Report": "AO" , "Prop Subs Inp Report": "AL" }
        },
        "MT Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "AP" , "Prop HWDP Inp Report": "AP" , "Prop Subs Inp Report": "AM" },
            "PIN": {"Prop Drill Pipe Inp Report": "AQ" , "Prop HWDP Inp Report": "AQ" , "Prop Subs Inp Report": "AN" },
        },
        "MS Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "AR" , "Prop HWDP Inp Report": "AR" , "Prop Subs Inp Report": "AO" },
            "PIN": {"Prop Drill Pipe Inp Report": "AS" , "Prop HWDP Inp Report": "AS" , "Prop Subs Inp Report": "AP" },
        },
        "MOD Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "AT" , "Prop HWDP Inp Report": "AT" , "Prop Subs Inp Report": "AQ" },
            "PIN": {"Prop Drill Pipe Inp Report": "AT" , "Prop HWDP Inp Report": "AT" , "Prop Subs Inp Report": "AQ" },
            "TUBE": {"Prop Drill Pipe Inp Report": "AT" , "Prop HWDP Inp Report": "AT" , "Prop Subs Inp Report": "AQ" },
        },
        "DAM Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "AU" , "Prop HWDP Inp Report": "AU" , "Prop Subs Inp Report": "AR" },
            "PIN": {"Prop Drill Pipe Inp Report": "AU" , "Prop HWDP Inp Report": "AU" , "Prop Subs Inp Report": "AR" },
            "TUBE": {"Prop Drill Pipe Inp Report": "AU" , "Prop HWDP Inp Report": "AU" , "Prop Subs Inp Report": "AR" },
        },
        "EMI Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "AV" , "Prop HWDP Inp Report": "AV" , "Prop Subs Inp Report": "AS" },
            "PIN": {"Prop Drill Pipe Inp Report": "AV" , "Prop HWDP Inp Report": "AV" , "Prop Subs Inp Report": "AS" },
            "TUBE": {"Prop Drill Pipe Inp Report": "AV" , "Prop HWDP Inp Report": "AV" , "Prop Subs Inp Report": "AS" },
        },
        "OTHER Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "AW" , "Prop HWDP Inp Report": "AW" , "Prop Subs Inp Report": "AT" },
            "PIN": {"Prop Drill Pipe Inp Report": "AW" , "Prop HWDP Inp Report": "AW" , "Prop Subs Inp Report": "AT" },
            "TUBE": {"Prop Drill Pipe Inp Report": "AW" , "Prop HWDP Inp Report": "AW" , "Prop Subs Inp Report": "AT" },
        },
        "TR Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "AX" , "Prop HWDP Inp Report": "AX" , "Prop Subs Inp Report": "AU" },
            "PIN": {"Prop Drill Pipe Inp Report": "AY" , "Prop HWDP Inp Report": "AY" , "Prop Subs Inp Report": "AV" },
        },
        "BVR Tally": {
            "BOX": {"Prop Drill Pipe Inp Report": "AZ" , "Prop HWDP Inp Report": "AZ" , "Prop Subs Inp Report": "AW" },
            "PIN": {"Prop Drill Pipe Inp Report": "BA" , "Prop HWDP Inp Report": "BA" , "Prop Subs Inp Report": "AX" },
        },
        "Not Recognized Keywords": {
            "BOX": {"Prop Drill Pipe Inp Report": "BU" , "Prop HWDP Inp Report": "BU" , "Prop Subs Inp Report": "BU" },
            "PIN": {"Prop Drill Pipe Inp Report": "BW" , "Prop HWDP Inp Report": "BW" , "Prop Subs Inp Report": "BW" },
            "TUBE": {"Prop Drill Pipe Inp Report": "BV" , "Prop HWDP Inp Report": "BV" , "Prop Subs Inp Report": "BV" }
        }
    }

    def process_tube_col(col_name, col_value, row_num, sheet):
        keywords = col_value.split()
        n = len(keywords)
        for i in range(n):
            keyword = keywords[i]
            if keyword == "MW":
                sheet[f"{col_chooser['MW Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, keyword)
                update_keyword_tally(joint_dict, 'Scrap')
            elif keyword == "MOD":
                sheet[f"{col_chooser['MOD Tally'][col_name][reptyp]}{row_num}"] = "X"
                process_mod_keyword(col_name, keywords, keyword, row_num, sheet)
                update_keyword_tally(taldict, keyword)
                update_keyword_tally(joint_dict, 'Scrap')
            elif keyword == "DAM":
                sheet[f"{col_chooser['DAM Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, keyword)
                update_keyword_tally(joint_dict, 'Scrap')
            elif keyword == "EMI":
                sheet[f"{col_chooser['EMI Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, keyword)
                update_keyword_tally(joint_dict, 'Scrap')
            elif keyword == "OTHER":
                sheet[f"{col_chooser['OTHER Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, keyword)
                update_keyword_tally(joint_dict, 'Scrap')
            elif keyword == "BNT" or keyword == "BENT":
                sheet[f"{col_chooser['BNT Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, "BNT")
                update_keyword_tally(joint_dict, 'Repairable')
            elif not re.match(r'^(\d+(\.\d+)?|\d+|\d+\s\d+/\d+|\d+/\d+)$', keyword.replace('_', ' ')):
                if i + 1 < n:  
                    next_keyword = keywords[i + 1].replace('_', ' ')
                    if re.match(r'^(\d+(\.\d+)?|\d+|\d+\s\d+/\d+|\d+/\d+)$', next_keyword):
                        sheet[f"{col_chooser['Not Recognized Keywords'][col_name][reptyp]}{row_num}"] = f"{keyword} {next_keyword}"
                        continue  
                sheet[f"{col_chooser['Not Recognized Keywords'][col_name][reptyp]}{row_num}"] = keyword



    def process_box_pin_col(col_name, col_value, row_num, sheet):
        keywords = col_value.split()
        n = len(keywords)
        for i in range(n):
            keyword = keywords[i]
            if len(keyword) == 3 and keyword.isdigit():
                sheet[f"{col_chooser['Critical Lengths'][col_name][reptyp]}{row_num}"] = keyword
            elif keyword == "TJ":
                process_tj_keyword(col_name, keywords, keyword, row_num, sheet)
            elif keyword == "TS":
                process_ts_keyword(col_name, keywords, keyword, row_num, sheet)
            elif keyword == "BD":
                process_bd_keyword(col_name, keywords, keyword, row_num, sheet)
            elif keyword == "MT":
                sheet[f"{col_chooser['MT Tally'][col_name][reptyp]}{row_num}"] = "X"
                process_mtong_keyword(col_name, keywords, keyword, row_num, sheet)
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
                update_keyword_tally(joint_dict, 'Scrap')
            elif keyword == "MS":
                sheet[f"{col_chooser['MS Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
                update_keyword_tally(joint_dict, 'Scrap')
            elif keyword == "DS":
                sheet[f"{col_chooser['DS Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
                update_keyword_tally(joint_dict, 'Repairable')
            elif keyword == "DT":
                sheet[f"{col_chooser['DT Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
                update_keyword_tally(joint_dict, 'Repairable')
            elif keyword == "OR":
                sheet[f"{col_chooser['OR Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
                update_keyword_tally(joint_dict, 'Repairable')
            elif keyword == "DHB":
                sheet[f"{col_chooser['DHB Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "HB":
                sheet[f"{col_chooser['HB Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "DBRHB":
                sheet[f"{col_chooser['DBRHB Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "HBCP":
                sheet[f"{col_chooser['HBCP Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "MOD":
                sheet[f"{col_chooser['MOD Tally'][col_name][reptyp]}{row_num}"] = "X"
                process_mod_keyword(col_name, keywords, keyword, row_num, sheet)
                update_keyword_tally(taldict, keyword)
            elif keyword == "R":
                sheet[f"{col_chooser['R Tally'][col_name][reptyp]}{row_num}"] = "X"
                process_r_keyword(col_name, keywords, row_num, sheet)
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
            elif keyword == "SB" or keyword == "SP":
                sheet[f"{col_chooser['SB/SP Tally'][col_name][reptyp]}{row_num}"] = "X"
                process_short_keywords(col_name, keywords, keyword, row_num, sheet)
                update_keyword_tally(taldict, f"{keyword}")
                update_keyword_tally(joint_dict, 'Repairable')
            elif keyword == "LB" or keyword == "LP":
                sheet[f"{col_chooser['LB/LP Tally'][col_name][reptyp]}{row_num}"] = "X"
                process_long_keywords(col_name, keywords, keyword, row_num, sheet)
                update_keyword_tally(taldict, f"{keyword}")
                update_keyword_tally(joint_dict, 'Repairable')
            elif keyword == "ODAM":
                sheet[f"{col_chooser['Other Damages Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
                update_keyword_tally(joint_dict, 'Repairable')
            elif keyword == "OTH":
                sheet[f"{col_chooser['BOX/PIN Other Damages Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
                update_keyword_tally(joint_dict, 'Repairable')
            elif keyword == "TR":
                sheet[f"{col_chooser['TR Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
                update_keyword_tally(joint_dict, 'Repairable')
            elif keyword == "BVR":
                sheet[f"{col_chooser['BVR Tally'][col_name][reptyp]}{row_num}"] = "X"
                update_keyword_tally(taldict, f"{keyword}-{col_name}")
                update_keyword_tally(joint_dict, 'Repairable')
            elif not re.match(r'^(\d+(\.\d+)?|\d+|\d+\s\d+/\d+|\d+/\d+)$', keyword.replace('_', ' ')):
                if i + 1 < n:  
                    next_keyword = keywords[i + 1].replace('_', ' ')
                    if re.match(r'^(\d+(\.\d+)?|\d+|\d+\s\d+/\d+|\d+/\d+)$', next_keyword):
                        sheet[f"{col_chooser['Not Recognized Keywords'][col_name][reptyp]}{row_num}"] = f"{keyword} {next_keyword}"
                        continue  
                sheet[f"{col_chooser['Not Recognized Keywords'][col_name][reptyp]}{row_num}"] = keyword



    def process_ts_keyword(col_name, keywords, keyword, row_num, sheet):
        keywords = col_value.split()
        n = len(keywords)
        ts_index = keywords.index(keyword)
        if ts_index < n - 1:
            fraction_dec_pat = re.compile(r'((\d+_)?\d+/\d+|\d+(\.\d{3})?)')
            val_match = keywords[ts_index + 1]
            if fraction_dec_pat.match(val_match):
                sheet[f"{col_chooser['Tong Space'][col_name][reptyp]}{row_num}"] = val_match.replace('_', ' ')
    def process_tj_keyword(col_name, keywords, keyword, row_num, sheet):
        keywords = col_value.split()
        n = len(keywords)
        tj_index = keywords.index(keyword)
        if tj_index < n - 1:
            fraction_dec_pat = re.compile(r'((\d+_)?\d+/\d+|\d+(\.\d{3})?)')
            val_match = keywords[tj_index + 1]
            if fraction_dec_pat.match(val_match):
                sheet[f"{col_chooser['Tool Joint'][col_name][reptyp]}{row_num}"] = val_match.replace('_', ' ')
    def process_bd_keyword(col_name, keywords, keyword, row_num, sheet):
        keywords = col_value.split()
        n = len(keywords)
        bd_index = keywords.index(keyword)
        if bd_index < n - 1:
            fraction_dec_pat = re.compile(r'((\d+_)?\d+/\d+|\d+(\.\d{3})?)')
            val_match = keywords[bd_index + 1]
            if fraction_dec_pat.match(val_match):
                sheet[f"{col_chooser['Bevel Diameter'][col_name][reptyp]}{row_num}"] = val_match.replace('_', ' ')




    def process_short_keywords(col_name, keywords, keyword, row_num, sheet):
        sb_index = keywords.index(keyword) if keyword == "SB" else None
        sp_index = keywords.index(keyword) if keyword == "SP" else None 
        short_index = sb_index if sb_index is not None else sp_index
        if short_index < len(keywords) -1:
            fraction_dec_pat = re.compile(r'((\d+_)?\d+/\d+|\d+(\.\d{3})?)')
            val_match = keywords[short_index + 1]
            if fraction_dec_pat.match(val_match):
                sheet[f"{col_chooser['Tool Joint'][col_name][reptyp]}{row_num}"] = val_match.replace('_', ' ') 

    def process_long_keywords(col_name, keywords, keyword, row_num, sheet):
        lp_index = keywords.index(keyword) if keyword == "LP" else None  
        lb_index = keywords.index(keyword) if keyword == "LB" else None
        long_index = lp_index if lp_index is not None else lb_index
        if long_index < len(keywords) - 1:
            fraction_dec_pat = re.compile(r'((\d+_)?\d+/\d+|\d+(\.\d{3})?)')
            val_match = keywords[long_index + 1]
            if fraction_dec_pat.match(val_match):
                sheet[f"{col_chooser['Tool Joint'][col_name][reptyp]}{row_num}"] = val_match.replace('_', ' ') 


    def process_mtong_keyword(col_name, keywords, keyword, row_num, sheet):
        mt_index = keywords.index(keyword)
        if mt_index < len(keywords) - 1:
            fraction_dec_pat = re.compile(r'((\d+_)?\d+/\d+|\d+(\.\d{3})?)')
            val_match = keywords[mt_index + 1]
            if fraction_dec_pat.match(val_match):
                sheet[f"{col_chooser['Tong Space'][col_name][reptyp]}{row_num}"] = val_match.replace('_', ' ')


    def process_mod_keyword(col_name, keywords, keyword, row_num, sheet):
        mod_index = keywords.index(keyword)
        if mod_index < len(keywords) - 1:
            fraction_dec_pat = re.compile(r'((\d+_)?\d+/\d+|\d+(\.\d{3})?)')
            val_match = keywords[mod_index + 1]
            if fraction_dec_pat.match(val_match):
                sheet[f"{col_chooser['Tool Joint']['BOX'][reptyp]}{row_num}"]  = val_match.replace('_', ' ')

    def process_r_keyword(col_name, keywords, row_num, sheet):
        r_index = keywords.index("R")
        if r_index < len(keywords) - 1 and keywords[r_index + 1].isdigit() and len(keywords[r_index + 1]) == 6:
            sheet[f"{col_chooser['R Value'][col_name][reptyp]}{row_num}"] = keywords[r_index + 1]

    for col_name, col_value in joint_values.items():
        col_value = re.sub(r'(\d+)\s(\d+/\d+)', r'\1_\2', col_value)
        if col_name in ["BOX", "PIN"]:
            process_box_pin_col(col_name, col_value, row_num, sheet)
        elif col_name == "UT":
            sheet[f"{col_chooser['UT Value'][col_name][reptyp]}{row_num}"] = col_value
        elif col_name == "SERIAL":
            sheet[f"{col_chooser['Serial #'][col_name][reptyp]}{row_num}"]  = col_value
        elif col_name == "TUBE":
            process_tube_col(col_name, col_value, row_num, sheet)
        elif col_name == "ID":
            if col_value != "":
                sheet[f"{col_chooser['Tool Joint'][col_name][reptyp]}{row_num}"] = col_value.replace('_', ' ')
        elif col_name == "OD":
            if col_value != "":
                sheet[f"{col_chooser['Tool Joint'][col_name][reptyp]}{row_num}"] = col_value.replace('_', ' ')
        elif col_name == "TS-PIN":
            if col_value != "":
                sheet[f"{col_chooser['Tong Space'][col_name][reptyp]}{row_num}"] = col_value.replace('_', ' ')
        elif col_name == "TS-BOX":
            if col_value != "":
                sheet[f"{col_chooser['Tong Space'][col_name][reptyp]}{row_num}"] = col_value.replace('_', ' ')
        elif col_name == "DESCRIPTION":
            sheet[f"C{row_num}"] = col_value
        elif col_name == "SUBS CONN/DATA":
            sheet[f"N{row_num}"] = col_value
        elif col_name == "COMMENTS":
            sheet[f"BX{row_num}"] = col_value

def generate_summary_entry_tx_fd(tot_joints, keyword_tally_dict, datmg):
    joint_tallies = keyword_tally_dict.get("Joint Tallies", {})
    keyword_tallies = keyword_tally_dict.get("Keyword Tallies", {})

    total_damages = (
        int(keyword_tallies.get('DT-BOX', 0)) + 
        int(keyword_tallies.get('DT-PIN', 0)) + 
        int(keyword_tallies.get('DS-PIN', 0)) + 
        int(keyword_tallies.get('DS-BOX', 0)) + 
        int(keyword_tallies.get('DTS-BOX', 0)) + 
        int(keyword_tallies.get('DTS-PIN', 0)) + 
        int(keyword_tallies.get('PIT-BOX', 0)) + 
        int(keyword_tallies.get('PIT-PIN', 0)) + 

        int(keyword_tallies.get('ODAM1', 0)) +
        int(keyword_tallies.get('ODAM2', 0)) + 
        int(keyword_tallies.get('ODAM3', 0)) + 
        int(keyword_tallies.get('ODAM4', 0)) +  

        int(keyword_tallies.get('OR-PIN', 0)) +
        int(keyword_tallies.get('OR-BOX', 0)) +
        int(keyword_tallies.get('BNT', 0))
    )

    total_dbrs = (
        int(keyword_tallies.get('MW', 0)) +
        int(keyword_tallies.get('MT-BOX', 0)) +
        int(keyword_tallies.get('MT-PIN', 0)) +
        int(keyword_tallies.get('MS-BOX', 0)) +
        int(keyword_tallies.get('MS-PIN', 0)) +
        int(keyword_tallies.get('MOD', 0)) +
        int(keyword_tallies.get('DAM', 0)) +
        int(keyword_tallies.get('OTHER', 0)) +
        int(keyword_tallies.get('EMI', 0)) +
        int(keyword_tallies.get('DBRHB-BOX', 0)) +
        int(keyword_tallies.get('DBRHB-PIN', 0))
    )

    header_val_dict = {
    "TOTAL JOINTS: ": tot_joints,
    "TOTAL DAMAGES: ": total_damages,
    "Damaged Seals: ": f"{(int(keyword_tallies.get('DS-BOX', 0)) + int(keyword_tallies.get('DS-PIN', 0)))} [BOX: {int(keyword_tallies.get('DS-BOX', 0))}] [PIN: {int(keyword_tallies.get('DS-PIN', 0))}]",
    "Damaged Threads: ": f"{int(keyword_tallies.get('DT-BOX', 0)) + int(keyword_tallies.get('DT-PIN', 0))} [BOX: {int(keyword_tallies.get('DT-BOX', 0))}] [PIN: {int(keyword_tallies.get('DT-PIN', 0))}]",
    "Damaged Torque Shoulders: ": f"[BOX: {int(keyword_tallies.get('DTS-BOX', 0))}] [PIN: {int(keyword_tallies.get('DTS-PIN', 0))}]",
    "Pitted: ": f"[BOX: {int(keyword_tallies.get('PIT-BOX', 0))}] [PIN: {int(keyword_tallies.get('PIT-PIN', 0))}]",
    "Over-Refaced: ": f"[BOX: {int(keyword_tallies.get('OR-BOX', 0))}] [PIN: {int(keyword_tallies.get('OR-PIN', 0))}]",
    "Other Damages: ": f"[BOX: {int(keyword_tallies.get('ODAM-BOX', 0))}] [PIN: {int(keyword_tallies.get('ODAM-PIN', 0))}]",
    "Bent Tube: ": int(keyword_tallies.get('BNT', 0)),


    "TOTAL DBRs: ":  total_dbrs,
    "Minimum Walls: ": int(keyword_tallies.get('MW', 0)),
    "Minimum Tongs: ": f"[BOX: {int(keyword_tallies.get('MT-BOX', 0))}] [PIN: {int(keyword_tallies.get('MT-PIN', 0))}]",
    "Minimum Seals: ": f"[BOX: {int(keyword_tallies.get('MS-BOX', 0))}] [PIN: {int(keyword_tallies.get('MS-PIN', 0))}]",
    "Minimum ODs: ": int(keyword_tallies.get('MOD', 0)),
    "Damaged Tubes: ": int(keyword_tallies.get('DAM', 0)),
    "EMI Rejects: ": int(keyword_tallies.get('EMI', 0)),
    "DBR Hardband - Box: ": int(keyword_tallies.get('DBRHB-BOX', 0)),
    "DBR Hardband - Pin: ": int(keyword_tallies.get('DBRHB-PIN', 0)),
    "Other Damage - Tube: ": int(keyword_tallies.get('OTHER', 0)),

    "TOTAL REFACES: ": int(keyword_tallies.get('R1-BOX', 0)) + int(keyword_tallies.get('R2-BOX', 0)) + int(keyword_tallies.get('R3-BOX', 0)) + int(keyword_tallies.get('R4-BOX', 0)) + int(keyword_tallies.get('R1-PIN', 0)) + int(keyword_tallies.get('R2-PIN', 0)) + int(keyword_tallies.get('R3-PIN', 0)) + int(keyword_tallies.get('R4-PIN', 0)),
    "Box Refaces: ": f"[R1: {int(keyword_tallies.get('R1-BOX', 0))}] | [R2: {int(keyword_tallies.get('R2-BOX', 0))}] | [R3: {int(keyword_tallies.get('R3-BOX', 0))}] | [R4: {int(keyword_tallies.get('R4-BOX', 0))}]",
    "Pin Refaces: ": f"[R1: {int(keyword_tallies.get('R1-PIN', 0))}] | [R2: {int(keyword_tallies.get('R2-PIN', 0))}] | [R3: {int(keyword_tallies.get('R3-PIN', 0))}] | [R4: {int(keyword_tallies.get('R4-PIN', 0))}]",

    "TOTAL HARDBANDS: ": int(keyword_tallies.get('HB-BOX', 0)) + int(keyword_tallies.get('HB-PIN', 0)),
    "Hardband - Box: ": int(keyword_tallies.get('HB-BOX', 0)),
    "Hardband - Pin: ": int(keyword_tallies.get('HB-PIN', 0)),
    "Hardband Centerpad #1: ": int(keyword_tallies.get('HBCP-BOX', 0)),
    "Hardband Centerpad #2: ": int(keyword_tallies.get('HBCP-PIN', 0)),
    "Damaged Hardband - Box: ": int(keyword_tallies.get('DHB-BOX', 0)),
    "Damaged Hardband - Pin: ": int(keyword_tallies.get('DHB-PIN', 0)),

    }


    return header_val_dict

def generate_summary_entry_tx_cl2dbr(tot_joints, keyword_tally_dict, datmg):
    joint_tallies = keyword_tally_dict.get("Joint Tallies", {})
    keyword_tallies = keyword_tally_dict.get("Keyword Tallies", {})
    rep_style = datmg.json_data_dict['report_user_metadata']['report_style'] 
    
    total_damages = (
        int(keyword_tallies.get('DT-BOX', 0)) + 
        int(keyword_tallies.get('DT-PIN', 0)) + 
        int(keyword_tallies.get('DS-PIN', 0)) + 
        int(keyword_tallies.get('DS-BOX', 0)) + 
        int(keyword_tallies.get('DTS-BOX', 0)) + 
        int(keyword_tallies.get('DTS-PIN', 0)) + 
        int(keyword_tallies.get('PIT-BOX', 0)) + 
        int(keyword_tallies.get('PIT-PIN', 0)) + 
        int(keyword_tallies.get('SW', 0)) +
        int(keyword_tallies.get('SP', 0)) +
        int(keyword_tallies.get('SB', 0)) +
        int(keyword_tallies.get('OR-PIN', 0)) +
        int(keyword_tallies.get('OR-BOX', 0)) +
        int(keyword_tallies.get('BNT', 0))
    )

    total_dbrs = (
        int(keyword_tallies.get('MW', 0)) +
        int(keyword_tallies.get('MT-BOX', 0)) +
        int(keyword_tallies.get('MT-PIN', 0)) +
        int(keyword_tallies.get('MS-BOX', 0)) +
        int(keyword_tallies.get('MS-PIN', 0)) +
        int(keyword_tallies.get('MOD', 0)) +
        int(keyword_tallies.get('DAM', 0)) +
        int(keyword_tallies.get('OTHER', 0)) +
        int(keyword_tallies.get('EMI', 0)) +
        int(keyword_tallies.get('DBRHB-BOX', 0)) +
        int(keyword_tallies.get('DBRHB-PIN', 0))
    )

    if rep_style == "Class 2 DBR":
        total_dbrs += int(keyword_tallies.get('CL2', 0))
    elif rep_style == "Class 2 NOT DBR":
        total_damages += int(keyword_tallies.get('CL2', 0))

    header_val_dict_1 = {
        "TOTAL JOINTS: ": tot_joints,
        "TOTAL DAMAGES: ": total_damages
    }

    header_val_dict_2 = {
        "Damaged Seals: ": f"{(int(keyword_tallies.get('DS-BOX', 0)) + int(keyword_tallies.get('DS-PIN', 0)))} [BOX: {int(keyword_tallies.get('DS-BOX', 0))}] [PIN: {int(keyword_tallies.get('DS-PIN', 0))}]",
        "Damaged Threads: ": f"{int(keyword_tallies.get('DT-BOX', 0)) + int(keyword_tallies.get('DT-PIN', 0))} [BOX: {int(keyword_tallies.get('DT-BOX', 0))}] [PIN: {int(keyword_tallies.get('DT-PIN', 0))}]",
        "Damaged Torque Shoulders: ": f"[BOX: {int(keyword_tallies.get('DTS-BOX', 0))}] [PIN: {int(keyword_tallies.get('DTS-PIN', 0))}]",
        "Pitted: ": f"[BOX: {int(keyword_tallies.get('PIT-BOX', 0))}] [PIN: {int(keyword_tallies.get('PIT-PIN', 0))}]",
        "Over-Refaced: ": f"[BOX: {int(keyword_tallies.get('OR-BOX', 0))}] [PIN: {int(keyword_tallies.get('OR-PIN', 0))}]",
        "Other Damages: ": f"[BOX: {int(keyword_tallies.get('ODAM-BOX', 0))}] [PIN: {int(keyword_tallies.get('ODAM-PIN', 0))}]",
        "Swelled Box: ": int(keyword_tallies.get('SW', 0)),
        "Short Pin: ": int(keyword_tallies.get('SP', 0)),
        "Short Box: ": int(keyword_tallies.get('SB', 0)),
        "Bent Tube: ": int(keyword_tallies.get('BNT', 0))
    }

    header_val_dict_3 = {
        "TOTAL DBRs: ": total_dbrs
    }

    header_val_dict_cl2 = {
        "Class 2: ": int(keyword_tallies.get('CL2', 0))
    }

    header_val_dict_4 = {
        "Minimum Walls: ": int(keyword_tallies.get('MW', 0)),
        "Minimum Tongs: ": f"[BOX: {int(keyword_tallies.get('MT-BOX', 0))}] [PIN: {int(keyword_tallies.get('MT-PIN', 0))}]",
        "Minimum Seals: ": f"[BOX: {int(keyword_tallies.get('MS-BOX', 0))}] [PIN: {int(keyword_tallies.get('MS-PIN', 0))}]",
        "Min. Bevel Diameters: ": f"[BOX: {int(keyword_tallies.get('MBD-BOX', 0))}] [PIN: {int(keyword_tallies.get('MBD-PIN', 0))}]",
        "Cracked Tubes: ": int(keyword_tallies.get('CRK', 0)),
        "Damaged Tubes: ": int(keyword_tallies.get('DAM', 0)),
        "EMI Rejects: ": int(keyword_tallies.get('EMI', 0)),
        "DBR Hardband - Box: ": int(keyword_tallies.get('DBRHB-BOX', 0)),
        "DBR Hardband - Pin: ": int(keyword_tallies.get('DBRHB-PIN', 0)),
        "Other Damage - Tube: ": int(keyword_tallies.get('OTHER', 0)),

        "TOTAL REFACES: ": int(keyword_tallies.get('R1-BOX', 0)) + int(keyword_tallies.get('R2-BOX', 0)) + int(keyword_tallies.get('R3-BOX', 0)) + int(keyword_tallies.get('R4-BOX', 0)) + int(keyword_tallies.get('R1-PIN', 0)) + int(keyword_tallies.get('R2-PIN', 0)) + int(keyword_tallies.get('R3-PIN', 0)) + int(keyword_tallies.get('R4-PIN', 0)),
        "Box Refaces: ": f"[R1: {int(keyword_tallies.get('R1-BOX', 0))}] | [R2: {int(keyword_tallies.get('R2-BOX', 0))}] | [R3: {int(keyword_tallies.get('R3-BOX', 0))}] | [R4: {int(keyword_tallies.get('R4-BOX', 0))}]",
        "Pin Refaces: ": f"[R1: {int(keyword_tallies.get('R1-PIN', 0))}] | [R2: {int(keyword_tallies.get('R2-PIN', 0))}] | [R3: {int(keyword_tallies.get('R3-PIN', 0))}] | [R4: {int(keyword_tallies.get('R4-PIN', 0))}]",

        "TOTAL HARDBANDS: ": int(keyword_tallies.get('HB-BOX', 0)) + int(keyword_tallies.get('HB-PIN', 0)),
        "Hardband - Box: ": int(keyword_tallies.get('HB-BOX', 0)),
        "Hardband - Pin: ": int(keyword_tallies.get('HB-PIN', 0)),
        "Hardband Centerpad #1: ": int(keyword_tallies.get('HBCP-BOX', 0)),
        "Hardband Centerpad #2: ": int(keyword_tallies.get('HBCP-PIN', 0)),
        "Damaged Hardband - Box: ": int(keyword_tallies.get('DHB-BOX', 0)),
        "Damaged Hardband - Pin: ": int(keyword_tallies.get('DHB-PIN', 0)),
        "Damaged Hardband - CWP: ": int(keyword_tallies.get('DHB-TUBE', 0))
    }

    # Combine dictionaries based on the report style
    header_val_dict = {}
    
    if rep_style == "Class 2 DBR":
        header_val_dict.update(header_val_dict_1)
        header_val_dict.update(header_val_dict_2)
        header_val_dict.update(header_val_dict_3)
        header_val_dict.update(header_val_dict_cl2)
        header_val_dict.update(header_val_dict_4)
    elif rep_style == "Class 2 NOT DBR":
        header_val_dict.update(header_val_dict_1)
        header_val_dict.update(header_val_dict_cl2)
        header_val_dict.update(header_val_dict_2)
        header_val_dict.update(header_val_dict_3)
        header_val_dict.update(header_val_dict_4)

    return header_val_dict

def generate_summary_entry_nd_dp(tot_joints, keyword_tally_dict, datmg):
    joint_tallies = keyword_tally_dict.get("Joint Tallies", {})
    keyword_tallies = keyword_tally_dict.get("Keyword Tallies", {})

    total_damages = (
        int(keyword_tallies.get('DT-BOX', 0)) + 
        int(keyword_tallies.get('DT-PIN', 0)) + 
        int(keyword_tallies.get('DS-PIN', 0)) + 
        int(keyword_tallies.get('DS-BOX', 0)) + 
        int(keyword_tallies.get('SB', 0)) + 
        int(keyword_tallies.get('LB', 0)) + 
        int(keyword_tallies.get('SP', 0)) +
        int(keyword_tallies.get('LP', 0)) +
        int(keyword_tallies.get('OR-PIN', 0)) +
        int(keyword_tallies.get('OR-BOX', 0)) +
        int(keyword_tallies.get('BNT', 0)) +
        int(keyword_tallies.get('ODAM-BOX', 0)) +
        int(keyword_tallies.get('ODAM-PIN', 0)) +
        int(keyword_tallies.get('OTH-BOX', 0)) +
        int(keyword_tallies.get('OTH-PIN', 0))
    )

    total_dbrs = (
        int(keyword_tallies.get('MW', 0)) +
        int(keyword_tallies.get('MT-BOX', 0)) +
        int(keyword_tallies.get('MT-PIN', 0)) +
        int(keyword_tallies.get('MS-BOX', 0)) +
        int(keyword_tallies.get('MS-PIN', 0)) +
        int(keyword_tallies.get('MOD', 0)) +
        int(keyword_tallies.get('DAM', 0)) +
        int(keyword_tallies.get('OTHER', 0)) +
        int(keyword_tallies.get('EMI', 0)) +
        int(keyword_tallies.get('DBRHB-BOX', 0)) +
        int(keyword_tallies.get('DBRHB-PIN', 0))
    )

    header_val_dict = {
    "TOTAL JOINTS: ": tot_joints,
    
    "TOTAL DAMAGES: ": total_damages,
    "Damaged Seals: ": f"{(int(keyword_tallies.get('DS-BOX', 0)) + int(keyword_tallies.get('DS-PIN', 0)))} [BOX: {int(keyword_tallies.get('DS-BOX', 0))}] [PIN: {int(keyword_tallies.get('DS-PIN', 0))}]",
    "Damaged Threads: ": f"{int(keyword_tallies.get('DT-BOX', 0)) + int(keyword_tallies.get('DT-PIN', 0))} [BOX: {int(keyword_tallies.get('DT-BOX', 0))}] [PIN: {int(keyword_tallies.get('DT-PIN', 0))}]",
    "Short Box/Long Pin: ": f"[SB: {int(keyword_tallies.get('SB', 0))}] [LP: {int(keyword_tallies.get('LP', 0))}]",
    "Other Damages Box/Pin: ": f"[BOX: {int(keyword_tallies.get('OTH-BOX', 0))}] [PIN: {int(keyword_tallies.get('OTH-PIN', 0))}]",
    "Other Damages: ": f"[BOX: {int(keyword_tallies.get('ODAM-BOX', 0))}] [PIN: {int(keyword_tallies.get('ODAM-PIN', 0))}]",
    "Box Over-Refaces: ": int(keyword_tallies.get('OR-BOX', 0)),
    "Pin Over-Refaces: ": int(keyword_tallies.get('OR-PIN', 0)),
    "Bent Tube: ": int(keyword_tallies.get('BNT', 0)),

    "TOTAL DBRs: ":  total_dbrs,
    "Minimum Walls: ": int(keyword_tallies.get('MW', 0)),
    "Minimum Tongs: ": f"[BOX: {int(keyword_tallies.get('MT-BOX', 0))}] [PIN: {int(keyword_tallies.get('MT-PIN', 0))}]",
    "Minimum Seals: ": f"[BOX: {int(keyword_tallies.get('MS-BOX', 0))}] [PIN: {int(keyword_tallies.get('MS-PIN', 0))}]",
    "Minimum ODs: ": int(keyword_tallies.get('MOD', 0)),
    "Damaged Tubes: ": int(keyword_tallies.get('DAM', 0)),
    "EMI Rejects: ": int(keyword_tallies.get('EMI', 0)),
    "Other Damage - Tube: ": int(keyword_tallies.get('OTHER', 0)),
    "TOTAL REFACES: ": int(keyword_tallies.get('R-BOX', 0)) + int(keyword_tallies.get('R-PIN', 0)),
    "Box Refaces: ": int(keyword_tallies.get('R-BOX', 0)),
    "Pin Refaces: ": int(keyword_tallies.get('R-PIN', 0)),

    "TOTAL HARDBANDS: ": int(keyword_tallies.get('HB-BOX', 0)) + int(keyword_tallies.get('HB-PIN', 0)),
    "Hardband - Box: ": int(keyword_tallies.get('HB-BOX', 0)),
    "Hardband - Pin: ": int(keyword_tallies.get('HB-PIN', 0)),
    "Hardband Centerpad #1: ": int(keyword_tallies.get('HBCP-BOX', 0)),
    "Hardband Centerpad #2: ": int(keyword_tallies.get('HBCP-PIN', 0)),
    "Damaged Hardband - Box: ": int(keyword_tallies.get('DHB-BOX', 0)),
    "Damaged Hardband - Pin: ": int(keyword_tallies.get('DHB-PIN', 0)),
    "DBR Hardband - Box: ": int(keyword_tallies.get('DBRHB-BOX', 0)),
    "DBR Hardband - Pin: ": int(keyword_tallies.get('DBRHB-PIN', 0))
    }


    return header_val_dict


def write_summary_notes_nd_tubing(datmg, workbook):
    if 'Sum PH6 Tubing' in workbook.sheetnames:
        sheet = workbook['Sum PH6 Tubing']
        write_cell_caps = 'A45'
        write_cell_hbs = 'A46'
    elif 'Sum EUE Tubing' in workbook.sheetnames:
        sheet = workbook['Sum EUE Tubing']
        write_cell_caps = 'A44'
        write_cell_hbs = 'A41'
    else:
        return  # Exit the function if neither sheet is found

    # Define your search pattern with regular expression
    search_pattern_caps = "Missing Caps: () BOX ; () PIN"
    search_pattern_hbs = "Hardbands: () BOX ; () PIN"
    regex_pattern_caps = re.escape(search_pattern_caps).replace(r'\(\)', r'\(\d*\)')
    regex_pattern_hbs = re.escape(search_pattern_hbs).replace(r'\(\)', r'\(\d*\)')



    # Check if the string in notes_entry matches the regex pattern
    notes_entry = datmg.json_data_dict['report_user_metadata'].get('notes_entry', '')
    summary_entry = datmg.json_data_dict['report_user_metadata'].get('summary_entry', '')

    match_caps = re.search(regex_pattern_caps, notes_entry)
    match_hbs = re.search(regex_pattern_hbs, summary_entry)

    if match_caps:
        matched_text = match_caps.group(0)
        sheet[write_cell_caps] = matched_text
    if match_hbs:
        matched_text = match_hbs.group(0)
        sheet[write_cell_hbs] = matched_text

def generate_summary_entry_ndtube(keyword_tally_dict, datmg):
    connection_size = datmg.json_data_dict['report_user_metadata']['connection_size_selection']
    joint_tallies = keyword_tally_dict.get("Joint Tallies", {})
    keyword_tallies = keyword_tally_dict.get("Keyword Tallies", {})

    total_damages = int(keyword_tallies.get('DB', 0)) + int(keyword_tallies.get('DP', 0)) + int(keyword_tallies.get('BNT', 0))
    total_dbrs = int(keyword_tallies.get('BNTDBR', 0)) + int(keyword_tallies.get('SC', 0)) + int(keyword_tallies.get('TC', 0)) + int(keyword_tallies.get('GOU', 0)) + int(keyword_tallies.get('MW', 0)) + int(keyword_tallies.get('RW', 0)) + int(keyword_tallies.get('PIT', 0)) + int(keyword_tallies.get('MASH', 0)) + int(keyword_tallies.get('NODRIFT', 0)) + int(keyword_tallies.get('EMI', 0)) + int(keyword_tallies.get('OTHER', 0))
    total_joints = int(joint_tallies.get('Ready', 0)) + int(joint_tallies.get('Repairable', 0)) + int(joint_tallies.get('Scrap', 0))

    header_val_dict_1 = {
        "TOTAL JOINTS: ": total_joints,
    }
    header_val_dict_rjoints1 = {
        "Ready Joints: ": (joint_tallies.get('Ready', 0) - joint_tallies.get('Hardband Joints', 0))
    }
    header_val_dict_rjoints2 = {
        "Ready Joints: ": joint_tallies.get('Ready', 0)
    }
    header_val_dict_2 = {
        "Repairable Joints: ": joint_tallies.get('Repairable', 0),
        "Scrap Joints: ": joint_tallies.get('Scrap', 0),
    }
    header_val_dict_hbjoint = {
        "Hardband Joints: ": joint_tallies.get('Hardband Joints', 0)
    }

    header_val_dict = {
        "TOTAL DAMAGES: ": total_damages,
        "Damaged Box: ": keyword_tallies.get('DB', 0),
        "Damaged Pin: ": keyword_tallies.get('DP', 0),
        "Bent Tube: ": keyword_tallies.get('BNT', 0),
        "TOTAL DBRs: ": total_dbrs,
        "Bent DBR Tube: ": keyword_tallies.get('BNTDBR', 0),
        "Slip Cuts: ": keyword_tallies.get('SC', 0),
        "Slip Cut Repairs: ": keyword_tallies.get('SCR', 0),
        "Tong Cuts: ": keyword_tallies.get('TC', 0),
        "Gouges: ": keyword_tallies.get('GOU', 0),
        "Min Walls: ": keyword_tallies.get('MW', 0),
        "Rod Wears: ": keyword_tallies.get('RW', 0),
        "Pitted: ": keyword_tallies.get('PIT', 0),
        "Mashed Tube: ": keyword_tallies.get('MASH', 0),
        "No Drift: ": keyword_tallies.get('NODRIFT', 0),
        "EMI Reject: ": keyword_tallies.get('EMI', 0),
        "Other Damage - DBR: ": keyword_tallies.get('OTHER', 0),
        "Yellow Bands: ": keyword_tallies.get('YB', 0),
        "Blue Bands: ": keyword_tallies.get('BB', 0),
        "Hardbands: ": f"({keyword_tallies.get('HB-BOX', 0)}) BOX ; ({keyword_tallies.get('HB-PIN', 0)}) PIN"
    }

    summary_entry = f"Hardbands: ({keyword_tallies.get('HB-BOX', 0)}) BOX ; ({keyword_tallies.get('HB-PIN', 0)}) PIN"

    if connection_size.strip() == '2 3/8"':
        merged_dict = {**header_val_dict_1, **header_val_dict_rjoints1, **header_val_dict_2, **header_val_dict_hbjoint, **header_val_dict}
    else:
        merged_dict = {**header_val_dict_1, **header_val_dict_rjoints2, **header_val_dict_2, **header_val_dict}

    datmg.json_data_dict['report_user_metadata']['summary_entry'] = summary_entry
    return merged_dict






def process_for_write_report_nd_tubing(joint_values, row_num, sheet, datmg):
    taldict = datmg.keyword_tally_dict["Keyword Tallies"]
    joint_dict = datmg.keyword_tally_dict["Joint Tallies"]

    def update_keyword_tally(tally_dict, keyword):
        if keyword not in tally_dict:
            tally_dict[keyword] = 0
        tally_dict[keyword] += 1


    scrap_col = False  
    repair_col = False 
    ready_col = True
    hb_col = False

    #START HERE -- NEED TO IMPLEMENT LOGIC FOR THE FINAL CLASS KEYS AND YB, BB, and TALLIES
    def process_tube_ser_for_tally(col_name, col_value, row_num, sheet):
        keywords = col_value.split()
        for keyword in keywords:
            if len(keyword) == 4 and keyword.isdigit():
                sheet[f"AU{row_num}"] = keyword

    def process_box_pin_col(col_name, col_value, row_num, sheet):
        nonlocal scrap_col, repair_col, hb_col
        keywords = col_value.split()
        n = len(keywords)
        for i in range(n):
            keyword = keywords[i]
            if keyword == "DP":
                repair_col = True
                sheet[f"M{row_num}"] = "DP"
                update_keyword_tally(taldict, keyword)
            elif keyword == "DB":
                repair_col = True
                sheet[f"K{row_num}"] = "DB"
                update_keyword_tally(taldict, keyword)
            elif keyword == "HB":
                hb_col = True
                if col_name == "BOX":
                    sheet[f"J{row_num}"] = "HB"
                    update_keyword_tally(taldict, "HB-BOX")
                elif col_name == "PIN":
                    sheet[f"L{row_num}"] = "HB"
                    update_keyword_tally(taldict, "HB-PIN")
            elif not re.match(r'^(\d+(\.\d+)?|\d+|\d+\s\d+/\d+|\d+/\d+)$', keyword.replace('_', ' ')):
                if i + 1 < n:  
                    next_keyword = keywords[i + 1].replace('_', ' ')
                    if re.match(r'^(\d+(\.\d+)?|\d+|\d+\s\d+/\d+|\d+/\d+)$', next_keyword):
                        sheet[f"AV{row_num}"] = f"{keyword} {next_keyword}"
                        continue  
                sheet[f"AV{row_num}"] = keyword
            else:
                fraction_dec_pat = re.compile(r'((\d+_)?\d+/\d+|\d+(\.\d{3})?)')
                if col_name == "BOX" and fraction_dec_pat.match(keyword):
                    sheet[f"J{row_num}"] = keyword.replace('_', ' ') 
                elif col_name == "PIN" and fraction_dec_pat.match(keyword):
                    sheet[f"L{row_num}"] = keyword.replace('_', ' ') 

    def process_tube_col(col_name, col_value, row_num, sheet):
        nonlocal scrap_col, repair_col
        keywords = col_value.split()
        n = len(keywords)
        for i in range(n):
            keyword = keywords[i]
            if keyword == "BNT" or keyword == "BT":
                sheet[f"E{row_num}"] = "BENT"
                repair_col = True
                update_keyword_tally(taldict, "BNT")
            elif keyword == "DP":
                repair_col = True
                sheet[f"M{row_num}"] = "DP"
                update_keyword_tally(taldict, keyword)
            elif keyword == "DB":
                repair_col = True
                sheet[f"K{row_num}"] = "DB"
                update_keyword_tally(taldict, keyword)
            elif keyword == "BNTDBR" or keyword == "BTDBR":
                sheet[f"E{row_num}"] = "BENT DBR"
                scrap_col = True
                update_keyword_tally(taldict, "BNTDBR")
            elif keyword == "SC":
                scrap_col = True
                sheet[f"F{row_num}"] = "SC"
                process_sc_measure_keyword(col_name, keywords, keyword, row_num, sheet)
                update_keyword_tally(taldict, keyword)
            elif keyword == "SCR":
                sheet[f"F{row_num}"] = "SC-R"
                process_sc_measure_keyword(col_name, keywords, keyword, row_num, sheet)
                update_keyword_tally(taldict, keyword)
            elif keyword == "GOU":
                scrap_col = True
                sheet[f"F{row_num}"] = "GOU"
                process_sc_measure_keyword(col_name, keywords, keyword, row_num, sheet)
                update_keyword_tally(taldict, keyword)
            elif keyword == "TC": 
                scrap_col = True 
                sheet[f"F{row_num}"] = "TC"
                process_sc_measure_keyword(col_name, keywords, keyword, row_num, sheet)
                update_keyword_tally(taldict, keyword)
            elif keyword == "MW":
                scrap_col = True
                sheet[f"G{row_num}"] = "MW"
                update_keyword_tally(taldict, keyword)
            elif keyword == "RW":
                scrap_col = True
                sheet[f"G{row_num}"] = "RW"
                update_keyword_tally(taldict, keyword)   
            elif keyword == "PIT":
                scrap_col = True
                sheet[f"G{row_num}"] = "Pit"
                update_keyword_tally(taldict, keyword)
            elif keyword == "MASH":
                scrap_col = True
                sheet[f"G{row_num}"] = "Mashed"
                update_keyword_tally(taldict, keyword)
            elif keyword == "NODRIFT":
                scrap_col = True
                sheet[f"H{row_num}"] = "NO"
                update_keyword_tally(taldict, keyword)
            elif keyword == "EMI":
                scrap_col = True
                sheet[f"I{row_num}"] = "EMI"
                update_keyword_tally(taldict, keyword)
            elif keyword == "OTHER":
                scrap_col = True 
                sheet[f"G{row_num}"] = "OTHER"
                update_keyword_tally(taldict, keyword)
            elif keyword == "BB":
                sheet[f"Q{row_num}"] = "X"
                update_keyword_tally(taldict, keyword)
            elif keyword == "YB":
                sheet[f"P{row_num}"] = "X"
                update_keyword_tally(taldict, keyword)
            elif not re.match(r'^(\d+(\.\d+)?|\d+|\d+\s\d+/\d+|\d+/\d+)$', keyword.replace('_', ' ')):
                if i + 1 < n:  
                    next_keyword = keywords[i + 1].replace('_', ' ')
                    if re.match(r'^(\d+(\.\d+)?|\d+|\d+\s\d+/\d+|\d+/\d+)$', next_keyword):
                        sheet[f"AV{row_num}"] = f"{keyword} {next_keyword}"
                        continue  
                sheet[f"AV{row_num}"] = keyword

    def process_sc_measure_keyword(col_name, keywords, keyword, row_num, sheet):
        measure_index = keywords.index(keyword)
        if measure_index < len(keywords) - 1:
            fraction_dec_pat = re.compile(r'((\d+_)?\d+/\d+|\d+(\.\d{3})?)')
            val_match = keywords[measure_index + 1]
            if fraction_dec_pat.match(val_match):
                sheet[f"O{row_num}"] = val_match.replace('_', ' ')  

    def write_final_class_status(scrap_col, repair_col, ready_col, hb_col, row_num):
        # Check if the item is marked as scrap
        if scrap_col:
            sheet[f"N{row_num}"] = 'Scrap'
            update_keyword_tally(joint_dict, "Scrap")
        elif repair_col:
            sheet[f"N{row_num}"] = 'Repairable'
            update_keyword_tally(joint_dict, "Repairable")
        elif ready_col:
            update_keyword_tally(joint_dict, "Ready")
            if hb_col == True:
                update_keyword_tally(joint_dict, "Hardband Joints")

    for col_name, col_value in joint_values.items():
        col_value = re.sub(r'(\d+)\s(\d+/\d+)', r'\1_\2', col_value)
        if col_name in ["BOX", "PIN"]:
            process_box_pin_col(col_name, col_value, row_num, sheet)
        elif col_name == "UT":
            sheet[f"AT{row_num}"] = col_value
        elif col_name == "TUBE":
            process_tube_col(col_name, col_value, row_num, sheet)
            process_tube_ser_for_tally(col_name, col_value, row_num, sheet)
        elif col_name == "COMMENTS":
            sheet[f"S{row_num}"] = col_value
        elif col_name == "Visual OD":
            sheet[f"AX{row_num}"] = col_value.replace('_', ' ')

    write_final_class_status(scrap_col, repair_col, ready_col, hb_col, row_num)


def generate_pdf_copy(summary_data, filename, widmg, datmg, root):
    pdf = FPDF()
    pdf.add_page()

    # Add the logo image at the top of the page
    pdf.image('Pathfinder Logo.png', x=(215.9 * 0.4), y=(279.4 * 0.001), h=(215.9 * 0.1))

    metadata = datmg.json_data_dict['report_user_metadata']
    header_labels = {
        'operator_entry': 'Operator: ',
        'date_entry': 'Date: ',
        'contractor_entry': 'Contractor/Rig: ',
        'invoice_entry': 'Invoice: ',
        'location_entry': 'Location: ',
        'inspected_by_entry': 'Inspected By: ',
        'inspection_type_selection': 'Inspection Type: ',
        'connection_size_selection': 'Connection Size: ',
        'connection_type_selection': 'Connection Type: '
    }

    additional_1 = "Actual OD's, ID's, and Tong Space" if metadata.get('inspection_type_addodid', False) else ""
    additional_2 = f"{metadata.get('inspection_type_additional', '')}" if 'inspection_type_additional' in metadata else ""

    def print_meta_table():
        pdf.set_font("Arial", 'B', size=12)
        frame_x = pdf.l_margin
        frame_y = 23
        frame_width = pdf.w - pdf.l_margin - pdf.r_margin  # Adjusted frame width to account for both margins
        frame_height = pdf.h * 0.17

        # Draw the frame border
        pdf.set_xy(frame_x, frame_y)
        pdf.cell(frame_width, frame_height, border=1)

        label_positions = [
            ('operator_entry', 0.0, 0.0, 0.20),  # (key, x position header, y position, width)
            ('date_entry', 0.5, 0.0, 0.20),
            ('contractor_entry', 0.0, 0.16, 0.20),
            ('invoice_entry', 0.5, 0.16, 0.20),
            ('location_entry', 0.0, 0.32, 0.20),
            ('inspected_by_entry', 0.5, 0.32, 0.20),
            ('inspection_type_selection', 0.0, 0.48, 0.20),
            ('connection_size_selection', 0.0, 0.84, 0.20),
            ('connection_type_selection', 0.5, 0.84, 0.20)
        ]

        for key, relx, rely, relwidth in label_positions:
            header_text = header_labels[key]
            value_text = metadata.get(key, '')

            if key == 'inspection_type_selection':
                if additional_1:
                    value_text += f" {additional_1}"
                if additional_2:
                    value_text += f" {additional_2}"
                wraplength = frame_width * 0.425
                relheight = 0.36
                align = 'C'
            elif key == 'connection_size_selection':
                value_text = metadata.get(key, '') + " " + metadata.get('grade_info', '')
                wraplength = None
                relheight = 0.16
                align = 'L'
            else:
                wraplength = None
                relheight = 0.16
                align = 'L'

            # Header Label
            pdf.set_xy(frame_x + relx * frame_width, frame_y + rely * frame_height)
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(relwidth * frame_width, relheight * frame_height, header_text, border=1, ln=0, align='L')

            # Value Label
            pdf.set_xy(frame_x + (relx + relwidth) * frame_width, frame_y + rely * frame_height)
            pdf.set_font("Helvetica", 'I', 11)
            if wraplength:
                value_label_x = frame_x + (relx + relwidth) * frame_width
                value_label_y = frame_y + rely * frame_height + (relheight * frame_height / 2) - 5  # Adjust Y position to center vertically
                
                # Draw borders for the value cell
                pdf.set_xy(frame_x + (relx + relwidth) * frame_width, frame_y + rely * frame_height)
                pdf.cell((1 - relwidth) * frame_width, relheight * frame_height, '', border=1)

                # Adjust position for the value text inside the cell
                pdf.set_xy(value_label_x, value_label_y)
                pdf.multi_cell((1 - relwidth) * frame_width, relheight * frame_height / 3, value_text, border=0, align=align)
            else:
                pdf.cell((0.5 - relwidth) * frame_width, relheight * frame_height, value_text, border=1, ln=0, align='L')


    def print_summary_section(summary_data):
        pdf.ln(10)  # Add space before the SUMMARY section
        pdf.set_font("Arial", 'B', size=12)

        # Define the width for the cells
        frame_width = pdf.w - pdf.l_margin - pdf.r_margin
        header_height = 10

        # Draw the header cell
        pdf.set_xy(pdf.l_margin, pdf.get_y())
        pdf.set_font("Arial", 'BI', size=22)
        pdf.cell(frame_width, header_height, "S U M M A R Y", border=1, ln=1, align='L')

        bold_ul_items = ["TOTAL DAMAGES: ", "TOTAL DBRs: ", "TOTAL JOINTS: ", "TOTAL HARDBANDS: ", "TOTAL REFACES: "]

        # Iterate over the summary data (header-value pairs)
        for header, value in summary_data.items():
            # Bold and underlined items (e.g., TOTAL DAMAGES, TOTAL DBRs)
            if header in bold_ul_items:
                pdf.ln(2)
                pdf.set_font("Helvetica", 'B', size=16)

                # Calculate the width of the header
                header_width = pdf.get_string_width(header)

                # Reserve space for the value
                pdf.set_font("Arial", 'I', size=16)
                value_width = pdf.get_string_width(str(value))

                # Calculate the remaining space for the dots
                remaining_width = frame_width - header_width - value_width - 10  # Adjust 10 as padding for space

                # Draw the header
                pdf.set_font("Helvetica", 'BU', size=18)
                pdf.cell(header_width, 5, header, border=0, ln=0)

                # Draw the dots dynamically
                dot_string = '.' * int(remaining_width / pdf.get_string_width('.'))
                pdf.set_font("Arial", 'I', size=18)
                pdf.cell(remaining_width, 5, dot_string, border=0, ln=0)

                # Draw the value at the right
                pdf.set_font("Arial", 'B', size=20)
                pdf.cell(value_width, 5, str(value), border=0, ln=1)

                pdf.ln(1)

            # Non-bolded items (sub-categories)

            else:
                if value in [0, "0 [BOX: 0] [PIN: 0]", "[BOX: 0] [PIN: 0]", "[SB: 0] [LP: 0]", "(0) BOX ; (0) PIN"]:
                    continue
                # Calculate the width of the header
                header_width = pdf.get_string_width(header)

                # Reserve space for the value
                pdf.set_font("Arial", 'I', size=15)
                value_width = pdf.get_string_width(str(value))

                # Shorter dot sequence for sub-items (sub-categories)
                remaining_width = frame_width - header_width - value_width - 20  # Adjust for sub-category

                # Draw the header with indentation
                indent = 10  # You can adjust the indentation as needed
                pdf.set_x(pdf.get_x() + indent)

                # Draw the header
                pdf.set_font("Helvetica", 'I', size=15)
                pdf.cell(header_width, 4.5, header, border=0, ln=0)

                # Draw the shorter dots dynamically
                dot_string = '.' * int(remaining_width / pdf.get_string_width('.'))
                pdf.set_font("Arial", 'I', size=15)
                pdf.cell(remaining_width, 4.5, dot_string, border=0, ln=0)

                # Draw the value at the right
                pdf.set_font("Arial", 'I', size=15)
                pdf.cell(value_width, 4.5, str(value), border=0, ln=1)

                # Add space after specific sub-items
                if header in ["Bent Tube: ", "Other Damage - DBR: ", "Other Damage - Tube: ", "Scrap Joints: ", "Hardband Joints: ", "Pin Refaces: "]:
                    if "Scrap Joints: " in summary_data and "Hardband Joints: " in summary_data and header == "Hardband Joints: ":
                        pdf.ln(4)
                    elif "Scrap Joints: " in summary_data and "Hardband Joints: " not in summary_data and header == "Scrap Joints: ":
                        pdf.ln(4)
                    elif header not in ["Scrap Joints: ", "Hardband Joints: "]:
                        pdf.ln(4)


        pdf.ln(1)  # Add space after the SUMMARY section



    def print_notes_section():
        notes = metadata.get('notes_entry', '')
        pdf.ln(1)  # Add space before the NOTES section
        pdf.set_font("Arial", 'B', size=12)
        
        # Define the width and header height for the cells
        frame_width = pdf.w - pdf.l_margin - pdf.r_margin
        header_height = 10

        # Calculate the available height left on the page (from the current Y position to the bottom margin)
        available_height = pdf.h - pdf.get_y() - pdf.b_margin

        # Draw the header cell for the "NOTES" section
        pdf.set_xy(pdf.l_margin, pdf.get_y())
        pdf.set_font("Arial", 'BI', size=22)  # Set font for the "N O T E S" header
        pdf.cell(frame_width, header_height, "N O T E S", border=1, ln=1, align='L')

        # Reset the font size for the actual notes content
        pdf.set_font("Arial", size=12)  # Set font size to 12 for the notes content

        extra_lines = "\n" * 3
        notes_with_extra_space = notes + extra_lines

        # Calculate the remaining height for the "NOTES" text area
        notes_height = available_height - header_height

        # Estimate how many lines can fit in the remaining height on the first page
        line_height = 3.5  # Approximate line height for multi-cell text
        max_lines_on_first_page = int(notes_height / line_height)

        # Split the notes if they exceed the space available on the first page
        notes_lines = notes_with_extra_space.split("\n")
        
        if len(notes_lines) <= max_lines_on_first_page:
            # Draw all notes if they fit within the remaining height
            pdf.multi_cell(frame_width, line_height, notes_with_extra_space, border=1)
        else:
            # Fit as many lines as possible on the first page
            notes_to_fit_first_page = "\n".join(notes_lines[:max_lines_on_first_page])
            pdf.multi_cell(frame_width, line_height, notes_to_fit_first_page, border=1)

            # Move to the second page for the remaining notes
            pdf.add_page()
            
            # Draw the "Notes Continued" header
            pdf.set_xy(pdf.l_margin, pdf.get_y())
            pdf.set_font("Arial", 'BI', size=22)  # Font for the "Notes Continued" header
            pdf.cell(frame_width, header_height, "N O T E S   C O N T I N U E D", border=1, ln=1, align='L')

            # Reset font size again for the remaining notes text
            pdf.set_font("Arial", size=12)

            # Draw a bordered frame for the entire second page (within margins)
            pdf.set_xy(pdf.l_margin, pdf.get_y())  # Start at the top left corner inside margins
            available_height_for_notes = pdf.h - pdf.get_y() - pdf.b_margin  # Full height minus margins and header height
            pdf.cell(frame_width, available_height_for_notes, '', border=1)  # Draw a full-page frame for the notes section

            # Now write the remaining notes inside this full-page frame
            pdf.set_xy(pdf.l_margin, pdf.get_y() + 1)  # Adjust y position to start writing text inside the frame
            remaining_notes = "\n".join(notes_lines[max_lines_on_first_page:])
            pdf.multi_cell(frame_width, line_height, remaining_notes, border=0)  # No border for the text, since the frame is drawn




    print_meta_table()
    print_summary_section(summary_data)
    print_notes_section()
    pdf.add_page()


    # Prepare the table
    active_tab = datmg.json_data_dict['active_tab']
    columns = datmg.json_data_dict['report_data'][active_tab]['users_column_select']
    report_data = datmg.json_data_dict['report_data'][active_tab]['joint_data']
    table_columns = ["JOINT #"] + columns

    # Set default margins
    pdf.set_left_margin(10)
    pdf.set_right_margin(10)
    table_width = pdf.w - 2 * pdf.l_margin  # Available width for table
    col_width = table_width / len(table_columns)  # Calculate uniform column width

    def print_table_headers():
        pdf.set_font("Arial", 'B', size=12)  # Start with a default size
        pdf.set_fill_color(200, 200, 200)
        pdf.set_text_color(0)
        
        for col in table_columns:
            # Calculate the width of the text with the initial font size
            text_width = pdf.get_string_width(col)
            
            # Initial font size
            font_size = 12
            pdf.set_font("Arial", 'B', font_size)
            
            # Reduce font size if the text width is greater than the cell width
            while text_width > col_width - 5.2:  # subtract a small margin to ensure text fits comfortably
                font_size -= 1
                pdf.set_font("Arial", 'B', font_size)
                text_width = pdf.get_string_width(col)
            
            # Ensure font size does not reduce too much, set a minimum font size if necessary
            if font_size < 8:
                font_size = 8
                pdf.set_font("Arial", 'B', font_size)
            
            # Draw the cell with the text
            pdf.cell(col_width, 10, col, 1, 0, 'C', fill=True)
        pdf.ln()


    # Print table headers
    print_table_headers()

    def print_table_row(joint, data):
        pdf.set_font("Arial", size=12)
        joint_number = joint.replace("Joint_", "")  # Remove the prefix
        pdf.cell(col_width, 10, joint_number, 1, 0, 'C')

        for col in columns:
            value = data.get(col, '')

            # Ensure text width is accurately calculated and reduce font size appropriately
            font_size = 12
            pdf.set_font("Arial", size=font_size)  # Ensure font size is set before width calculation
            text_width = pdf.get_string_width(value)

            # Check if the text fits within the cell width, including a margin
            while text_width > col_width - 1 and font_size > 8:  # Reduced margin to 1
                font_size -= 1
                pdf.set_font("Arial", size=font_size)
                text_width = pdf.get_string_width(value)

            # Wrap the text if it still doesn't fit after reducing font size
            if text_width > col_width - 1:
                lines = []
                words = value.split(' ')
                current_line = ""
                for word in words:
                    if pdf.get_string_width(current_line + " " + word) <= col_width:
                        current_line += " " + word
                    else:
                        lines.append(current_line.strip())
                        current_line = word
                lines.append(current_line.strip())

                # Check if the wrapped text fits within the cell height
                max_lines = 10 // (pdf.font_size * 2.5 / 3)  # Max number of lines that fit in the cell
                if len(lines) > max_lines:
                    while len(lines) > max_lines and font_size > 5:
                        font_size -= 1
                        pdf.set_font("Arial", size=font_size)
                        lines = []
                        current_line = ""
                        for word in words:
                            if pdf.get_string_width(current_line + " " + word) <= col_width:
                                current_line += " " + word
                            else:
                                lines.append(current_line.strip())
                                current_line = word
                        lines.append(current_line.strip())

                # Print the wrapped lines within the cell
                cell_start_y = pdf.get_y()
                cell_start_x = pdf.get_x()
                line_height = 10 / max(len(lines), 1)
                for i, line in enumerate(lines):
                    y_position = cell_start_y + (i * line_height)
                    pdf.set_xy(cell_start_x, y_position)
                    pdf.cell(col_width, line_height, line, 0, 0, 'C')
                pdf.set_xy(cell_start_x, cell_start_y)
                pdf.cell(col_width, 10, "", 1, 0, 'C')  # Draw the cell border

            else:
                pdf.cell(col_width, 10, value, 1, 0, 'C')

    # Print table rows
    for joint, data in report_data.items():
        # Add a new page and headers if space is running out
        if pdf.get_y() + 2 * (1 + len(columns)) > pdf.page_break_trigger:
            pdf.add_page()
            print_table_headers()
        
        print_table_row(joint, data)
        pdf.ln()

    # Save the generated PDF
    pdf.output(filename)

def main():
    try:
        root = initialize_main_window()
        datmg = DataManager()
        widmg = WidgetManager(root)
        datmg.set_widget_manager(widmg)
        widmg.set_data_manager(datmg)
        branch_select_screen(widmg, datmg, root)
        root.bind("<Configure>", widmg.schedule_resize_fonts)
        root.mainloop()
    except Exception as e:
        logging.error("An error occurred", exc_info=True)
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        messagebox.showerror("Error", "An unexpected error occurred. Please check the log file.")
        root.destroy()

if __name__ == "__main__":
    main()

